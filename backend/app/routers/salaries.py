from typing import List, Optional
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, delete, update, func
from app.database import get_db
from app.models.salary import MonthlySalary, MonthlyWorkdayConfig, AdvancePayment, AdvanceLoan
from app.models.employee import Employee
from app.models.user import AppUser, UserRole
from app.middleware.auth import require_roles, get_current_user
from pydantic import BaseModel
import openpyxl
from io import BytesIO
from app.utils.audit_helper import log_audit
from app.utils.lock_helper import check_month_locked

router = APIRouter(prefix="/salaries", tags=["Salaries - Lương"])


class BaseSalaryRow(BaseModel):
    employee_id: int
    employee_code: str
    full_name: str
    department: Optional[str] = None
    base_salary: float
    allowance: float

class BaseSalaryResponse(BaseModel):
    month_key: str
    standard_days: float
    is_locked: bool
    rows: List[BaseSalaryRow]


class SalaryHistoryRow(BaseModel):
    month_key: str
    base_salary: float
    allowance: float
    base_daily_wage: float
    pay_method: Optional[str] = None
    salary_coefficient: float
    updated_at: Optional[datetime] = None


@router.get("/base", response_model=BaseSalaryResponse)
async def get_base_salaries(
    month_key: str = Query(..., description="YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Lấy danh sách lương cơ bản và phụ cấp của tất cả nhân viên trong tháng"""
    # Lấy config tháng
    config_result = await db.execute(select(MonthlyWorkdayConfig).where(MonthlyWorkdayConfig.month_key == month_key))
    config = config_result.scalar_one_or_none()
    standard_days = float(config.company_work_days) if config else 26.0
    is_locked = bool(config.is_locked) if config else False

    # Lấy lương
    query = select(MonthlySalary, Employee).join(Employee, MonthlySalary.employee_id == Employee.id)\
        .where(MonthlySalary.month_key == month_key)
    if current_user.role == UserRole.WORKER:
        query = query.where(MonthlySalary.employee_id == current_user.employee_id)
    query = query.order_by(Employee.employee_code)
    
    result = await db.execute(query)
    records = result.all()

    rows = []
    for sal, emp in records:
        rows.append(BaseSalaryRow(
            employee_id=sal.employee_id,
            employee_code=emp.employee_code,
            full_name=emp.full_name,
            department=emp.department,
            base_salary=float(sal.base_salary or 0),
            allowance=float(sal.allowance or 0),
        ))

    return BaseSalaryResponse(
        month_key=month_key,
        standard_days=standard_days,
        is_locked=is_locked,
        rows=rows,
    )


@router.get("/history", response_model=List[SalaryHistoryRow])
async def get_salary_history(
    employee_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Lịch sử lương theo tháng của 1 nhân viên"""
    if current_user.role == UserRole.WORKER and employee_id != current_user.employee_id:
        raise HTTPException(status_code=403, detail="Không có quyền xem lịch sử lương của nhân viên khác")

    result = await db.execute(
        select(MonthlySalary)
        .where(MonthlySalary.employee_id == employee_id)
        .order_by(MonthlySalary.month_key.desc())
    )
    rows = []
    for sal in result.scalars().all():
        rows.append(SalaryHistoryRow(
            month_key=sal.month_key,
            base_salary=float(sal.base_salary or 0),
            allowance=float(sal.allowance or 0),
            base_daily_wage=float(sal.base_daily_wage or 0),
            pay_method=sal.pay_method,
            salary_coefficient=float(sal.salary_coefficient or 1),
            updated_at=sal.updated_at,
        ))
    return rows


@router.get("/export-template")
async def export_salaries_template(
    month_key: str = Query(..., description="YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Xuất file Excel mẫu lương cơ bản cố định chứa danh sách nhân viên hiện tại để điền và import lại."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    # 1. Query month workday config to get standard days
    config_result = await db.execute(
        select(MonthlyWorkdayConfig).where(MonthlyWorkdayConfig.month_key == month_key)
    )
    config = config_result.scalar_one_or_none()
    standard_days = float(config.company_work_days) if config else 26.0

    # 2. Query active employees
    emp_q = select(Employee).where(Employee.is_active == True).order_by(Employee.employee_code)
    emp_res = await db.execute(emp_q)
    employees = emp_res.scalars().all()
    
    # 3. Query existing monthly salaries for this month to prefill if exists
    salary_q = select(MonthlySalary).where(MonthlySalary.month_key == month_key)
    salary_res = await db.execute(salary_q)
    sal_map = {s.employee_id: s for s in salary_res.scalars().all()}

    # Initialize workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bang Luong"
    ws.views.sheetView[0].showGridLines = True

    # Styling
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    font_bold = Font(name='Arial', size=10, bold=True)
    font_regular = Font(name='Arial', size=10)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Row 1: Title
    ws['A1'] = f"BẢNG LƯƠNG CỐ ĐỊNH - THÁNG {month_key}"
    ws['A1'].font = Font(name='Arial', size=12, bold=True)

    # Row 2: Workdays coefficient configuration
    ws['F2'] = "He so luong"
    ws['F2'].font = font_bold
    ws['G2'] = standard_days
    ws['G2'].font = font_regular
    ws['G2'].number_format = '0.0'

    # Row 3: Headers
    headers = ["STT", "Mã NV", "Họ Tên", "Lương cơ bản", "Phụ cấp"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = thin_border

    # Row 4 onwards: Data
    stt = 1
    current_row = 4
    for emp in employees:
        sal = sal_map.get(emp.id)
        # Fallback to employee profile salary if monthly salary not present
        base_val = float(sal.base_salary) if (sal and sal.base_salary is not None) else float(emp.base_salary or 0.0)
        allowance_val = float(sal.allowance) if (sal and sal.allowance is not None) else 0.0

        row_data = [
            stt,
            emp.employee_code,
            emp.full_name,
            base_val,
            allowance_val
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if col_idx in (1, 2):
                cell.alignment = align_center
            elif col_idx == 3:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = '#,##0'
                
        current_row += 1
        stt += 1

    # Column widths
    column_widths = {1: 8, 2: 12, 3: 25, 4: 18, 5: 18, 6: 12, 7: 12}
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Mau_nhap_luong_{month_key}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import-base")
async def import_base_salaries(
    month_key: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN)),
):
    """Import file lương cố định (Lương cơ bản + Phụ cấp) cho toàn bộ nhân viên. Chỉ admin."""
    # Kiểm tra khóa tháng
    config_result = await db.execute(select(MonthlyWorkdayConfig).where(MonthlyWorkdayConfig.month_key == month_key))
    config = config_result.scalar_one_or_none()
    
    if config and config.is_locked:
        raise HTTPException(400, "Tháng này đã bị khóa (chốt dữ liệu), không thể import lại lương.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "File không hợp lệ. Vui lòng upload file Excel (.xlsx)")
    
    ws = None
    if "Bang Luong" in wb.sheetnames:
        ws = wb["Bang Luong"]
    else:
        ws = wb[wb.sheetnames[0]]

    # Tìm hệ số lương
    standard_days = 26.0
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "he so luong" in v.lower().replace("ệ", "e").replace("ố", "o").replace("ươ", "uo"):
                val = ws.cell(r, c + 1).value
                if val is not None:
                    try:
                        standard_days = float(val)
                    except ValueError:
                        pass
                break

    # Lưu cấu hình tháng
    if not config:
        config = MonthlyWorkdayConfig(month_key=month_key, company_work_days=standard_days)
        db.add(config)
    else:
        config.company_work_days = standard_days
    # 1. Thu thap thong tin tu file Excel
    file_emps = {} # code -> {name, base_salary, allowance}
    for r in range(3, ws.max_row + 1):
        c_raw = ws.cell(r, 2).value
        n_raw = ws.cell(r, 3).value
        if c_raw and n_raw:
            code = str(int(c_raw) if isinstance(c_raw, float) else c_raw).strip().lstrip("'")
            file_emps[code] = {
                "name": str(n_raw).strip(),
                "base_salary": float(ws.cell(r, 4).value or 0),
                "allowance": float(ws.cell(r, 5).value or 0)
            }

    # Load lai nhan vien dang hoat dong
    emp_result = await db.execute(select(Employee).where(Employee.is_active == True))
    emp_map = {str(e.employee_code).lstrip("'"): e for e in emp_result.scalars().all()}

    processed = 0
    # Đọc dữ liệu từ dòng 3 (sau header)
    for r in range(3, ws.max_row + 1):
        emp_code = ws.cell(r, 2).value
        if not emp_code:
            continue
        
        emp_code = str(int(emp_code) if isinstance(emp_code, float) else emp_code).strip().lstrip("'")
        if emp_code not in emp_map:
            continue
            
        emp = emp_map[emp_code]
        info = file_emps.get(emp_code, {})
        base_salary = info.get("base_salary", 0)
        allowance = info.get("allowance", 0)

        # Upsert
        sal_result = await db.execute(select(MonthlySalary).where(
            and_(MonthlySalary.employee_id == emp.id, MonthlySalary.month_key == month_key)
        ))
        sal = sal_result.scalar_one_or_none()

        # Dong bo nguoc lai bang Employee (Luong co ban hien tai)
        emp.base_salary = base_salary

        if sal:
            sal.base_salary = base_salary
            sal.allowance = allowance
            sal.base_daily_wage = base_salary / standard_days if standard_days > 0 else 0
        else:
            sal = MonthlySalary(
                employee_id=emp.id,
                month_key=month_key,
                base_salary=base_salary,
                allowance=allowance,
                base_daily_wage=base_salary / standard_days if standard_days > 0 else 0
            )
            db.add(sal)
        processed += 1

    await db.commit()

    # Ghi nhật ký
    await log_audit(
        db, "monthly_salaries", month_key, "IMPORT", current_user.username,
        notes=f"Import luong thang {month_key} tu {file.filename}. {processed} NV. He so: {standard_days}"
    )
    await db.commit()

    return {"message": f"Import thành công lương cho {processed} nhân viên (Hệ số: {standard_days} ngày)", "processed": processed}


@router.post("/lock-month")
async def lock_month(
    month_key: str = Form(...),
    action: str = Form(...),  # 'lock' or 'unlock'
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN)),
):
    """Khóa/Mở khóa dữ liệu tháng. Chỉ admin."""
    config_result = await db.execute(select(MonthlyWorkdayConfig).where(MonthlyWorkdayConfig.month_key == month_key))
    config = config_result.scalar_one_or_none()

    if not config:
        config = MonthlyWorkdayConfig(month_key=month_key)
        db.add(config)

    config.is_locked = (action == 'lock')
    await db.commit()
    return {"message": "Đã chốt (khóa) dữ liệu tháng" if config.is_locked else "Đã mở khóa dữ liệu tháng"}


@router.get("/advances")
async def get_advances_summary(
    month_key: str = Query(..., description="YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Tổng tạm ứng theo nhân viên trong tháng"""
    query = select(AdvancePayment, Employee).join(Employee, AdvancePayment.employee_id == Employee.id)\
        .where(AdvancePayment.month_key == month_key)
    
    if current_user.role == UserRole.WORKER:
        query = query.where(AdvancePayment.employee_id == current_user.employee_id)
        
    query = query.order_by(Employee.employee_code)
    result = await db.execute(query)
    rows = result.all()
    from collections import defaultdict
    emp_advances = defaultdict(float)
    emp_info = {}
    for adv, emp in rows:
        emp_advances[emp.id] += float(adv.amount or 0)
        emp_info[emp.id] = {
            "employee_id": emp.id,
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
        }
    return [{"total_advance": emp_advances[eid], **emp_info[eid]} for eid in emp_advances]


class UpdateDependentsRequest(BaseModel):
    employee_id: int
    dependents: int


@router.put("/dependents")
async def update_dependents(
    req: UpdateDependentsRequest,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Cập nhật số người phụ thuộc của nhân viên"""
    emp = await db.get(Employee, req.employee_id)
    if not emp:
        raise HTTPException(404, "Không tìm thấy nhân viên")
    emp.dependents = max(0, req.dependents)
    await db.commit()
    return {"message": f"Đã cập nhật {emp.full_name}: {emp.dependents} người phụ thuộc"}


class UpdateBaseSalaryRequest(BaseModel):
    employee_id: int
    month_key: str
    base_salary: float


@router.put("/base")
async def update_base_salary(
    req: UpdateBaseSalaryRequest,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN)),
):
    """Admin cập nhật lương cơ bản của nhân viên cho tháng cụ thể."""
    await check_month_locked(db, req.month_key)

    emp = await db.get(Employee, req.employee_id)
    if not emp:
        raise HTTPException(404, "Không tìm thấy nhân viên")

    # Lấy standard_days từ config tháng
    config_res = await db.execute(
        select(MonthlyWorkdayConfig).where(MonthlyWorkdayConfig.month_key == req.month_key)
    )
    config = config_res.scalar_one_or_none()
    standard_days = float(config.company_work_days) if config else 26.0

    # Upsert MonthlySalary
    sal_res = await db.execute(
        select(MonthlySalary).where(
            and_(MonthlySalary.employee_id == req.employee_id, MonthlySalary.month_key == req.month_key)
        )
    )
    sal = sal_res.scalar_one_or_none()

    if sal:
        sal.base_salary = req.base_salary
        sal.base_daily_wage = req.base_salary / standard_days if standard_days > 0 else 0
        sal.updated_at = datetime.utcnow()
    else:
        sal = MonthlySalary(
            employee_id=req.employee_id,
            month_key=req.month_key,
            base_salary=req.base_salary,
            allowance=0,
            base_daily_wage=req.base_salary / standard_days if standard_days > 0 else 0,
        )
        db.add(sal)

    # Đồng bộ ngược lại bảng Employee
    emp.base_salary = req.base_salary

    await db.commit()

    await log_audit(
        db, "monthly_salaries", req.month_key, "UPDATE_BASE",
        current_user.username,
        notes=f"Sửa lương cơ bản {emp.full_name} ({emp.employee_code}): {req.base_salary:,.0f}đ"
    )
    await db.commit()

    return {
        "message": f"Đã cập nhật lương cơ bản {emp.full_name}: {req.base_salary:,.0f}đ",
        "base_salary": req.base_salary,
    }


# ─── Advance Loans ────────────────────────────────────────────────────────────

class CreateLoanRequest(BaseModel):
    employee_id: int
    loan_date: date
    total_amount: float
    advance_type: str = 'cash'          # cash | half_month | full_month | multi_month
    repayment_months: int = 1           # Số tháng trả
    monthly_repayment: Optional[float] = None  # Nếu None → auto = total/months
    start_month: str                    # YYYY-MM tháng đầu tiên bị trừ
    notes: Optional[str] = None


@router.get("/loans")
async def get_loans(
    employee_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Danh sách khoản tạm ứng (có thể lọc theo nhân viên / trạng thái)"""
    q = select(AdvanceLoan, Employee).join(Employee, AdvanceLoan.employee_id == Employee.id)
    if current_user.role == UserRole.WORKER:
        employee_id = current_user.employee_id
    if employee_id:
        q = q.where(AdvanceLoan.employee_id == employee_id)
    if status:
        q = q.where(AdvanceLoan.status == status)
    q = q.order_by(AdvanceLoan.loan_date.desc())
    result = await db.execute(q)

    rows = []
    for loan, emp in result.all():
        # Tính tổng đã trả từ advance_payments
        paid_res = await db.execute(
            select(func.sum(AdvancePayment.amount))
            .where(AdvancePayment.loan_id == loan.id)
        )
        paid = float(paid_res.scalar() or 0)
        rows.append({
            "id": loan.id,
            "employee_id": emp.id,
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
            "department": emp.department,
            "loan_date": loan.loan_date.isoformat(),
            "total_amount": float(loan.total_amount),
            "advance_type": loan.advance_type,
            "repayment_months": loan.repayment_months,
            "monthly_repayment": float(loan.monthly_repayment or 0),
            "start_month": loan.start_month,
            "paid_amount": paid,
            "remaining": max(0, float(loan.total_amount) - paid),
            "status": loan.status,
            "notes": loan.notes,
            "created_at": loan.created_at.isoformat() if loan.created_at else None,
        })
    return rows


@router.post("/loans")
async def create_loan(
    req: CreateLoanRequest,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Tạo khoản tạm ứng + tự động sinh advance_payments theo kế hoạch trả"""
    # Kiểm tra khóa tháng cho toàn bộ các kỳ trả sắp tạo
    months = max(1, req.repayment_months)
    def add_months(d: date, n: int) -> date:
        m = d.month - 1 + n
        return d.replace(year=d.year + m // 12, month=m % 12 + 1, day=1)

    start_dt = datetime.strptime(req.start_month, "%Y-%m").date().replace(day=1)
    for i in range(months):
        month_dt = add_months(start_dt, i)
        mk = month_dt.strftime("%Y-%m")
        await check_month_locked(db, mk)

    emp = await db.get(Employee, req.employee_id)
    if not emp:
        raise HTTPException(404, "Không tìm thấy nhân viên")

    months = max(1, req.repayment_months)
    per_month = req.monthly_repayment if req.monthly_repayment else round(req.total_amount / months)

    loan = AdvanceLoan(
        employee_id=req.employee_id,
        loan_date=req.loan_date,
        total_amount=req.total_amount,
        advance_type=req.advance_type,
        repayment_months=months,
        monthly_repayment=per_month,
        start_month=req.start_month,
        paid_amount=0,
        status='active',
        notes=req.notes,
        created_by=current_user.username,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(loan)
    await db.flush()  # để có loan.id

    def add_months(d: date, n: int) -> date:
        m = d.month - 1 + n
        return d.replace(year=d.year + m // 12, month=m % 12 + 1, day=1)

    # Sinh advance_payment cho từng kỳ
    start = datetime.strptime(req.start_month, "%Y-%m").date().replace(day=1)
    remaining = float(req.total_amount)
    for i in range(months):
        month_dt = add_months(start, i)
        mk = month_dt.strftime("%Y-%m")
        # Kỳ cuối trả phần còn lại để tránh sai số làm tròn
        amt = per_month if i < months - 1 else remaining
        remaining -= amt
        pay = AdvancePayment(
            employee_id=req.employee_id,
            loan_id=loan.id,
            advance_date=month_dt,
            month_key=mk,
            amount=round(amt),
            installment_no=i + 1,
            input_mode='amount',
            notes=f"Kỳ {i+1}/{months} — {req.notes or ''}".strip(" —"),
        )
        db.add(pay)

    await db.commit()
    return {"message": f"Tạo thành công khoản ứng {req.total_amount:,.0f}đ cho {emp.full_name} — {months} kỳ", "loan_id": loan.id}


@router.delete("/loans/{loan_id}")
async def cancel_loan(
    loan_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Hủy khoản tạm ứng — xóa các kỳ chưa đến hạn"""
    loan = await db.get(AdvanceLoan, loan_id)
    if not loan:
        raise HTTPException(404, "Không tìm thấy khoản ứng")

    today_mk = datetime.utcnow().strftime("%Y-%m")
    
    # Kiểm tra khóa tháng đối với các kỳ trả sắp bị xóa (kỳ ở tương lai)
    result = await db.execute(
        select(AdvancePayment.month_key).where(
            AdvancePayment.loan_id == loan_id,
            AdvancePayment.month_key > today_mk,
        )
    )
    target_mks = result.scalars().all()
    for mk in target_mks:
        await check_month_locked(db, mk)
    # Xóa các kỳ ở tháng tương lai
    await db.execute(
        delete(AdvancePayment).where(
            AdvancePayment.loan_id == loan_id,
            AdvancePayment.month_key > today_mk,
        )
    )
    loan.status = 'cancelled'
    loan.updated_at = datetime.utcnow()
    await db.commit()
    return {"message": "Đã hủy các kỳ trả chưa đến hạn"}


@router.get("/loans/{loan_id}/installments")
async def get_loan_installments(
    loan_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Danh sách kỳ trả của 1 khoản ứng"""
    result = await db.execute(
        select(AdvancePayment)
        .where(AdvancePayment.loan_id == loan_id)
        .order_by(AdvancePayment.month_key)
    )
    today_mk = datetime.utcnow().strftime("%Y-%m")
    rows = []
    for p in result.scalars().all():
        rows.append({
            "id": p.id,
            "month_key": p.month_key,
            "amount": float(p.amount or 0),
            "installment_no": p.installment_no,
            "paid": p.month_key <= today_mk,
            "notes": p.notes,
        })
    return rows


def calc_tncn(taxable: float) -> int:
    if taxable <= 0:
        return 0
    tax = 0.0
    tax += min(taxable, 10000000.0) * 0.05
    tax += min(max(taxable - 10000000.0, 0.0), 20000000.0) * 0.10
    tax += min(max(taxable - 30000000.0, 0.0), 30000000.0) * 0.20
    tax += min(max(taxable - 60000000.0, 0.0), 40000000.0) * 0.30
    tax += max(taxable - 100000000.0, 0.0) * 0.35
    return int(round(tax))


def vietnamese_number_to_words(number: int) -> str:
    if number == 0:
        return "Không đồng"
    
    units = ["", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
    
    def read_three_digits(n: int, show_zero_hundred: bool = False) -> str:
        hundred = n // 100
        ten = (n % 100) // 10
        unit = n % 10
        
        res = []
        if hundred > 0 or show_zero_hundred:
            res.append(units[hundred] + " trăm")
            
        if ten > 1:
            res.append(units[ten] + " mươi")
            if unit == 1:
                res.append("mốt")
            elif unit == 5:
                res.append("lăm")
            elif unit > 0:
                res.append(units[unit])
        elif ten == 1:
            res.append("mười")
            if unit == 5:
                res.append("lăm")
            elif unit > 0:
                res.append(units[unit])
        else: # ten == 0
            if (hundred > 0 or show_zero_hundred) and unit > 0:
                res.append("lẻ")
            if unit > 0:
                res.append(units[unit])
                
        return " ".join(res)

    groups = []
    temp = number
    while temp > 0:
        groups.append(temp % 1000)
        temp //= 1000
        
    scales = ["", "nghìn", "triệu", "tỷ", "nghìn tỷ", "triệu tỷ"]
    words = []
    
    for idx, val in enumerate(groups):
        if val > 0:
            show_zero = (idx < len(groups) - 1)
            group_word = read_three_digits(val, show_zero)
            scale_word = scales[idx]
            if scale_word:
                words.append(group_word + " " + scale_word)
            else:
                words.append(group_word)
                
    words.reverse()
    result_str = ", ".join(words).replace("  ", " ").strip()
    return result_str[0].upper() + result_str[1:] + " đồng"


@router.get("/export-detail")
async def export_payroll_detail(
    month_key: str = Query(..., description="YYYY-MM"),
    ot_style: Optional[str] = Query("old", description="old or new"),
    night_allowance_rate: Optional[float] = Query(100000.0),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Xuất Excel bảng kê tiền lương chi tiết song ngữ Trung - Việt theo mẫu"""
    from app.routers.attendance import get_attendance
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from app.models.department import Department
    
    # Fetch department mapping for codes
    dept_q = select(Department)
    dept_res = await db.execute(dept_q)
    dept_map = {d.name: d for d in dept_res.scalars().all()}
    
    # 1. Fetch data using the shared get_attendance endpoint logic
    att_res = await get_attendance(
        month_key=month_key,
        night_allowance_rate=night_allowance_rate,
        ot_style=ot_style,
        db=db,
        current_user=current_user
    )
    
    # 2. Fetch Base Salaries
    salary_q = select(MonthlySalary).where(MonthlySalary.month_key == month_key)
    salary_res = await db.execute(salary_q)
    salary_map = {s.employee_id: s for s in salary_res.scalars().all()}
    
    # 3. Fetch Advances
    year, month = map(int, month_key.split("-"))
    adv_q = select(AdvancePayment).where(AdvancePayment.month_key == month_key)
    adv_res = await db.execute(adv_q)
    advances = adv_res.scalars().all()
    adv_map = {}
    for a in advances:
        adv_map[a.employee_id] = adv_map.get(a.employee_id, 0.0) + float(a.amount or 0)
        
    # 4. Fetch Employees
    emp_q = select(Employee)
    emp_res = await db.execute(emp_q)
    employees_db = {e.id: e for e in emp_res.scalars().all()}
    
    # 5. Month Config
    config_result = await db.execute(select(MonthlyWorkdayConfig).where(MonthlyWorkdayConfig.month_key == month_key))
    config = config_result.scalar_one_or_none()
    standard_days = float(config.company_work_days) if config else 26.0
    
    # 6. Initialize workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Bang Ke Luong {month}-{year}"
    ws.views.sheetView[0].showGridLines = True
    
    # Border & Font styles
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    double_bottom_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    font_header = Font(name='Arial', size=9, bold=True)
    font_data = Font(name='Arial', size=9)
    font_data_bold = Font(name='Arial', size=9, bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    fill_header = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid") # soft light orange/yellow
    fill_gross = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # soft yellow
    fill_advance = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid") # soft pink
    fill_net = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") # soft green for actual pay!
    
    # Title Block
    ws['A1'] = "CÔNG TY TNHH HIỆP LỢI - 越南協利有限公司"
    ws['A1'].font = Font(name='Arial', size=11, bold=True)
    
    ws.merge_cells('A2:Z2')
    title_cell = ws['A2']
    title_cell.value = f"BẢNG KÊ TIỀN LƯƠNG VÀ CÁC KHOẢN THU NHẬP KHÁC THÁNG {month} NĂM {year}***{year} 年 {month} 月份文房薪資表"
    title_cell.font = Font(name='Arial', size=13, bold=True)
    title_cell.alignment = align_center
    ws.row_dimensions[2].height = 30
    
    # Headers definition: (text, start_col, end_col, start_row, end_row)
    headers = [
        ("STT\n序號", 1, 1, 3, 4),
        ("HỌ VÀ TÊN\n姓名", 2, 2, 3, 4),
        ("NGÀY VÀO\n入廠日期", 3, 3, 3, 4),
        ("CHỨC VỤ\n職務", 4, 5, 3, 4),
        ("基本薪資\nLƯƠNG CĂN\nBẢN", 6, 6, 3, 4),
        ("日數\nNGÀY\nCÔNG", 7, 7, 3, 4),
        ("金額\nTHÀNH\nTIỀN", 8, 8, 3, 4),
        
        ("加 班 平 常 T.CA THƯỜNG", 9, 11, 3, 3),
        ("總小時\nTỔNG\nGIỜ", 9, 9, 4, 4),
        ("加班費\nTIỀN\nTĂNG CA", 10, 10, 4, 4),
        ("金額\nTHÀNH\nTIỀN", 11, 11, 4, 4),
        
        ("星期天 T.CA CN", 12, 14, 3, 3),
        ("總小時\nTỔNG\nGIỜ", 12, 12, 4, 4),
        ("加班費\nTIỀN\nTĂNG CA", 13, 13, 4, 4),
        ("金額\nTHÀNH\nTIỀN", 14, 14, 4, 4),
        
        ("職務獎\nBỒI\nDƯỠNG &\nTRÁCH\nNHIỆM", 15, 15, 3, 4),
        ("補助育兒\nBỒI DƯỠNG\n/PHỤ CẤP\nNUÔI\nCON NHỎ\n<6 TUỔI", 16, 16, 3, 4),
        ("補助、交通、電話、夜班\nXĂNG,\nĐIỆN\nTHOẠI,\nNHÀ Ở,\nCA ĐÊM\n(PC KO\nCHỊU\nTHUẾ)", 17, 17, 3, 4),
        ("全勤獎\nTIỀN\nCHUYÊN\nCẦN", 18, 18, 3, 4),
        ("合計\nTỔNG\nCỘNG", 19, 19, 3, 4),
        ("Trừ BHXH,\nBHYT,\nBHTN,\nBHTNLĐ-\nBNN\n10.5%", 20, 20, 3, 4),
        ("個人\n所得稅\nThuế\nTNCN", 21, 21, 3, 4),
        ("工會費\nTIỀN\nCÔNG\nĐOÀN\nPHÍ CĐ", 22, 22, 3, 4),
        ("績效獎金\nTHƯỞNG\nNĂNG\nSUẤT", 23, 23, 3, 4),
        ("借支\nTẠM\nỨNG", 24, 24, 3, 4),
        ("實發金額\nTHỰC\nLÃNH", 25, 25, 3, 4),
        ("簽收\nKÝ TÊN", 26, 26, 3, 4),
    ]
    
    for h_text, sc, ec, sr, er in headers:
        for r_idx in range(sr, er + 1):
            for c_idx in range(sc, ec + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                cell.font = font_header
                cell.alignment = align_center
                cell.fill = fill_header
                if r_idx == sr and c_idx == sc:
                    cell.value = h_text
        if sc != ec or sr != er:
            ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)
                
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 42
    
    # Row data
    current_row = 5
    stt = 1
    sum_net = 0.0

    for row in att_res.rows:
        emp_id = row.employee_id
        emp = employees_db.get(emp_id)
        sal = salary_map.get(emp_id)
        
        base_salary = float(sal.base_salary) if (sal and sal.base_salary is not None) else float(emp.base_salary or 0.0) if emp else 0.0
        fixed_allowance = float(sal.allowance or 0) if sal else 0.0
        dependents = emp.dependents if emp else 0
        summary = row.summary or {}
        
        actual_days = float(summary.get("total_present") or 0) + float(summary.get("total_paid_leave") or 0)
        
        ot_wd = float(summary.get("total_ot_weekday") or summary.get("total_ot") or 0)
        ot_sun = float(summary.get("total_ot_sunday") or 0)
        ot_hol = float(summary.get("total_ot_holiday") or 0)
        
        meal_allowance = float(summary.get("total_meal_allowance") or 0)
        night_allowance = float(summary.get("total_night_allowance") or 0)
        advance = float(adv_map.get(emp_id, 0.0))
        
        daily_rate = base_salary / standard_days if standard_days > 0 else 0.0
        hourly_rate = daily_rate / 8
        
        is_usd = 0.0 < base_salary < 100000
        coef = float(sal.salary_coefficient or 1.0) if sal else 1.0
        
        salary_from_days = round(coef * base_salary) if is_usd else round(actual_days * daily_rate)
        ot_pay_wd = 0 if is_usd else round(ot_wd * hourly_rate * 1.5)
        ot_pay_sun = 0 if is_usd else round(ot_sun * hourly_rate * 2.0)
        ot_pay_hol = 0 if is_usd else round(ot_hol * hourly_rate * 3.0)
        ot_pay = ot_pay_wd + ot_pay_sun + ot_pay_hol
        
        gross = salary_from_days + ot_pay + fixed_allowance + night_allowance
        bhxh = 0 if is_usd else round((base_salary + fixed_allowance) * 0.105)
        union_fee = 0 if is_usd else round(base_salary * 0.01)
        taxable = max(0, gross - bhxh - 11000000 - dependents * 4400000)
        tncn = calc_tncn(taxable)
        net = round(gross - bhxh - union_fee - tncn - advance)
        
        sum_net += net
        
        join_date_str = ""
        if emp and emp.join_date:
            join_date_str = emp.join_date.strftime("%d/%m/%Y")
            
        # Map department name to department code if possible
        dept_obj = dept_map.get(emp.department) if emp else None
        dept_code = dept_obj.code if dept_obj else (emp.department if emp else "")
        
        row_data = [
            stt,
            emp.full_name if emp else "",
            join_date_str,
            dept_code,
            emp.position if (emp and emp.position) else "",
            base_salary, # F
            coef if is_usd else actual_days, # G
            f"=F{current_row}*G{current_row}" if is_usd else f"=ROUND(F{current_row}*G{current_row}/{standard_days}, 0)", # H
            0 if is_usd else ot_wd, # I
            0 if is_usd else f"=ROUND(F{current_row}/{standard_days}/8*1.5, 0)", # J
            0 if is_usd else f"=ROUND(I{current_row}*F{current_row}/{standard_days}/8*1.5, 0)", # K
            0 if is_usd else ot_sun, # L
            0 if is_usd else f"=ROUND(F{current_row}/{standard_days}/8*2, 0)", # M
            0 if is_usd else f"=ROUND(L{current_row}*F{current_row}/{standard_days}/8*2, 0)", # N
            fixed_allowance, # O
            0, # P
            night_allowance, # Q
            0, # R
            f"=H{current_row}+K{current_row}+N{current_row}+O{current_row}+P{current_row}+Q{current_row}+R{current_row}", # S
            0 if is_usd else f"=ROUND((F{current_row}+O{current_row})*0.105, 0)", # T
            tncn, # U
            0 if is_usd else f"=ROUND(F{current_row}*0.01, 0)", # V
            0, # W
            advance, # X
            f"=S{current_row}-T{current_row}-U{current_row}-V{current_row}-W{current_row}-X{current_row}", # Y
            "" # Z
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.border = thin_border
            c.font = font_data
            
            # Alignments
            if col_idx in (1, 3, 4, 7, 9, 12): # STT, Date, Dept, Days, OT Hours
                c.alignment = align_center
            elif col_idx in (2, 5): # Name, Position
                c.alignment = align_left
            else:
                c.alignment = align_right
                
            # Number formats and color fills
            if col_idx in (6, 8, 10, 11, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25): # Money
                c.number_format = '#,##0;-#,##0;"-"'
                if col_idx == 19: # Gross
                    c.fill = fill_gross
                elif col_idx == 24: # Advance
                    c.fill = fill_advance
                elif col_idx == 25: # Net
                    c.fill = fill_net
            elif col_idx in (7, 9, 12): # Decimals
                c.number_format = '0.00;-0.00;"-"'
                    
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        stt += 1

    # Total Row
    tot_row = current_row
    for col in range(1, 6):
        c = ws.cell(row=tot_row, column=col)
        c.border = double_bottom_border
        c.font = font_data_bold
        if col == 1:
            c.value = "TỔNG CỘNG\n合計"
            c.alignment = align_center
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=5)
            
    for col in range(6, 27):
        col_letter = get_column_letter(col)
        c = ws.cell(row=tot_row, column=col)
        c.border = double_bottom_border
        c.font = font_data_bold
        
        if col in (10, 13, 26): # OT rates and signature are empty
            c.value = ""
        else:
            c.value = f"=SUM({col_letter}5:{col_letter}{tot_row-1})"
            if col in (7, 9, 12):
                c.alignment = align_center
                c.number_format = '0.00;-0.00;"-"'
            else:
                c.alignment = align_right
                c.number_format = '#,##0;-#,##0;"-"'
                
                # Apply column fills to total row
                if col == 19: # Gross
                    c.fill = fill_gross
                elif col == 24: # Advance
                    c.fill = fill_advance
                elif col == 25: # Net
                    c.fill = fill_net

    ws.row_dimensions[tot_row].height = 24
    
    # Financial details & Signatures below
    current_row = tot_row + 2
    ws.cell(row=current_row, column=2, value="Thành tiền:").font = font_data_bold
    ws.cell(row=current_row, column=6, value=f"=Y{tot_row}").font = font_data_bold
    ws.cell(row=current_row, column=6).number_format = '#,##0'
    ws.cell(row=current_row, column=6).alignment = align_left
    
    current_row += 1
    ws.cell(row=current_row, column=2, value="Bằng chữ:").font = font_data_bold
    try:
        words = vietnamese_number_to_words(int(sum_net))
        ws.cell(row=current_row, column=6, value=words).font = Font(name='Arial', size=9, italic=True)
        ws.cell(row=current_row, column=6).alignment = align_left
    except Exception:
        pass
        
    current_row += 2
    ws.cell(row=current_row, column=2, value="Kế toán").font = font_data_bold
    ws.cell(row=current_row, column=2).alignment = align_center
    
    ws.cell(row=current_row, column=12, value="Chủ quản").font = font_data_bold
    ws.cell(row=current_row, column=12).alignment = align_center
    
    import datetime as dt_mod
    today = dt_mod.date.today()
    date_str = f"Ngày {today.day:02d} tháng {today.month:02d} năm {today.year}"
    ws.cell(row=current_row, column=20, value=f"{date_str}\nTổng giám đốc").font = font_data_bold
    ws.cell(row=current_row, column=20).alignment = align_center
    
    # Set widths
    column_widths = {
        1: 6, 2: 24, 3: 12, 4: 12, 5: 14, 6: 14, 7: 9, 8: 14, 9: 8, 10: 12,
        11: 14, 12: 8, 13: 12, 14: 14, 15: 14, 16: 14, 17: 16, 18: 12, 19: 16,
        20: 14, 21: 12, 22: 12, 23: 12, 24: 12, 25: 16, 26: 14
    }
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"Bang_luong_chi_tiet_{month_key}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
