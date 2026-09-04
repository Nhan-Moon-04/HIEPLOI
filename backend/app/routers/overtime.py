import math
from io import BytesIO
from typing import List, Optional
from datetime import date, datetime, time, timedelta
import calendar
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import get_db
from app.models.schedule import WorkSchedule
from app.models.employee import Employee
from app.models.shift import ShiftTemplate
from app.models.attendance import AttendanceDaily, AttendanceLog
from app.models.holiday import CompanyHoliday
from app.models.x_overtime import XOvertimeConfig
from app.models.user import AppUser, UserRole
from app.middleware.auth import get_current_user
from app.routers.attendance import evaluate_attendance, parse_time, check_holiday_applies_to_employee
from app.services.nu_shift import (
    is_nu_dynamic_shift_code, build_nu_shift_day_results,
    XNU_MODE_1, XNU_MODE_2, XNU_MODE_3, NU_MORNING_MODE, NU_NIGHT_MODE, NU_STANDARD_HOURS,
)
from pydantic import BaseModel

router = APIRouter(prefix="/overtime", tags=["Overtime - Tang Ca"])

DOW_VN = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]

_TX_CODES = {"TX1", "TX2"}


def _nu_display_ot_hours(mode: str, check_in_dt: datetime, check_out_dt: datetime) -> float:
    """Hiển thị OT cho ca NU: ca sáng 3h, ca tối 4h, thấp hơn nếu ra sớm."""
    baseline = 4.0 if mode == NU_NIGHT_MODE else 3.0
    worked_hours = max((check_out_dt - check_in_dt).total_seconds() / 3600.0, 0.0)
    raw_ot = max(worked_hours - NU_STANDARD_HOURS, 0.0)
    return min(baseline, _to_half(raw_ot))

# XNU mode → (shift_start, shift_end, ends_next_day)
_XNU_SCHEDULE = {
    XNU_MODE_1: (time(6, 0),  time(14, 0), False),
    XNU_MODE_2: (time(14, 0), time(22, 0), False),
    XNU_MODE_3: (time(22, 0), time(6, 0),  True),
}


def _round_ot_minutes(raw_minutes: float) -> float:
    """15-min threshold, then round up to nearest 0.5h slot."""
    if raw_minutes <= 15:
        return 0.0
    return math.ceil((raw_minutes - 15) / 30) * 0.5


def _to_half(h: float) -> float:
    """Round any hours value to the nearest 0.5h (standard round-half-up)."""
    return math.floor(h * 2 + 0.5) / 2


def _ceil_30min(dt: datetime) -> datetime:
    """Round a datetime UP to the next 30-minute boundary."""
    total = dt.hour * 60 + dt.minute
    if total % 30 == 0:
        return dt.replace(second=0, microsecond=0)
    rounded = math.ceil(total / 30) * 30
    return dt.replace(hour=(rounded // 60) % 24, minute=rounded % 60,
                      second=0, microsecond=0)


def _round_nearest_30min(dt: datetime) -> datetime:
    """Làm tròn đến 30 phút gần nhất.
    0-15p → :00, 16-30p → :30, 31-45p → :30, 46-59p → giờ chẵn kế."""
    m = dt.minute
    if m <= 15:
        rounded_min = 0
        h = dt.hour
    elif m <= 45:
        rounded_min = 30
        h = dt.hour
    else:
        rounded_min = 0
        h = (dt.hour + 1) % 24
    return dt.replace(hour=h, minute=rounded_min, second=0, microsecond=0)


async def _compute_actual_ot(month_key: str, db: AsyncSession, current_user) -> list:
    """Tính OT thực tế dựa trên giờ chấm công so với giờ ca quy định."""
    year, month = map(int, month_key.split("-"))
    days_in_month = calendar.monthrange(year, month)[1]
    first_day = date(year, month, 1)
    last_day = date(year, month, days_in_month)

    from sqlalchemy import or_

    # ── Shifts ─────────────────────────────────────────────────────────────
    shift_res = await db.execute(select(ShiftTemplate))
    shifts_by_id = {}
    shifts_by_code = {}
    for s in shift_res.scalars().all():
        shifts_by_id[s.id] = s
        shifts_by_code[s.code] = s

    # ── Employees ───────────────────────────────────────────────────────────
    emp_q = select(Employee).where(and_(
        or_(Employee.join_date.is_(None), Employee.join_date <= last_day),
        or_(Employee.leave_date.is_(None), Employee.leave_date >= first_day),
        or_(Employee.is_active == True, Employee.leave_date.is_not(None)),
    ))
    if current_user.role == UserRole.WORKER:
        emp_q = emp_q.where(Employee.id == current_user.employee_id)
    employees = list((await db.execute(emp_q)).scalars().all())

    # Sắp xếp theo bộ phận (sort_order) → nhân viên (sort_order) → mã NV
    from app.models.department import Department
    dept_order_q = await db.execute(select(Department.name, Department.sort_order))
    dept_order_map = {row[0]: row[1] for row in dept_order_q.all() if row[0]}

    def _emp_sort_key(e):
        d_order = dept_order_map.get(e.department, 9999) if e.department else 9999
        e_order = e.sort_order if e.sort_order is not None else 9999
        try:
            code_num = int(e.employee_code)
        except ValueError:
            code_num = 999999
        return (d_order, e.department or "", e_order, code_num)

    employees.sort(key=_emp_sort_key)

    # ── Schedule overrides ──────────────────────────────────────────────────
    override_map = {}
    for ws in (await db.execute(
        select(WorkSchedule).where(WorkSchedule.month_key == month_key)
    )).scalars().all():
        override_map[(ws.employee_id, ws.work_date.day)] = ws.shift_id

    # ── Attendance daily ────────────────────────────────────────────────────
    att_map = {}
    for a in (await db.execute(
        select(AttendanceDaily).where(and_(
            AttendanceDaily.work_date >= first_day,
            AttendanceDaily.work_date <= last_day,
        ))
    )).scalars().all():
        att_map[(a.employee_id, a.work_date)] = a

    # ── Holidays ───────────────────────────────────────────────────────────
    holiday_q = select(CompanyHoliday).where(
        and_(CompanyHoliday.holiday_date >= first_day, CompanyHoliday.holiday_date <= last_day, CompanyHoliday.is_active == True)
    )
    holiday_result = await db.execute(holiday_q)
    holidays_in_range = list(holiday_result.scalars().all())

    # Load holiday targets (for employee scope)
    from app.models.holiday import HolidayTargetEmployee
    holiday_ids = [h.id for h in holidays_in_range]
    holiday_targets_map = {}
    if holiday_ids:
        target_res = await db.execute(
            select(HolidayTargetEmployee.holiday_id, HolidayTargetEmployee.employee_id)
            .where(HolidayTargetEmployee.holiday_id.in_(holiday_ids))
        )
        for h_id, emp_id in target_res.all():
            if h_id not in holiday_targets_map:
                holiday_targets_map[h_id] = set()
            holiday_targets_map[h_id].add(emp_id)

    # ── Holiday exceptions ──────────────────────────────────────────────────
    from app.models.holiday import HolidayException
    exc_q = select(HolidayException.employee_id, CompanyHoliday.holiday_date).join(
        CompanyHoliday, HolidayException.holiday_id == CompanyHoliday.id
    ).where(
        and_(
            CompanyHoliday.holiday_date >= first_day,
            CompanyHoliday.holiday_date <= last_day,
            CompanyHoliday.is_active == True
        )
    )
    exc_res = await db.execute(exc_q)
    holiday_exceptions = {(row[0], row[1]) for row in exc_res.all()}

    # ── NU raw logs (for mode detection) ───────────────────────────────────
    log_res = await db.execute(
        select(AttendanceLog).where(and_(
            AttendanceLog.event_time >= datetime.combine(first_day, time(0, 0)),
            AttendanceLog.event_time <= datetime.combine(last_day + timedelta(days=1), time(12, 0)),
        ))
    )
    emp_code_to_id = {str(e.employee_code).lstrip("'"): e.id for e in employees}
    logs_with_id = []
    for lg in log_res.scalars().all():
        eid = emp_code_to_id.get(str(lg.employee_code).lstrip("'"))
        if eid:
            lg.employee_id = eid
            logs_with_id.append(lg)

    # ── Build NU shift results ──────────────────────────────────────────────
    nu_shift_code_map = {}
    emp_id_list = [e.id for e in employees]
    for emp in employees:
        default_shift = shifts_by_code.get(emp.default_shift_code)
        for d in range(1, days_in_month + 1):
            dt = date(year, month, d)
            ov_id = override_map.get((emp.id, d))
            if ov_id:
                s = shifts_by_id.get(ov_id)
                if s:
                    nu_shift_code_map[(emp.id, dt)] = s.code
            elif default_shift and is_nu_dynamic_shift_code(default_shift.code):
                nu_shift_code_map[(emp.id, dt)] = default_shift.code

    nu_results = build_nu_shift_day_results(
        nu_shift_code_map=nu_shift_code_map,
        employee_id_list=emp_id_list,
        attendance_log_rows=logs_with_id,
    )

    # ── Build rows ──────────────────────────────────────────────────────────
    rows = []
    stt = 0

    for emp in employees:
        default_shift = shifts_by_code.get(emp.default_shift_code)

        for d in range(1, days_in_month + 1):
            dt = date(year, month, d)
            if (emp.join_date and dt < emp.join_date) or (emp.leave_date and dt > emp.leave_date):
                continue
            is_sunday = dt.weekday() == 6
            dow = DOW_VN[dt.weekday()]

            # Determine shift
            ov_id = override_map.get((emp.id, d))
            shift = shifts_by_id.get(ov_id) if ov_id else default_shift
            if shift and shift.is_leave_code:
                continue

            # Check if there is an active holiday on this date that applies to the employee
            active_holiday = None
            for h in holidays_in_range:
                if h.holiday_date == dt:
                    is_exception = (emp.id, dt) in holiday_exceptions
                    if not is_exception and check_holiday_applies_to_employee(h, emp.id, emp.department, holiday_targets_map):
                        active_holiday = h
                        break

            # Attendance times — prefer NU-corrected times
            nu_res = nu_results.get((emp.id, dt))
            att = att_map.get((emp.id, dt))
            if nu_res:
                check_in_dt = nu_res.check_in
                check_out_dt = nu_res.check_out
                shift_code = nu_res.shift_code
            else:
                check_in_dt = att.first_check_in if att else None
                check_out_dt = att.last_check_out if att else None
                shift_code = shift.code if shift else None

            is_holiday = False
            is_half_day_worked = False

            if active_holiday:
                if active_holiday.duration == "half":
                    has_punches = False
                    if nu_res:
                        has_punches = bool(nu_res.check_in and nu_res.check_out)
                    else:
                        has_punches = bool(check_in_dt and check_out_dt)
                    
                    if has_punches:
                        is_half_day_worked = True
                        is_holiday = False
                    else:
                        is_holiday = True
                        check_in_dt = None
                        check_out_dt = None
                        nu_res = None
                else:
                    is_holiday = True
                    check_in_dt = None
                    check_out_dt = None
                    nu_res = None

            if not check_in_dt or not check_out_dt:
                continue

            if is_half_day_worked:
                continue

            ot_hours = 0.0
            shift_hours_str = ""

            # ── Sunday: tính giờ OT thực tế (không ×2 — kế toán tự nhân) ──────
            if is_sunday:
                is_tx  = bool(shift and (shift.code or "").upper() in _TX_CODES)
                is_xnu = bool(shift and (shift.code or "").upper() == "XNU")
                s_code = (shift_code or "").upper()
                is_nu_sun = is_nu_dynamic_shift_code(s_code) or bool(nu_res)

                # Xác định lịch ca cho chủ nhật
                if is_xnu and nu_res:
                    sched = _XNU_SCHEDULE.get(nu_res.mode)
                    if not sched:
                        continue
                    s_start_t, s_end_t, s_next = sched
                elif shift and shift.start_time and shift.end_time:
                    s_start_t = shift.start_time if isinstance(shift.start_time, time) else parse_time(shift.start_time)
                    s_end_t   = shift.end_time   if isinstance(shift.end_time,   time) else parse_time(shift.end_time)
                    s_next    = bool(getattr(shift, 'is_night_shift', False))
                else:
                    s_start_t = s_end_t = None
                    s_next = False

                if s_start_t and s_end_t:
                    s_start_dt = datetime.combine(dt, s_start_t)
                    s_end_dt   = datetime.combine(dt + timedelta(days=1) if s_next else dt, s_end_t)
                    shift_hours_str = f"{s_start_t.strftime('%H:%M')}-{s_end_t.strftime('%H:%M')}"
                elif is_nu_sun:
                    mode = nu_res.mode if nu_res else (
                        NU_NIGHT_MODE if (check_in_dt.hour >= 17 or check_out_dt < check_in_dt) else NU_MORNING_MODE
                    )
                    shift_hours_str = "18:00-06:00" if mode == NU_NIGHT_MODE else "06:00-18:00"
                else:
                    s_start_dt = s_end_dt = None
                    shift_hours_str = "Chủ nhật"

                # Giờ vào hiệu lực:
                #   TX: làm tròn 30p gần nhất
                #   Non-TX vào sớm hoặc trễ ≤15p: lấy đúng giờ bắt đầu ca
                #   Non-TX vào trễ >15p: làm tròn lên 30p
                if is_tx:
                    effective_in = _round_nearest_30min(check_in_dt)
                elif s_start_dt and check_in_dt <= s_start_dt + timedelta(minutes=15):
                    effective_in = s_start_dt
                else:
                    effective_in = _ceil_30min(check_in_dt)
                raw_h = max((check_out_dt - effective_in).total_seconds() / 3600.0, 0.0)

                # Trừ 30p nghỉ giữa giờ nếu:
                #   - Ca có span >= 8.5h (tức ca 9h, ví dụ X, TX — span 7h-16h=9h)
                #   - Và giờ vào hiệu lực < giờ kết thúc ca - 5h (tức có làm buổi sáng)
                #   - XNU span = 8h → không đủ điều kiện, không trừ
                if s_start_dt and s_end_dt:
                    span_h = (s_end_dt - s_start_dt).total_seconds() / 3600.0
                    if span_h >= 8.5 and effective_in < (s_end_dt - timedelta(hours=5)):
                        raw_h = max(raw_h - 0.5, 0.0)

                ot_hours = _to_half(raw_h)  # không ×2; kế toán tự có công thức trong Excel

                if ot_hours <= 0:
                    continue

            # ── Ngày thường ────────────────────────────────────────────────
            else:
                s_code = (shift_code or "").upper()
                is_tx = s_code in _TX_CODES
                is_nu = is_nu_dynamic_shift_code(s_code) or bool(nu_res)

                if is_nu and s_code != "XNU":
                    mode = nu_res.mode if nu_res else (
                        NU_NIGHT_MODE if (check_in_dt.hour >= 17 or check_out_dt < check_in_dt) else NU_MORNING_MODE
                    )
                    ot_hours = _nu_display_ot_hours(mode, check_in_dt, check_out_dt)
                    if nu_res and mode == NU_NIGHT_MODE:
                        shift_hours_str = "18:00-06:00"
                    elif nu_res and mode == NU_MORNING_MODE:
                        shift_hours_str = "06:00-18:00"
                    elif shift and shift.start_time and shift.end_time:
                        shift_start_t = shift.start_time if isinstance(shift.start_time, time) else parse_time(shift.start_time)
                        shift_end_t = shift.end_time if isinstance(shift.end_time, time) else parse_time(shift.end_time)
                        shift_hours_str = f"{shift_start_t.strftime('%H:%M')}-{shift_end_t.strftime('%H:%M')}"
                    else:
                        shift_hours_str = "18:00-06:00" if mode == NU_NIGHT_MODE else "06:00-18:00"
                else:
                    if not shift:
                        continue

                    # Xác định giờ kết thúc ca
                    if is_nu and nu_res and s_code == "XNU":
                        sched = _XNU_SCHEDULE.get(nu_res.mode)
                        if not sched:
                            continue
                        _, shift_end_t, end_next_day = sched
                        shift_hours_str = f"{sched[0].strftime('%H:%M')}-{shift_end_t.strftime('%H:%M')}"
                    elif shift.start_time and shift.end_time:
                        shift_end_t = shift.end_time if isinstance(shift.end_time, time) else parse_time(shift.end_time)
                        shift_start_t = shift.start_time if isinstance(shift.start_time, time) else parse_time(shift.start_time)
                        end_next_day = bool(getattr(shift, 'is_night_shift', False))
                        # NU night-mode: end time is next day
                        if is_nu and nu_res and nu_res.mode == NU_NIGHT_MODE:
                            end_next_day = True
                        shift_hours_str = f"{shift_start_t.strftime('%H:%M')}-{shift_end_t.strftime('%H:%M')}"
                    else:
                        continue

                    shift_end_dt = datetime.combine(
                        dt + timedelta(days=1) if end_next_day else dt,
                        shift_end_t,
                    )

                    # OT từ ra trễ
                    ot_checkout = 0.0
                    if check_out_dt > shift_end_dt:
                        raw_min = (check_out_dt - shift_end_dt).total_seconds() / 60.0
                        if not is_tx:
                            raw_min -= 30   # 30p nghỉ ngơi trước OT
                        ot_checkout = _round_ot_minutes(raw_min)

                    # TX1/TX2: OT vào sớm — làm tròn 30p gần nhất
                    ot_early = 0.0
                    if is_tx and shift.start_time:
                        shift_start_t_local = shift.start_time if isinstance(shift.start_time, time) else parse_time(shift.start_time)
                        shift_start_dt = datetime.combine(dt, shift_start_t_local)
                        effective_in = _round_nearest_30min(check_in_dt)
                        if effective_in < shift_start_dt:
                            early_hours = (shift_start_dt - effective_in).total_seconds() / 3600.0
                            ot_early = _to_half(early_hours)
                            if ot_early < 1.0:
                                ot_early = 0.0

                    ot_hours = ot_checkout + ot_early

                if ot_hours <= 0:
                    continue

            stt += 1
            rows.append({
                "stt": stt,
                "employee_code": emp.employee_code,
                "full_name": emp.full_name,
                "work_date": dt.strftime("%d/%m/%Y"),
                "weekday": dow,
                "shift_code": shift_code,
                "shift_hours": shift_hours_str,
                "check_in":  check_in_dt.strftime("%H:%M")  if check_in_dt  else "",
                "check_out": check_out_dt.strftime("%H:%M") if check_out_dt else "",
                "ot_hours": ot_hours,
                "is_sunday": is_sunday,
            })

    return rows


class OvertimeRow(BaseModel):
    employee_id: int
    employee_code: str
    full_name: str
    department: Optional[str] = None
    default_shift_code: Optional[str] = None
    days: dict  # {1: {shift: "D", ot: 2.0, is_sunday: false, is_holiday: false}, ...}
    total_ot_normal: float  # OT ngay thuong (x1.5)
    total_ot_sunday: float  # OT chu nhat (x2.0)
    total_ot_holiday: float # OT ngay le (x3.0)
    total_ot_hours: float


class OvertimeMonthResponse(BaseModel):
    month_key: str
    days_in_month: int
    weekdays: dict
    rows: List[OvertimeRow]
    summary: dict  # tong hop


@router.get("", response_model=OvertimeMonthResponse)
async def get_overtime(
    month_key: str = Query(..., description="YYYY-MM"),
    ot_style: Optional[str] = Query("old", description="old or new"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Tinh OT theo thang dua tren lich lam + ma ca + thuc te cham cong"""
    try:
        year, month = map(int, month_key.split("-"))
    except ValueError:
        raise HTTPException(400, "month_key phai la YYYY-MM")

    try:
        days_in_month = calendar.monthrange(year, month)[1]
        first_day = date(year, month, 1)
        last_day = date(year, month, days_in_month)

        # Weekday labels + sunday detection
        weekdays = {}
        sundays = set()
        for d in range(1, days_in_month + 1):
            dt = date(year, month, d)
            dow = DOW_VN[dt.weekday()]
            weekdays[d] = dow
            if dow == "CN":
                sundays.add(d)

        # Load all shifts
        shift_result = await db.execute(select(ShiftTemplate))
        shifts_by_id = {s.id: s for s in shift_result.scalars().all()}
        shifts_by_code = {s.code: s for s in shift_result.scalars().all()}

        # Load holidays
        holiday_q = select(CompanyHoliday).where(
            and_(CompanyHoliday.holiday_date >= first_day, CompanyHoliday.holiday_date <= last_day, CompanyHoliday.is_active == True)
        )
        holiday_result = await db.execute(holiday_q)
        holidays_in_range = list(holiday_result.scalars().all())
        holiday_dates = {h.holiday_date for h in holidays_in_range}

        # Load holiday targets (for employee scope)
        from app.models.holiday import HolidayTargetEmployee
        holiday_ids = [h.id for h in holidays_in_range]
        holiday_targets_map = {}
        if holiday_ids:
            target_res = await db.execute(
                select(HolidayTargetEmployee.holiday_id, HolidayTargetEmployee.employee_id)
                .where(HolidayTargetEmployee.holiday_id.in_(holiday_ids))
            )
            for h_id, emp_id in target_res.all():
                if h_id not in holiday_targets_map:
                    holiday_targets_map[h_id] = set()
                holiday_targets_map[h_id].add(emp_id)

        # Load holiday exceptions
        from app.models.holiday import HolidayException
        exc_q = select(HolidayException.employee_id, CompanyHoliday.holiday_date).join(
            CompanyHoliday, HolidayException.holiday_id == CompanyHoliday.id
        ).where(
            and_(
                CompanyHoliday.holiday_date >= first_day,
                CompanyHoliday.holiday_date <= last_day,
                CompanyHoliday.is_active == True
            )
        )
        exc_res = await db.execute(exc_q)
        holiday_exceptions = {(row[0], row[1]) for row in exc_res.all()}

        # Load active employees
        from sqlalchemy import or_
        emp_q = select(Employee).where(
            and_(
                or_(Employee.join_date.is_(None), Employee.join_date <= last_day),
                or_(Employee.leave_date.is_(None), Employee.leave_date >= first_day),
                or_(Employee.is_active == True, Employee.leave_date.is_not(None)),
            )
        )
        if current_user.role == UserRole.WORKER:
            emp_q = emp_q.where(Employee.id == current_user.employee_id)
        emp_q = emp_q.order_by(Employee.employee_code)
        emp_result = await db.execute(emp_q)
        employees = emp_result.scalars().all()

        # Load schedule overrides
        schedule_q = select(WorkSchedule).where(WorkSchedule.month_key == month_key)
        schedule_result = await db.execute(schedule_q)
        override_map = {}
        for ws in schedule_result.scalars().all():
            override_map[(ws.employee_id, ws.work_date.day)] = ws.shift_id

        # Load attendance data
        att_q = select(AttendanceDaily).where(
            and_(AttendanceDaily.work_date >= first_day, AttendanceDaily.work_date <= last_day)
        )
        att_result = await db.execute(att_q)
        att_map = {}
        for a in att_result.scalars().all():
            att_map[(a.employee_id, a.work_date)] = a

        # Load raw logs for NU mode detection
        log_q = select(AttendanceLog).where(
            and_(AttendanceLog.event_time >= datetime.combine(first_day, time(0, 0)), 
                 AttendanceLog.event_time <= datetime.combine(last_day + timedelta(days=1), time(12, 0)))
        )
        log_result = await db.execute(log_q)
        all_logs = log_result.scalars().all()
        
        # Map logs to employee_id
        emp_code_to_id = {str(e.employee_code).lstrip("'"): e.id for e in employees}
        logs_with_id = []
        for l in all_logs:
            eid = emp_code_to_id.get(str(l.employee_code).lstrip("'"))
            if eid:
                l.employee_id = eid
                logs_with_id.append(l)

        # Build NU results
        nu_shift_code_map = {}
        emp_id_list = [e.id for e in employees]
        for emp in employees:
            default_shift = shifts_by_code.get(emp.default_shift_code)
            for d in range(1, days_in_month + 1):
                dt = date(year, month, d)
                override_id = override_map.get((emp.id, d))
                if override_id:
                    s = shifts_by_id.get(override_id)
                    if s: nu_shift_code_map[(emp.id, dt)] = s.code
                elif default_shift and is_nu_dynamic_shift_code(default_shift.code):
                    nu_shift_code_map[(emp.id, dt)] = default_shift.code

        nu_results = build_nu_shift_day_results(
            nu_shift_code_map=nu_shift_code_map,
            employee_id_list=emp_id_list,
            attendance_log_rows=logs_with_id
        )

        # Load X overtime configs cho tháng
        xot_map = {}
        if emp_id_list:
            xot_q = select(XOvertimeConfig).where(
                and_(
                    XOvertimeConfig.work_date >= first_day,
                    XOvertimeConfig.work_date <= last_day,
                    XOvertimeConfig.employee_id.in_(emp_id_list),
                )
            )
            xot_result = await db.execute(xot_q)
            xot_map = {(c.employee_id, c.work_date): c for c in xot_result.scalars().all()}

        # Build OT data
        rows = []
        grand_ot_normal = 0
        grand_ot_sunday = 0
        grand_ot_holiday = 0

        for emp in employees:
            days_data = {}
            emp_ot_normal = 0.0
            emp_ot_sunday = 0.0
            emp_ot_holiday = 0.0

            default_shift = shifts_by_code.get(emp.default_shift_code)

            for d in range(1, days_in_month + 1):
                dt = date(year, month, d)
                is_sunday = dt.weekday() == 6
                if (emp.join_date and dt < emp.join_date) or (emp.leave_date and dt > emp.leave_date):
                    days_data[d] = {
                        "shift": None,
                        "ot": 0.0,
                        "is_sunday": is_sunday,
                        "is_holiday": False
                    }
                    continue
                # Check if there is an active holiday on this date that applies to the employee
                active_holiday = None
                for h in holidays_in_range:
                    if h.holiday_date == dt:
                        is_exception = (emp.id, dt) in holiday_exceptions
                        if not is_exception and check_holiday_applies_to_employee(h, emp.id, emp.department, holiday_targets_map):
                            active_holiday = h
                            break

                # Determine shift
                override_id = override_map.get((emp.id, d))
                if override_id:
                    shift = shifts_by_id.get(override_id)
                elif default_shift and is_nu_dynamic_shift_code(default_shift.code):
                    shift = default_shift
                else:
                    shift = None if is_sunday else default_shift

                # Get attendance
                att = att_map.get((emp.id, dt))
                check_in_dt = att.first_check_in if att else None
                check_out_dt = att.last_check_out if att else None
                shift_code = shift.code if shift else None
                
                nu_res = nu_results.get((emp.id, dt))

                is_holiday = False
                is_half_day_worked = False

                if active_holiday:
                    if active_holiday.duration == "half":
                        has_punches = False
                        if nu_res:
                            has_punches = bool(nu_res.check_in and nu_res.check_out)
                        else:
                            has_punches = bool(check_in_dt and check_out_dt)
                        
                        if has_punches:
                            is_half_day_worked = True
                            is_holiday = False
                        else:
                            is_holiday = True
                            check_in_dt = None
                            check_out_dt = None
                            nu_res = None
                    else:
                        is_holiday = True
                        check_in_dt = None
                        check_out_dt = None
                        nu_res = None

                if nu_res:
                    check_in_dt = nu_res.check_in
                    check_out_dt = nu_res.check_out
                    ev = evaluate_attendance(shift, check_in_dt, check_out_dt, dt, is_sunday, is_holiday, is_night_override=(nu_res.mode=="night"))
                    shift_code = nu_res.shift_code
                    if shift and is_nu_dynamic_shift_code(shift.code) and (shift.default_overtime_hours is not None):
                        ot_hours = float(shift.default_overtime_hours or 0)
                    else:
                        ot_hours = float(nu_res.total_ot_hours)
                else:
                    ev = evaluate_attendance(shift, check_in_dt, check_out_dt, dt, is_sunday, is_holiday)
                    ot_hours = float(ev["ot_hours"])

                    if is_half_day_worked:
                        ot_hours = 0.0

                    if ot_hours <= 0 and shift and is_nu_dynamic_shift_code(shift.code):
                        ot_hours = float(shift.default_overtime_hours or 0)

                if nu_res and shift and is_nu_dynamic_shift_code(shift.code):
                    shift_hours_str = "18:00-06:00" if nu_res.mode == NU_NIGHT_MODE else "06:00-18:00"
                    ot_hours = _nu_display_ot_hours(nu_res.mode, check_in_dt, check_out_dt)
                    if ot_style == "new":
                        xot = xot_map.get((emp.id, dt))
                        ot_hours = float(xot.ot_hours) if (xot and xot.ot_hours is not None) else 0.0

                    if is_holiday:
                        emp_ot_holiday += ot_hours
                    elif is_sunday:
                        emp_ot_sunday += ot_hours
                    else:
                        emp_ot_normal += ot_hours

                    days_data[d] = {
                        "shift": shift_code,
                        "ot": ot_hours,
                        "is_sunday": is_sunday,
                        "is_holiday": is_holiday,
                    }

                    continue

                # Apply new ot_style override if needed
                if ot_style == "new":
                    xot = xot_map.get((emp.id, dt))
                    ot_hours = float(xot.ot_hours) if (xot and xot.ot_hours is not None) else 0.0

                # Categorize OT
                if is_holiday:
                    emp_ot_holiday += ot_hours
                elif is_sunday:
                    emp_ot_sunday += ot_hours
                else:
                    emp_ot_normal += ot_hours

                days_data[d] = {
                    "shift": shift_code,
                    "ot": ot_hours,
                    "is_sunday": is_sunday,
                    "is_holiday": is_holiday
                }

            total_ot = emp_ot_normal + emp_ot_sunday + emp_ot_holiday
            grand_ot_normal += emp_ot_normal
            grand_ot_sunday += emp_ot_sunday
            grand_ot_holiday += emp_ot_holiday

            rows.append(OvertimeRow(
                employee_id=emp.id,
                employee_code=emp.employee_code,
                full_name=emp.full_name,
                department=emp.department,
                default_shift_code=emp.default_shift_code,
                days=days_data,
                total_ot_normal=emp_ot_normal,
                total_ot_sunday=emp_ot_sunday,
                total_ot_holiday=emp_ot_holiday,
                total_ot_hours=total_ot,
            ))

        return OvertimeMonthResponse(
            month_key=month_key,
            days_in_month=days_in_month,
            weekdays=weekdays,
            rows=rows,
            summary={
                "total_employees": len(rows),
                "employees_with_ot": len([r for r in rows if r.total_ot_hours > 0]),
                "total_ot_normal": grand_ot_normal,
                "total_ot_sunday": grand_ot_sunday,
                "total_ot_holiday": grand_ot_holiday,
                "total_ot_hours": grand_ot_normal + grand_ot_sunday + grand_ot_holiday,
            },
        )
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}")


# ── Actual OT endpoints ───────────────────────────────────────────────────────

@router.get("/actual-ot")
async def get_actual_ot(
    month_key: str = Query(..., description="YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """OT thực tế dựa trên giờ chấm công so với ca quy định (dành cho kế toán)."""
    try:
        rows = await _compute_actual_ot(month_key, db, current_user)
        return {"month_key": month_key, "total": len(rows), "rows": rows}
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(500, f"Internal Server Error: {str(e)}")


@router.get("/actual-ot/export")
async def export_actual_ot(
    month_key: str = Query(..., description="YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Xuất Excel OT thực tế — nhóm theo nhân viên, có dòng tổng từng người."""
    try:
        rows = await _compute_actual_ot(month_key, db, current_user)

        from decimal import Decimal
        # Map employee code to ID
        emp_r = await db.execute(select(Employee))
        emp_map = {str(e.employee_code).lstrip("'"): e.id for e in emp_r.scalars().all()}

        # Load X overtime configs cho tháng
        year, month = map(int, month_key.split("-"))
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        xot_q = select(XOvertimeConfig).where(
            and_(
                XOvertimeConfig.work_date >= first_day,
                XOvertimeConfig.work_date <= last_day,
            )
        )
        xot_result = await db.execute(xot_q)
        xot_map = {(c.employee_id, c.work_date): c for c in xot_result.scalars().all()}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"OT Thuc Te {month_key}"

        # ── Styles ─────────────────────────────────────────────────────────
        hdr_fill  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sun_fill  = PatternFill(start_color="FFDCE0", end_color="FFDCE0", fill_type="solid")  # đỏ nhạt CN
        sun_font  = Font(bold=True, color="C0392B")

        sub_fill  = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # vàng tổng
        sub_font  = Font(bold=True, color="856404")

        center = Alignment(horizontal="center", vertical="center")
        left   = Alignment(horizontal="left",   vertical="center")

        # ── Header ─────────────────────────────────────────────────────────
        headers = ["STT", "Mã NV", "Họ tên", "Ngày", "Thứ", "Giờ ca", "Giờ vào", "Giờ ra", "OT (giờ)", "Chấp Nhận"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = hdr_align
        ws.row_dimensions[1].height = 24

        for i, w in enumerate([6, 9, 28, 11, 6, 16, 10, 10, 10, 12], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        # ── Group by employee, write rows + subtotal ────────────────────────
        from itertools import groupby
        excel_row = 2

        for emp_code, grp in groupby(rows, key=lambda r: r["employee_code"]):
            grp_list = list(grp)
            emp_name = grp_list[0]["full_name"]
            total_ot = 0.0

            for row in grp_list:
                is_sun = row["is_sunday"]
                # Format "HH:MM - HH:MM" with spaces for readability
                shift_disp = row["shift_hours"].replace("-", " - ") if "-" in row["shift_hours"] else row["shift_hours"]
                
                # Check current approved config
                work_date_obj = datetime.strptime(row["work_date"], "%d/%m/%Y").date()
                eid = emp_map.get(str(row["employee_code"]).lstrip("'"))
                xot = xot_map.get((eid, work_date_obj)) if eid else None

                calc_ot_hours = float(row["ot_hours"] or 0)
                display_ot_hours = calc_ot_hours
                approve_val = ""

                if xot and xot.ot_hours is not None and float(xot.ot_hours) > 0:
                    approved_ot = float(xot.ot_hours)
                    display_ot_hours = approved_ot
                    if Decimal(str(approved_ot)) == Decimal(str(calc_ot_hours)):
                        approve_val = "x"
                    else:
                        approve_val = approved_ot

                total_ot += display_ot_hours

                vals = [
                    row["stt"], row["employee_code"], row["full_name"],
                    row["work_date"], row["weekday"], shift_disp,
                    row["check_in"], row["check_out"], display_ot_hours,
                    approve_val
                ]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=excel_row, column=col, value=val)
                    c.alignment = left if col == 3 else center
                    if is_sun:
                        c.fill = sun_fill
                        c.font = sun_font
                excel_row += 1

            # Subtotal row
            c_code = ws.cell(row=excel_row, column=2, value=emp_code)
            c_name = ws.cell(row=excel_row, column=3, value=f"TỔNG - {emp_name}")
            c_ot   = ws.cell(row=excel_row, column=9, value=total_ot)
            for c in (c_code, c_name, c_ot):
                c.fill = sub_fill
                c.font = sub_font
            c_name.alignment = left
            c_code.alignment = center
            c_ot.alignment   = center
            # Fill remaining cells of subtotal row with same bg
            for col in [1, 4, 5, 6, 7, 8, 10]:
                ws.cell(row=excel_row, column=col).fill = sub_fill
            excel_row += 1
            excel_row += 1  # dòng trống ngăn cách giữa các nhân viên

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"OT_thuc_te_{month_key}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(500, f"Internal Server Error: {str(e)}")

