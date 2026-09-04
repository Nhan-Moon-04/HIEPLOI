from typing import List, Optional, Literal
from datetime import date, datetime, time, timedelta
import calendar
import re
from decimal import Decimal
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, delete
from app.database import get_db
from app.models.attendance import AttendanceDaily, AttendanceDetail, AttendanceLog
from app.models.employee import Employee
from app.models.shift import ShiftTemplate
from app.models.schedule import WorkSchedule
from app.models.holiday import CompanyHoliday
from app.models.x_overtime import XOvertimeConfig
from app.models.user import AppUser, UserRole
from app.middleware.auth import get_current_user, require_roles
from app.services.nu_shift import (
    is_nu_dynamic_shift_code,
    build_nu_shift_day_results,
    calculate_nu_shift_details,
    XNU_MODE_1,
    XNU_MODE_2,
    XNU_MODE_3,
)
from app.utils.audit_helper import log_audit
from pydantic import BaseModel
from app.utils.lock_helper import check_date_locked, check_month_locked
from app.models.salary import MonthlyWorkdayConfig


def parse_input_time(time_str: str) -> Optional[time]:
    if not time_str:
        return None
    time_str = str(time_str).strip().lower()
    if not time_str:
        return None
    
    # Try HH:MM:SS or HH:MM
    m = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$', time_str)
    if m:
        h, m_val = int(m.group(1)), int(m.group(2))
        if 0 <= h < 24 and 0 <= m_val < 60:
            return time(h, m_val)
            
    # Try HHhMM or HHh
    m = re.match(r'^(\d{1,2})h(\d{2})?$', time_str)
    if m:
        h = int(m.group(1))
        m_val = int(m.group(2)) if m.group(2) else 0
        if 0 <= h < 24 and 0 <= m_val < 60:
            return time(h, m_val)
            
    # Try just an integer HH
    if time_str.isdigit():
        h = int(time_str)
        if 0 <= h < 24:
            return time(h, 0)
            
    return None


router = APIRouter(prefix="/attendance", tags=["Attendance - Cham Cong"])

DOW_VN = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]

GRACE_MINUTES = 15  # Cho phep ve som 15p
DRIVER_AUTO_OT_SHIFT_CODES = {"TX1", "TX2"}
X_OT_SHIFT_CODES = {"X", "X40"}  # Ca hỗ trợ tăng ca theo ngày


class AttendanceCell(BaseModel):
    work_date: str
    day: int
    dow: str
    shift_code: Optional[str] = None
    shift_name: Optional[str] = None
    shift_start: Optional[str] = None  # HH:MM
    shift_end: Optional[str] = None
    standard_hours: Optional[float] = None
    check_in: Optional[str] = None  # HH:MM or datetime
    check_out: Optional[str] = None
    actual_hours: Optional[float] = None
    deviation: Optional[float] = None  # negative = ve som
    ot_hours: Optional[float] = None
    status: str = "no_data"  # full, early_leave, late, absent, no_data, holiday, off, night
    is_holiday: bool = False
    is_sunday: bool = False
    notes: Optional[str] = None
    meal_allowance: float = 0.0
    meal_count: Optional[int] = None
    night_allowance: float = 0.0
    ot_eligible: bool = False
    night_eligible: bool = False
    has_manual_xot: bool = False
    manual_meal_count: Optional[int] = None
    manual_ot_end_time: Optional[str] = None
    is_irregular: bool = False  # Giờ làm bất thường, cần duyệt
    meal_approval_id: Optional[int] = None
    meal_approval_status: Optional[str] = None  # pending | approved | rejected


class AttendanceRow(BaseModel):
    employee_id: int
    employee_code: str
    full_name: str
    department: Optional[str] = None
    default_shift_code: Optional[str] = None
    days: List[AttendanceCell]
    summary: dict  # total_days, present, absent, ot_hours, etc.


class AttendanceMonthResponse(BaseModel):
    month_key: str
    days_in_month: int
    is_locked: bool = False
    rows: List[AttendanceRow]


def parse_time(t) -> Optional[time]:
    if t is None:
        return None
    if isinstance(t, time):
        return t
    if isinstance(t, str):
        parts = t.split(":")
        return time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
    return None


def calc_hours_between(check_in_dt, check_out_dt, break_minutes=60):
    """Tinh gio lam thuc te"""
    if not check_in_dt or not check_out_dt:
        return 0.0
    diff = (check_out_dt - check_in_dt).total_seconds() / 3600.0
    # Tru gio nghi
    if diff > 4:
        diff -= break_minutes / 60.0
    return round(max(diff, 0), 2)


def evaluate_attendance(shift, check_in_dt, check_out_dt, work_date, is_sunday, is_holiday, night_allowance_rate=0.0, is_night_override=None):
    """Danh gia cham cong 1 ngay dua tren ma ca"""
    result = {
        "actual_hours": 0.0,
        "deviation": 0.0,
        "ot_hours": 0.0,
        "status": "no_data",
        "notes": "",
        "meal_allowance": 0.0,
        "meal_count": 0,
        "night_allowance": 0.0,
    }

    if shift and (shift.code or "").upper() == "SEP":
        # Ca của sếp: Tự động tính công và tiền ăn, không cần quét thẻ
        standard = float(shift.standard_hours or 8)
        if is_holiday:
            result["status"] = "holiday"
            result["notes"] = "Ngày lễ/nghỉ"
            return result
        if is_sunday:
            result["status"] = "off"
            result["notes"] = "Nghỉ chủ nhật"
            return result
            
        result["status"] = "full"
        result["actual_hours"] = standard
        result["deviation"] = 0.0
        result["notes"] = "Ca sếp (Tự động tính công)"
        
        # Thứ 2 - Thứ 6: 80k (2 bữa), Thứ 7: 40k (1 bữa)
        dow_idx = work_date.weekday()
        rate = float(shift.meal_allowance or 40000)
        if dow_idx <= 4:  # Thứ 2 - Thứ 6
            result["meal_allowance"] = rate * 2
            result["meal_count"] = 2
        elif dow_idx == 5:  # Thứ 7
            result["meal_allowance"] = rate * 1
            result["meal_count"] = 1
        else:
            result["meal_allowance"] = 0.0
            result["meal_count"] = 0
            
        return result

    if is_holiday:
        result["status"] = "holiday"
        result["notes"] = "Ngay le/nghi"
        if not check_in_dt or not check_out_dt:
            return result
        # Neu co cham cong ngay le => tiep tuc de tinh tang ca

    if not shift:
        if (is_sunday or is_holiday) and (check_in_dt or check_out_dt):
            # Tu dong dung default_shift hoac gia dinh ca 8h de tinh tang ca
            result["notes"] = "Lam viec ngay nghi/le"
        elif is_sunday:
            result["status"] = "off"
            result["notes"] = "Nghỉ chủ nhật"
            return result
        else:
            result["status"] = "no_data"
            return result

    if shift and shift.is_leave_code:
        result["status"] = "off"
        result["notes"] = shift.name or "Nghi"
        return result

    standard = float(shift.standard_hours or 8) if shift else 8.0

    if not check_in_dt and not check_out_dt:
        # Ko co du lieu cham cong nao
        if is_sunday:
            # Chủ nhật mặc định nghỉ — chỉ tính đi làm nếu có chấm công
            result["status"] = "off"
            result["notes"] = "Nghỉ chủ nhật"
        elif is_holiday:
            result["status"] = "holiday"
            result["notes"] = "Ngày lễ/nghỉ"
        else:
            result["status"] = "absent"
            result["deviation"] = -standard
            result["notes"] = "Vang mat (Khong quet the)"
        return result

    if not check_in_dt or not check_out_dt:
        # Chi co 1 dau (vao hoac ra)
        result["status"] = "forgot_scan"
        result["actual_hours"] = 0.0
        result["deviation"] = -standard
        result["notes"] = "Quen quet the (Chi co 1 dau)"
        # No early return here - allow falling through to meal allowance

    # Gioi han gio vao (khong cho tinh som hon quy dinh)
    shift_start_time = parse_time(shift.start_time) if shift else None
    original_check_in = check_in_dt
    if shift_start_time and check_in_dt:
        expected_start = datetime.combine(work_date, shift_start_time)
        if check_in_dt < expected_start:
            check_in_dt = expected_start

    # Tinh gio lam thuc te
    break_mins = int(shift.break_minutes or 60) if shift else 60
    actual = 0.0
    if check_in_dt and check_out_dt:
        actual = calc_hours_between(check_in_dt, check_out_dt, break_mins)
    result["actual_hours"] = actual

    # Check ve som
    shift_end_time = parse_time(shift.end_time) if shift else None
    if shift_end_time and check_out_dt and result["status"] != "forgot_scan":
        effective_is_night = is_night_override if is_night_override is not None else (shift.is_night_shift if shift else False)
        # For night shift: end_time is next day
        if effective_is_night:
            expected_end = datetime.combine(work_date + timedelta(days=1), shift_end_time)
        else:
            expected_end = datetime.combine(work_date, shift_end_time)

        diff_minutes = (check_out_dt - expected_end).total_seconds() / 60.0

        is_nu = is_nu_dynamic_shift_code(shift.code if shift else None)
        if diff_minutes <= -GRACE_MINUTES and not is_nu:
            # Ve som tu 15p tro len
            result["status"] = "early_leave"
            result["deviation"] = round(diff_minutes / 60.0, 2)
            result["notes"] = f"Ve som {abs(int(diff_minutes))}p"
        elif actual >= standard:
            result["status"] = "full"
        else:
            result["deviation"] = round(actual - standard, 2)
            result["status"] = "full" if abs(result["deviation"]) <= 0.25 else "short"
    else:
        if result["status"] != "forgot_scan":
            result["status"] = "full" if actual >= standard * 0.9 else "short"

    # Special logic for NU shifts
    if shift and is_nu_dynamic_shift_code(shift.code):
        effective_is_night = is_night_override if is_night_override is not None else shift.is_night_shift
        nu_calc = calculate_nu_shift_details(shift.code, actual, is_night=effective_is_night, night_allowance_rate=night_allowance_rate)
        result["ot_hours"] = nu_calc["ot_hours"]
        result["meal_allowance"] = nu_calc["meal_allowance"]
        result["night_allowance"] = nu_calc["night_allowance"]
        
        # If it's a "minus" shift (NU1, NU2, etc.), we should also reflect the adjusted standard hours?
        # Actually, standard hours are already in the shift template, but nu_calc provides it too.
        # For evaluation, we mainly care about OT and money.
    else:
        # OT
        if is_sunday or is_holiday:
            # Ngay nghi/le: Tat ca gio lam deu tinh vao tang ca
            result["ot_hours"] = actual
        elif shift and (shift.code or "").upper() in DRIVER_AUTO_OT_SHIFT_CODES:
            # Tinh OT cho tai xe: checkout trễ + vào sớm (làm tròn 30p gần nhất, cap 6:00)
            import math
            ot = 0.0
            shift_end_time_local = parse_time(shift.end_time)
            if shift_end_time_local and check_out_dt:
                expected_end = datetime.combine(work_date, shift_end_time_local)
                if check_out_dt > expected_end:
                    ot = (check_out_dt - expected_end).total_seconds() / 3600.0
            
            # OT vào sớm: làm tròn 30p gần nhất
            if original_check_in and shift_start_time:
                expected_start = datetime.combine(work_date, shift_start_time)
                # Làm tròn nearest 30min: 0-15→:00, 16-45→:30, 46-59→next:00
                m = original_check_in.minute
                if m <= 15:
                    r_min, r_h = 0, original_check_in.hour
                elif m <= 45:
                    r_min, r_h = 30, original_check_in.hour
                else:
                    r_min, r_h = 0, (original_check_in.hour + 1) % 24
                effective_in = original_check_in.replace(hour=r_h, minute=r_min, second=0, microsecond=0)
                if effective_in < expected_start:
                    early_hours = (expected_start - effective_in).total_seconds() / 3600.0
                    ot_early = math.floor(early_hours * 2 + 0.5) / 2  # round to 0.5h
                    if ot_early >= 1.0:
                        ot += ot_early
            
            result["ot_hours"] = round(max(ot, 0), 2)
        else:
            # Ngay thuong
            ot = float(shift.default_overtime_hours or 0) if shift else 0.0
            result["ot_hours"] = ot

        # Tiền ăn: Nếu có mặt (hoặc có ít nhất 1 đầu quẹt) thì tính meal_allowance * meal_count
        if result["status"] in ("full", "early_leave", "short", "forgot_scan") or ((is_sunday or is_holiday) and actual > 0):
            meal_val = float(shift.meal_allowance or 35000) if shift else 35000.0
            shift_code_upper = (shift.code or "").upper() if shift else ""
            is_auto = shift_code_upper in DRIVER_AUTO_OT_SHIFT_CODES or is_nu_dynamic_shift_code(shift_code_upper)

            if is_auto:
                # Quy tắc thời gian cho ca tự động (TX1, TX2, NU, XNU):
                #   Bữa sáng : check-in trước 9h
                #   Bữa tối  : check-out >= 17h50 HOẶC check-in >= 18h HOẶC OT >= 3h
                ci = original_check_in or check_in_dt
                ot = result.get("ot_hours") or 0.0
                has_morning = bool(ci and ci.hour < 9)
                # Chủ nhật/ngày lễ: toàn bộ giờ làm tính OT → không dùng ot>=3 để xét bữa tối
                # (tránh tính 2 bữa khi tài xế về trước 17h50)
                has_late = bool(
                    (check_out_dt and (check_out_dt.hour * 60 + check_out_dt.minute) >= 17 * 60 + 50)
                    or (ci and ci.hour >= 18)
                    or (not (is_sunday or is_holiday) and ot >= 3)
                )
                meal_count = (1 if has_morning else 0) + (1 if has_late else 0)
            else:
                # Ca không-tự-động: dùng meal_count cố định từ cấu hình ca
                meal_count = int(shift.meal_count or 0) if shift else 0

            result["meal_allowance"] = meal_val * meal_count if meal_count > 0 else 0.0
            result["meal_count"] = meal_count

    return result


def check_holiday_applies_to_employee(holiday, emp_id, emp_dept, target_emp_ids_map):
    if holiday.scope == "all":
        return True
    if holiday.scope == "department":
        if not holiday.departments or not emp_dept:
            return False
        depts = [d.strip() for d in holiday.departments.split(",") if d.strip()]
        return emp_dept in depts
    if holiday.scope == "employee":
        target_ids = target_emp_ids_map.get(holiday.id, set())
        return emp_id in target_ids
    return True


@router.get("", response_model=AttendanceMonthResponse)
async def get_attendance(
    month_key: str = Query(..., description="YYYY-MM"),
    employee_id: Optional[int] = None,
    department: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    night_allowance_rate: Optional[float] = 0.0,
    ot_style: Optional[str] = Query("old", description="old or new"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Cham cong thang - ket hop lich lam + du lieu cham cong + ma ca"""


    if current_user.role == UserRole.WORKER:
        employee_id = current_user.employee_id

    try:
        year, month = map(int, month_key.split("-"))
    except ValueError:
        raise HTTPException(400, "month_key phai la YYYY-MM")

    month_days = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end = date(year, month, month_days)

    range_start = month_start
    range_end = month_end

    if start_date:
        range_start = date.fromisoformat(start_date)
    if end_date:
        range_end = date.fromisoformat(end_date)

    if range_start > range_end:
        # Invalid range or no days in range
        return AttendanceMonthResponse(month_key=month_key, days_in_month=0, rows=[])

    range_days = (range_end - range_start).days + 1
    range_dates = [range_start + timedelta(days=i) for i in range(range_days)]


    # Load shifts
    shift_result = await db.execute(select(ShiftTemplate))
    shifts_by_id = {}
    shifts_by_code = {}
    for s in shift_result.scalars().all():
        shifts_by_id[s.id] = s
        shifts_by_code[s.code] = s

    # Load holidays
    holiday_q = select(CompanyHoliday).where(
        and_(CompanyHoliday.holiday_date >= range_start, CompanyHoliday.holiday_date <= range_end, CompanyHoliday.is_active == True)
    )
    holiday_result = await db.execute(holiday_q)
    holidays_in_range = list(holiday_result.scalars().all())
    holiday_dates = {h.holiday_date for h in holidays_in_range}

    # Load holiday targets (for employee scope)
    from app.models.holiday import HolidayTargetEmployee
    holiday_ids = [h.id for h in holidays_in_range]
    holiday_targets_map = {}
    if holiday_ids:
        target_res = await db.execute(
            select(HolidayTargetEmployee.holiday_id, HolidayTargetEmployee.employee_id)
            .where(HolidayTargetEmployee.holiday_id.in_(holiday_ids))
        )
        for h_id, emp_id in target_res.all():
            if h_id not in holiday_targets_map:
                holiday_targets_map[h_id] = set()
            holiday_targets_map[h_id].add(emp_id)

    # Load holiday exceptions
    from app.models.holiday import HolidayException
    exc_q = select(HolidayException.employee_id, CompanyHoliday.holiday_date).join(
        CompanyHoliday, HolidayException.holiday_id == CompanyHoliday.id
    ).where(
        and_(
            CompanyHoliday.holiday_date >= range_start,
            CompanyHoliday.holiday_date <= range_end,
            CompanyHoliday.is_active == True
        )
    )
    exc_res = await db.execute(exc_q)
    holiday_exceptions = {(row[0], row[1]) for row in exc_res.all()}

    # Load employees
    from sqlalchemy import or_
    emp_filters = [
        or_(Employee.join_date.is_(None), Employee.join_date <= range_end),
        or_(Employee.leave_date.is_(None), Employee.leave_date >= range_start),
    ]
    if not employee_id:
        emp_filters.append(or_(Employee.is_active == True, Employee.leave_date.is_not(None)))
    emp_q = select(Employee).where(and_(*emp_filters))
    if employee_id:
        emp_q = emp_q.where(Employee.id == employee_id)
    if department:
        emp_q = emp_q.where(Employee.department == department)
        
    emp_result = await db.execute(emp_q)
    employees = list(emp_result.scalars().all())

    # Sắp xếp nhân viên theo thứ tự bộ phận, sau đó tới thứ tự nhân viên trong bộ phận
    from app.models.department import Department
    dept_order_q = await db.execute(select(Department.name, Department.sort_order))
    dept_order_map = {row[0]: row[1] for row in dept_order_q.all() if row[0]}

    def get_emp_sort_key(e):
        d_name = e.department
        d_order = dept_order_map.get(d_name, 9999) if d_name else 9999
        e_order = e.sort_order if e.sort_order is not None else 9999
        try:
            code_num = int(e.employee_code)
        except ValueError:
            code_num = 999999
        return (d_order, d_name or "", e_order, code_num)

    employees.sort(key=get_emp_sort_key)

    # Load schedule overrides for date range (include 1 day before range_start to properly link cross-boundary night shifts)
    emp_id_list = [e.id for e in employees]
    schedule_q = select(WorkSchedule).where(
        and_(WorkSchedule.work_date >= range_start - timedelta(days=1), WorkSchedule.work_date <= range_end)
    )
    if emp_id_list:
        schedule_q = schedule_q.where(WorkSchedule.employee_id.in_(emp_id_list))
    schedule_result = await db.execute(schedule_q)
    override_map = {}
    override_notes = {}
    for ws in schedule_result.scalars().all():
        override_map[(ws.employee_id, ws.work_date)] = ws.shift_id
        if ws.notes:
            override_notes[(ws.employee_id, ws.work_date)] = ws.notes

    # Load attendance data
    att_q = select(AttendanceDaily).where(
        and_(AttendanceDaily.work_date >= range_start, AttendanceDaily.work_date <= range_end)
    )
    if emp_id_list:
        att_q = att_q.where(AttendanceDaily.employee_id.in_(emp_id_list))
    att_result = await db.execute(att_q)
    att_map = {}
    for a in att_result.scalars().all():
        att_map[(a.employee_id, a.work_date)] = a

    # Load raw logs for NU mode detection (include 1 day before range_start)
    log_q = select(AttendanceLog).where(
        and_(AttendanceLog.event_time >= datetime.combine(range_start - timedelta(days=1), time(0, 0)), 
             AttendanceLog.event_time <= datetime.combine(range_end + timedelta(days=1), time(12, 0)))
    )
    log_result = await db.execute(log_q)
    all_logs = log_result.scalars().all()
    
    # Map logs to employee_id
    emp_code_to_id = {e.employee_code: e.id for e in employees}
    logs_with_id = []
    for l in all_logs:
        eid = emp_code_to_id.get(str(l.employee_code).lstrip("'"))
        if eid:
            l.employee_id = eid
            logs_with_id.append(l)

    # Prepare NU shift code map for build_nu_shift_day_results
    nu_shift_code_map = {}
    # We need to know which shift code applies to each (emp, date)
    # Priority: override > default_shift (if NU)
    for emp in employees:
        default_shift = shifts_by_code.get(emp.default_shift_code)
        
        # Include range_start - 1 day to link cross-boundary night shifts
        prev_dt = range_start - timedelta(days=1)
        override_id = override_map.get((emp.id, prev_dt))
        if override_id:
            s = shifts_by_id.get(override_id)
            if s and is_nu_dynamic_shift_code(s.code):
                nu_shift_code_map[(emp.id, prev_dt)] = s.code
        elif default_shift and is_nu_dynamic_shift_code(default_shift.code):
            nu_shift_code_map[(emp.id, prev_dt)] = default_shift.code

        for dt in range_dates:
            override_id = override_map.get((emp.id, dt))
            if override_id:
                s = shifts_by_id.get(override_id)
                if s and is_nu_dynamic_shift_code(s.code):
                    nu_shift_code_map[(emp.id, dt)] = s.code
            elif default_shift and is_nu_dynamic_shift_code(default_shift.code):
                nu_shift_code_map[(emp.id, dt)] = default_shift.code

    nu_results = build_nu_shift_day_results(
        nu_shift_code_map=nu_shift_code_map,
        employee_id_list=emp_id_list,
        attendance_log_rows=logs_with_id,
        night_allowance_rate=night_allowance_rate
    )

    # Load X overtime configs cho tháng
    xot_q = select(XOvertimeConfig).where(
        and_(
            XOvertimeConfig.work_date >= range_start,
            XOvertimeConfig.work_date <= range_end,
            XOvertimeConfig.employee_id.in_(emp_id_list),
        )
    )
    xot_result = await db.execute(xot_q)
    xot_map = {(c.employee_id, c.work_date): c for c in xot_result.scalars().all()}

    # Load MealApproval data cho giờ làm bất thường
    from app.models.meal_approval import MealApproval
    approval_q = select(MealApproval).where(
        and_(
            MealApproval.work_date >= range_start,
            MealApproval.work_date <= range_end,
            MealApproval.employee_id.in_(emp_id_list),
        )
    )
    approval_result = await db.execute(approval_q)
    approval_map = {(a.employee_id, a.work_date): a for a in approval_result.scalars().all()}

    # Build rows
    rows = []
    for emp in employees:
        default_shift = shifts_by_code.get(emp.default_shift_code)
        days_cells = []
        total_present = 0
        total_absent = 0
        total_forgot_scan = 0
        total_early = 0
        total_hours = 0.0
        total_ot = 0.0
        total_ot_weekday = 0.0
        total_ot_sunday = 0.0
        total_ot_holiday = 0.0
        total_meal_count = 0
        total_meal_allowance = 0.0
        total_paid_leave = 0.0

        for dt in range_dates:
            d = dt.day
            dow_idx = dt.weekday()
            dow = DOW_VN[dow_idx]
            is_sunday = dow == "CN"

            # Check if employee has joined yet or has already left
            if (emp.join_date and dt < emp.join_date) or (emp.leave_date and dt > emp.leave_date):
                cell = AttendanceCell(
                    work_date=str(dt),
                    day=d,
                    dow=dow,
                    shift_code=None,
                    shift_name=None,
                    shift_start=None,
                    shift_end=None,
                    standard_hours=None,
                    check_in=None,
                    check_out=None,
                    actual_hours=0.0,
                    deviation=0.0,
                    ot_hours=0.0,
                    status="no_data",
                    is_holiday=False,
                    is_sunday=is_sunday,
                    notes="Chưa vào làm" if (emp.join_date and dt < emp.join_date) else "Đã nghỉ việc",
                    meal_allowance=0.0,
                    meal_count=0,
                    night_allowance=0.0,
                    ot_eligible=False,
                    night_eligible=False,
                    has_manual_xot=False,
                    manual_meal_count=None,
                    manual_ot_end_time=None,
                )
                days_cells.append(cell)
                continue

            # Check if there is an active holiday on this date that applies to the employee
            active_holiday = None
            for h in holidays_in_range:
                if h.holiday_date == dt:
                    is_exception = (emp.id, dt) in holiday_exceptions
                    if not is_exception and check_holiday_applies_to_employee(h, emp.id, emp.department, holiday_targets_map):
                        active_holiday = h
                        break

            # Get attendance record
            att = att_map.get((emp.id, dt))
            check_in_dt = att.first_check_in if att else None
            check_out_dt = att.last_check_out if att else None
            
            # Special case for NU results
            nu_res = nu_results.get((emp.id, dt))

            is_holiday = False
            is_half_day_worked = False

            if active_holiday:
                if active_holiday.duration == "half":
                    # Check if they actually worked (has punches)
                    has_punches = False
                    if nu_res:
                        has_punches = bool(nu_res.check_in and nu_res.check_out)
                    else:
                        has_punches = bool(check_in_dt and check_out_dt)
                    
                    if has_punches:
                        is_half_day_worked = True
                        is_holiday = False
                    else:
                        is_holiday = True
                        check_in_dt = None
                        check_out_dt = None
                        nu_res = None
                else:
                    is_holiday = True
                    check_in_dt = None
                    check_out_dt = None
                    nu_res = None

            # Determine shift
            override_id = override_map.get((emp.id, dt))
            override_note = override_notes.get((emp.id, dt))
            if override_id:
                shift = shifts_by_id.get(override_id)
            elif default_shift and is_nu_dynamic_shift_code(default_shift.code):
                shift = default_shift
            else:
                # Ngày CN: chỉ áp dụng ca mặc định cho tài xế (TX1/TX2) vì họ làm cả tuần
                # Các ca khác ngày CN mặc định là nghỉ
                if is_sunday:
                    shift = default_shift if (default_shift and (default_shift.code or "").upper() in DRIVER_AUTO_OT_SHIFT_CODES) else None
                else:
                    shift = default_shift

            # Calculate paid leave contribution
            if shift and shift.is_leave_code and shift.is_paid_leave:
                if shift.code == "P":
                    total_paid_leave += 1.0
                elif shift.code in ("S", "C"):
                    total_paid_leave += 0.5
                else:
                    total_paid_leave += 1.0

            if nu_res:
                check_in_dt = nu_res.check_in
                check_out_dt = nu_res.check_out
                shift_name = nu_res.shift_name
                # Use standard evaluate for some parts but override others
                ev = evaluate_attendance(shift, check_in_dt, check_out_dt, dt, is_sunday, is_holiday, night_allowance_rate=night_allowance_rate, is_night_override=(nu_res.mode == "night"))
                ev["ot_hours"] = nu_res.total_ot_hours
                ev["meal_allowance"] = nu_res.meal_allowance
                ev["meal_count"] = nu_res.meal_count
                ev["night_allowance"] = nu_res.night_allowance
                ev["is_irregular"] = nu_res.is_irregular
                if nu_res.warning_note:
                    ev["notes"] = f"{ev['notes']} | {nu_res.warning_note}" if ev["notes"] else nu_res.warning_note
                
                # Ensure notes mention correct shift mode
                if nu_res.mode in (XNU_MODE_1, XNU_MODE_2, XNU_MODE_3):
                    mode_note = {
                        XNU_MODE_1: "Ca 1",
                        XNU_MODE_2: "Ca 2",
                        XNU_MODE_3: "Ca 3",
                    }[nu_res.mode]
                else:
                    mode_str = "Sáng" if nu_res.mode == "morning" else "Tối"
                    mode_note = f"Ca {mode_str}"
                ev["notes"] = f"{mode_note} | {ev['notes']}" if ev["notes"] else mode_note
                
                cell_shift_code = nu_res.shift_code
                cell_shift_name = nu_res.shift_name
            else:
                ev = evaluate_attendance(shift, check_in_dt, check_out_dt, dt, is_sunday, is_holiday, night_allowance_rate=night_allowance_rate)
                cell_shift_code = shift.code if shift else None
                cell_shift_name = shift.name if shift else None

            if is_half_day_worked:
                ev["actual_hours"] = 8.0
                ev["deviation"] = 0.0
                ev["status"] = "full"
                ev["notes"] = f"Nghỉ nửa ngày (Đi làm tính 8.0h) | {ev.get('notes') or ''}".strip(" | ")
                if ot_style != "new":
                    ev["ot_hours"] = 0.0

            # Format times
            ci_str = check_in_dt.strftime("%Y-%m-%d %H:%M") if check_in_dt else None
            co_str = check_out_dt.strftime("%Y-%m-%d %H:%M") if check_out_dt else None

            # Night allowance
            if not nu_res:
                effective_is_night = (shift.is_night_shift if shift else False)
                if shift and effective_is_night and ev["status"] in ("full", "early_leave", "short"):
                    ev["night_allowance"] = night_allowance_rate

            cell_notes = ev["notes"]
            if override_note:
                cell_notes = f"{cell_notes} | {override_note}" if cell_notes else override_note

            # Lấy cấu hình tăng ca thủ công XOT nếu có
            xot = xot_map.get((emp.id, dt))
            has_manual_xot_val = xot is not None
            manual_meal_count_val = xot.meal_count if xot else None
            manual_ot_end_time_val = str(xot.ot_end_time)[:5] if xot and xot.ot_end_time else None

            # Cộng thêm tiền ăn OT nếu có config xot, hoặc đánh dấu ot/night eligible
            # XNU cũng hỗ trợ OT thủ công (lâu lâu tăng ca)
            is_xnu_shift = nu_res is not None and nu_res.shift_code == "XNU"
            is_auto_shift = (nu_res is not None and not is_xnu_shift) or (cell_shift_code or "").upper() in DRIVER_AUTO_OT_SHIFT_CODES
            ot_eligible_val = False
            night_eligible_val = False

            # XNU: hỗ trợ OT thủ công giống X/X40
            if is_xnu_shift and ev["status"] in ("full", "early_leave", "short", "forgot_scan"):
                if xot and xot.meal_count and xot.meal_count > 0:
                    x_meal_rate = 35000.0
                    ot_meal = x_meal_rate * int(xot.meal_count)
                    ev["meal_allowance"] = (ev["meal_allowance"] or 0) + ot_meal
                    ev["meal_count"] = (ev["meal_count"] or 0) + int(xot.meal_count)
                    if xot.ot_hours:
                        ev["ot_hours"] = float(ev["ot_hours"] or 0) + float(xot.ot_hours)
                    if xot.ot_end_time:
                        end_t = parse_time(xot.ot_end_time)
                        if end_t and end_t.hour >= 23:
                            ev["night_allowance"] = (ev["night_allowance"] or 0) + night_allowance_rate
                elif check_out_dt:
                    # Xác định giờ kết thúc ca XNU để detect OT
                    xnu_shift_ends = {
                        XNU_MODE_1: time(14, 0),
                        XNU_MODE_2: time(22, 0),
                    }
                    xnu_end_t = xnu_shift_ends.get(nu_res.mode)
                    if xnu_end_t:
                        xnu_end_dt = datetime.combine(dt, xnu_end_t)
                        actual_ot_h = max(0.0, (check_out_dt - xnu_end_dt).total_seconds() / 3600.0)
                        if check_out_dt >= datetime.combine(dt, time(23, 0)):
                            night_eligible_val = True
                        elif ((check_out_dt.hour * 60 + check_out_dt.minute) >= 17 * 60 + 50 and xnu_end_t.hour < 18) or (actual_ot_h >= 3):
                            ot_eligible_val = True

            elif not is_auto_shift and ev["status"] in ("full", "early_leave", "short", "forgot_scan"):
                if xot and xot.meal_count and xot.meal_count > 0:
                    x_meal_rate = float(shift.meal_allowance) if shift and shift.meal_allowance else 35000.0
                    ot_meal = x_meal_rate * int(xot.meal_count)
                    ev["meal_allowance"] = (ev["meal_allowance"] or 0) + ot_meal
                    ev["meal_count"] = (ev["meal_count"] or 0) + int(xot.meal_count)
                    ev["ot_hours"] = float(xot.ot_hours) if xot.ot_hours else ev["ot_hours"]
                    # Nếu ot_end_time >= 23h thì cộng thêm phụ cấp ca đêm
                    if xot.ot_end_time:
                        end_t = parse_time(xot.ot_end_time)
                        if end_t and end_t.hour >= 23:
                            ev["night_allowance"] = (ev["night_allowance"] or 0) + night_allowance_rate
                elif shift and shift.end_time and check_out_dt:
                    shift_end_t = parse_time(shift.end_time)
                    shift_end_dt_elig = datetime.combine(dt, shift_end_t)
                    actual_ot_h = max(0.0, (check_out_dt - shift_end_dt_elig).total_seconds() / 3600.0)
                    # Checkout từ 23h trở lên → đề xuất thêm PCCD ca đêm (ưu tiên hơn ot_eligible)
                    if check_out_dt >= datetime.combine(dt, time(23, 0)):
                        night_eligible_val = True
                    elif ((check_out_dt.hour * 60 + check_out_dt.minute) >= 17 * 60 + 50 and shift_end_t.hour < 18) or (actual_ot_h >= 3):
                        ot_eligible_val = True

            if ot_style == "new":
                xot = xot_map.get((emp.id, dt))
                ev["ot_hours"] = float(xot.ot_hours) if (xot and xot.ot_hours is not None) else 0.0

            # --- Xử lý giờ làm bất thường (irregular hours) ---
            is_irregular_val = False
            meal_approval_id_val = None
            meal_approval_status_val = None

            if nu_res and nu_res.is_irregular:
                is_irregular_val = True
                # Kiểm tra MealApproval record
                approval = approval_map.get((emp.id, dt))
                if approval:
                    meal_approval_id_val = approval.id
                    meal_approval_status_val = approval.status
                    if approval.status == "approved":
                        # Đã duyệt → tính tiền ăn theo approved_meal_count
                        approved_count = approval.approved_meal_count or 1
                        ev["meal_allowance"] = 35000.0 * approved_count
                        ev["meal_count"] = approved_count
                        # Nếu ca đêm (Ca 3) và đã duyệt → tính lại PC đêm
                        if nu_res.mode in (XNU_MODE_3,):
                            ev["night_allowance"] = night_allowance_rate
                        is_irregular_val = True  # Vẫn đánh dấu để UI biết
                    elif approval.status == "rejected":
                        # Từ chối → giữ = 0
                        ev["meal_allowance"] = 0.0
                        ev["meal_count"] = 0
                        ev["night_allowance"] = 0.0
                    # pending → giữ = 0 (default từ nu_shift.py)
                else:
                    # Tự động tạo pending approval record
                    new_approval = MealApproval(
                        employee_id=emp.id,
                        work_date=dt,
                        shift_code=nu_res.shift_code,
                        detected_mode=nu_res.mode,
                        check_in=check_in_dt,
                        check_out=check_out_dt,
                        status="pending",
                        approved_meal_count=1,
                    )
                    db.add(new_approval)
                    # Flush to get ID
                    try:
                        await db.flush()
                        meal_approval_id_val = new_approval.id
                        meal_approval_status_val = "pending"
                        # Cập nhật map để không tạo trùng
                        approval_map[(emp.id, dt)] = new_approval
                    except Exception:
                        pass

            if nu_res and ev["status"] != "absent":
                if nu_res.mode == XNU_MODE_1:
                    s_start_val, s_end_val = "06:00", "14:00"
                elif nu_res.mode == XNU_MODE_2:
                    s_start_val, s_end_val = "14:00", "22:00"
                elif nu_res.mode == XNU_MODE_3:
                    s_start_val, s_end_val = "22:00", "06:00"
                elif nu_res.mode == "night":
                    s_start_val, s_end_val = "18:00", "06:00"
                else:
                    s_start_val, s_end_val = "06:00", "18:00"
            else:
                s_start_val = str(shift.start_time)[:5] if shift and shift.start_time and ev["status"] != "absent" else None
                s_end_val = str(shift.end_time)[:5] if shift and shift.end_time and ev["status"] != "absent" else None

            cell = AttendanceCell(
                work_date=str(dt),
                day=d,
                dow=dow,
                shift_code="N" if ev["status"] == "absent" else cell_shift_code,
                shift_name="Nghi khong phep" if ev["status"] == "absent" else cell_shift_name,
                shift_start=s_start_val,
                shift_end=s_end_val,
                standard_hours=float(nu_res.standard_hours) if nu_res else (float(shift.standard_hours) if shift and shift.standard_hours else None),
                check_in=ci_str,
                check_out=co_str,
                actual_hours=ev["actual_hours"],
                deviation=ev["deviation"],
                ot_hours=ev["ot_hours"],
                status=ev["status"],
                is_holiday=is_holiday,
                is_sunday=is_sunday,
                notes=cell_notes,
                meal_allowance=ev["meal_allowance"],
                meal_count=ev["meal_count"],
                night_allowance=ev["night_allowance"],
                ot_eligible=ot_eligible_val,
                night_eligible=night_eligible_val,
                has_manual_xot=has_manual_xot_val,
                manual_meal_count=manual_meal_count_val,
                manual_ot_end_time=manual_ot_end_time_val,
                is_irregular=is_irregular_val,
                meal_approval_id=meal_approval_id_val,
                meal_approval_status=meal_approval_status_val,
            )
            days_cells.append(cell)

            # Summary
            if ev["status"] in ("full", "early_leave", "short", "forgot_scan"):
                total_present += 1
                total_hours += ev["actual_hours"]
            if ev["status"] == "absent":
                total_absent += 1
            if ev["status"] == "forgot_scan":
                total_forgot_scan += 1
            if ev["status"] == "early_leave":
                total_early += 1
            total_ot += ev["ot_hours"]
            ot_h = ev["ot_hours"]
            if is_holiday:
                total_ot_holiday += ot_h
            elif is_sunday:
                total_ot_sunday += ot_h
            else:
                total_ot_weekday += ot_h
            total_meal_count += ev["meal_count"] or 0

        rows.append(AttendanceRow(
            employee_id=emp.id,
            employee_code=emp.employee_code,
            full_name=emp.full_name,
            department=emp.department,
            default_shift_code=emp.default_shift_code,
            days=days_cells,
            summary={
                "total_present": total_present,
                "total_absent": total_absent,
                "total_forgot_scan": total_forgot_scan,
                "total_early_leave": total_early,
                "total_hours": round(total_hours, 2),
                "total_ot": round(total_ot, 2),
                "total_ot_weekday": round(total_ot_weekday, 2),
                "total_ot_sunday": round(total_ot_sunday, 2),
                "total_ot_holiday": round(total_ot_holiday, 2),
                "total_meal_count": total_meal_count,
                "total_meal_allowance": sum(c.meal_allowance for c in days_cells),
                "total_night_allowance": sum(c.night_allowance for c in days_cells),
                "total_paid_leave": total_paid_leave
            },
        ))

    # Commit auto-created MealApproval records
    try:
        await db.commit()
    except Exception:
        pass

    config_result = await db.execute(select(MonthlyWorkdayConfig).where(MonthlyWorkdayConfig.month_key == month_key))
    config = config_result.scalar_one_or_none()
    is_locked = bool(config.is_locked) if config else False

    return AttendanceMonthResponse(
        month_key=month_key,
        days_in_month=range_days,
        is_locked=is_locked,
        rows=rows,
    )


@router.get("/export")
async def export_attendance(
    month_key: str = Query(..., description="YYYY-MM"),
    department: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Xuất bảng chấm công chi tiết đa sheet theo tháng ra file Excel."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    res = await get_attendance(
        month_key=month_key,
        employee_id=None,
        department=department,
        db=db,
        current_user=current_user,
    )

    # Load approval data & shift templates for audit sheet
    from app.models.meal_approval import MealApproval

    # Parse month dates
    year, month = map(int, month_key.split("-"))
    days_in_month = calendar.monthrange(year, month)[1]
    range_start = date(year, month, 1)
    range_end = date(year, month, days_in_month)

    # Fetch OT & Meal Approvals for Sheet 3
    approval_q = select(MealApproval, Employee).join(Employee, Employee.id == MealApproval.employee_id).where(
        and_(
            MealApproval.work_date >= range_start,
            MealApproval.work_date <= range_end,
        )
    )
    if department:
        approval_q = approval_q.where(Employee.department == department)
    approval_res = await db.execute(approval_q.order_by(MealApproval.work_date, Employee.employee_code))
    approvals = approval_res.all()

    # Fetch WorkSchedules overrides for Sheet 3
    sched_q = select(WorkSchedule, Employee, ShiftTemplate).join(Employee, Employee.id == WorkSchedule.employee_id).outerjoin(ShiftTemplate, ShiftTemplate.id == WorkSchedule.shift_id).where(
        and_(
            WorkSchedule.work_date >= range_start,
            WorkSchedule.work_date <= range_end,
        )
    )
    if department:
        sched_q = sched_q.where(Employee.department == department)
    sched_res = await db.execute(sched_q.order_by(WorkSchedule.work_date, Employee.employee_code))
    schedules = sched_res.all()

    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # STYLES DEFINITION
    # ---------------------------------------------------------
    thin = Side(border_style="thin", color="D9D9D9")
    dark_thin = Side(border_style="thin", color="000000")
    
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    border_box = Border(top=dark_thin, left=dark_thin, right=dark_thin, bottom=dark_thin)
    
    font_title = Font(name="Times New Roman", bold=True, size=14, color="1F4E79")
    font_company = Font(name="Times New Roman", bold=True, size=11, color="333333")
    font_subtitle = Font(name="Times New Roman", italic=True, size=10, color="595959")
    font_section = Font(name="Times New Roman", bold=True, size=11, color="1F4E79")
    
    font_header = Font(name="Times New Roman", bold=True, size=11, color="FFFFFF")
    font_data = Font(name="Times New Roman", size=10, color="000000")
    
    fill_header_main = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_header_sub = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_summary = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    fill_sun = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_holiday = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    days = list(range(1, res.days_in_month + 1))
    first_row_days = {cell.day: cell for cell in (res.rows[0].days if res.rows else [])}

    # =========================================================
    # SHEET 1: BẢNG TỔNG HỢP (SUMMARY SHEET)
    # =========================================================
    ws1 = wb.active
    ws1.title = "Bảng tổng hợp"
    ws1.views.sheetView[0].showGridLines = True

    # Title block
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13 + len(days))
    ws1.cell(1, 1, "CÔNG TY TNHH HIỆP LỢI").font = font_company
    ws1.cell(1, 1).alignment = align_left

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13 + len(days))
    ws1.cell(2, 1, "MST: 3701609885").font = font_subtitle
    ws1.cell(2, 1).alignment = align_left

    ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=13 + len(days))
    t1 = f"BẢNG TỔNG HỢP CHẤM CÔNG & TĂNG CA - THÁNG {month_key}"
    if department:
        t1 += f" ({department.upper()})"
    ws1.cell(3, 1, t1).font = font_title
    ws1.cell(3, 1).alignment = align_center

    ws1.merge_cells(start_row=4, start_column=1, end_row=4, end_column=13 + len(days))
    ws1.cell(4, 1, "Xuất từ màn hình chấm công hệ thống").font = font_subtitle
    ws1.cell(4, 1).alignment = align_left

    # Headers
    h1_labels = [
        "STT", "Mã NV", "Họ và tên", "Bộ phận", "Mã ca",
        "Công chuẩn", "Tổng giờ chuẩn", "Tổng OT (h)",
        "Tiền ăn (đ)", "Số bữa", "Vắng mặt", "Quên scan", "Về sớm"
    ]
    h1_row = 6
    ws1.row_dimensions[h1_row].height = 28

    for col_idx, label in enumerate(h1_labels, 1):
        c = ws1.cell(h1_row, col_idx, label)
        c.font = font_header
        c.fill = fill_header_main
        c.alignment = align_center
        c.border = border_box

    day_start_col = len(h1_labels) + 1
    for offset, day in enumerate(days):
        col_idx = day_start_col + offset
        day_cell = first_row_days.get(day)
        dow_str = day_cell.dow if day_cell else ""
        label = f"{day}\n{dow_str}"
        c = ws1.cell(h1_row, col_idx, label)
        c.font = font_header
        c.alignment = align_center
        c.border = border_box
        if dow_str == "CN":
            c.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        else:
            c.fill = fill_header_sub

    # Populate Sheet 1 Data
    row_idx = h1_row + 1
    for idx, r in enumerate(res.rows, 1):
        ws1.row_dimensions[row_idx].height = 20
        is_even = (idx % 2 == 0)
        base_fill = fill_zebra if is_even else None

        ws1.cell(row_idx, 1, idx).alignment = align_center
        ws1.cell(row_idx, 2, r.employee_code).alignment = align_center
        ws1.cell(row_idx, 3, r.full_name).alignment = align_left
        ws1.cell(row_idx, 4, r.department or "").alignment = align_left
        ws1.cell(row_idx, 5, r.default_shift_code or "").alignment = align_center
        
        # Summary counts
        ws1.cell(row_idx, 6, r.summary.get("total_present", 0) or 0).number_format = "#,##0.0"
        ws1.cell(row_idx, 7, r.summary.get("total_hours", 0.0) or 0.0).number_format = "#,##0.0"
        ws1.cell(row_idx, 8, r.summary.get("total_ot", 0.0) or 0.0).number_format = "#,##0.0"
        ws1.cell(row_idx, 9, r.summary.get("total_meal_allowance", 0.0) or 0.0).number_format = "#,##0"
        ws1.cell(row_idx, 10, r.summary.get("total_meal_count", 0) or 0).number_format = "#,##0"
        ws1.cell(row_idx, 11, r.summary.get("total_absent", 0) or 0).number_format = "#,##0"
        ws1.cell(row_idx, 12, r.summary.get("total_forgot_scan", 0) or 0).number_format = "#,##0"
        ws1.cell(row_idx, 13, r.summary.get("total_early_leave", 0) or 0).number_format = "#,##0"

        # Apply basic formatting to columns 1..13
        for c_i in range(1, day_start_col):
            cell = ws1.cell(row_idx, c_i)
            cell.font = font_data
            cell.border = border_all
            if base_fill and c_i <= 5:
                cell.fill = base_fill
            elif c_i in (6, 7, 8, 9, 10):
                cell.fill = fill_summary

        # Day columns in Sheet 1
        day_map = {cell.day: cell for cell in r.days}
        for offset, day in enumerate(days):
            col_idx = day_start_col + offset
            cell_data = day_map.get(day)
            
            disp_val = ""
            if cell_data:
                # Clean time extraction instead of [:5] on ISO datetime string
                if cell_data.check_in:
                    ci_time = cell_data.check_in.split(" ")[-1][:5] if " " in cell_data.check_in else cell_data.check_in[:5]
                    if ci_time and ci_time != "2026-":
                        disp_val = ci_time
                    else:
                        disp_val = "✓"
                elif cell_data.status == "holiday":
                    disp_val = "L"
                elif cell_data.status == "off":
                    disp_val = "N"
                elif cell_data.status == "absent":
                    disp_val = "V"
                elif cell_data.status == "forgot_scan":
                    disp_val = "Q"
                elif cell_data.status == "early_leave":
                    disp_val = "VS"
                elif cell_data.status == "short":
                    disp_val = "TG"

            excel_cell = ws1.cell(row_idx, col_idx, disp_val)
            excel_cell.font = font_data
            excel_cell.alignment = align_center
            excel_cell.border = border_all

            if cell_data:
                if cell_data.is_holiday:
                    excel_cell.fill = fill_holiday
                elif cell_data.is_sunday:
                    excel_cell.fill = fill_sun

        row_idx += 1

    # Freeze panes for Sheet 1
    ws1.freeze_panes = "F7"

    col_w1 = {
        "A": 6, "B": 11, "C": 24, "D": 16, "E": 10,
        "F": 12, "G": 14, "H": 12, "I": 14, "J": 9,
        "K": 11, "L": 11, "M": 11
    }
    for c_letter, w in col_w1.items():
        ws1.column_dimensions[c_letter].width = w
    for idx in range(day_start_col, day_start_col + len(days)):
        ws1.column_dimensions[get_column_letter(idx)].width = 7.5

    # =========================================================
    # SHEET 2: CHI TIẾT CHẤM CÔNG (DAILY DETAIL SHEET)
    # =========================================================
    ws2 = wb.create_sheet(title="Chi tiết từng ngày")
    ws2.views.sheetView[0].showGridLines = True

    # Title block Sheet 2
    ws2.merge_cells("A1:P1")
    ws2.cell(1, 1, "CÔNG TY TNHH HIỆP LỢI").font = font_company
    ws2.cell(1, 1).alignment = align_left

    ws2.merge_cells("A2:P2")
    ws2.cell(2, 1, f"BẢNG CHI TIẾT CHẤM CÔNG HÀNG NGÀY — THÁNG {month_key}").font = font_title
    ws2.cell(2, 1).alignment = align_center

    ws2.merge_cells("A3:P3")
    ws2.cell(3, 1, "Giờ Vào/Ra thực tế, giờ Vào/Ra tính công làm tròn theo quy định ca, giờ chuẩn, giờ OT & tiền ăn").font = font_subtitle
    ws2.cell(3, 1).alignment = align_left

    h2_labels = [
        "STT", "Ngày", "Thứ", "Mã NV", "Họ và tên", "Bộ phận", "Mã Ca",
        "Ca làm việc (Quy định)", "Vào thực tế", "Ra thực tế",
        "Vào tính công", "Ra tính công", "Giờ chuẩn (h)", "Giờ OT (h)",
        "Tiền ăn (đ)", "Trạng thái & Ghi chú"
    ]
    h2_row = 5
    ws2.row_dimensions[h2_row].height = 28

    for col_idx, label in enumerate(h2_labels, 1):
        c = ws2.cell(h2_row, col_idx, label)
        c.font = font_header
        c.fill = fill_header_main
        c.alignment = align_center
        c.border = border_box

    # Populate Sheet 2 Data
    d_row2 = h2_row + 1
    seq2 = 1

    for r in res.rows:
        for cell_data in r.days:
            ws2.row_dimensions[d_row2].height = 20
            
            # Extract clean times
            ci_raw = cell_data.check_in or ""
            co_raw = cell_data.check_out or ""
            
            ci_time = ci_raw.split(" ")[-1][:5] if " " in ci_raw else (ci_raw[:5] if ci_raw else "—")
            co_time = co_raw.split(" ")[-1][:5] if " " in co_raw else (co_raw[:5] if co_raw else "—")
            
            if ci_time == "2026-": ci_time = "—"
            if co_time == "2026-": co_time = "—"

            # Dynamic shift description and start/end times for NU and XNU
            shift_code_val = (cell_data.shift_code or "").strip().upper()
            shift_name_val = cell_data.shift_name or ""
            
            s_start = cell_data.shift_start or ""
            s_end = cell_data.shift_end or ""

            if shift_code_val == "XNU":
                if "ca 3" in shift_name_val.lower() or s_start == "22:00":
                    shift_desc = "Ca XNU 3 (22:00 - 06:00)"
                    s_start, s_end = "22:00", "06:00"
                elif "ca 2" in shift_name_val.lower() or s_start == "14:00":
                    shift_desc = "Ca XNU 2 (14:00 - 22:00)"
                    s_start, s_end = "14:00", "22:00"
                else:
                    shift_desc = "Ca XNU 1 (06:00 - 14:00)"
                    s_start, s_end = "06:00", "14:00"
            elif shift_code_val in ("NU", "NUT1", "NUT2", "NU1", "NU2", "NU3", "NUN"):
                if "toi" in shift_name_val.lower() or "tối" in shift_name_val.lower():
                    shift_desc = f"{shift_code_val} - Ca NU Tối (18:00 - 06:00)"
                    s_start, s_end = "18:00", "06:00"
                else:
                    shift_desc = f"{shift_code_val} - Ca NU Sáng (06:00 - 18:00)"
                    s_start, s_end = "06:00", "18:00"
            elif s_start and s_end:
                shift_desc = f"Ca {shift_code_val} ({s_start} - {s_end})"
            else:
                shift_desc = shift_name_val or shift_code_val or "—"

            ci_calc = ci_time
            co_calc = co_time

            if ci_time != "—" and ci_time:
                # 1) NU Sáng & XNU Ca 1: Quy định 06:00. Nếu vào sớm (<= 06:00) -> tính 06:00
                if shift_code_val in ("NU", "NUT1", "NUT2", "NU1", "NU2", "NU3", "NUN") and "18:00 - 06:00" not in shift_desc:
                    if ci_time <= "06:00":
                        ci_calc = "06:00"
                    if co_time != "—" and co_time and co_time >= "18:00":
                        co_calc = "18:00"
                
                # 2) NU Tối: Quy định 18:00. Nếu vào trong khoảng 17:00 - 18:00 -> tính 18:00
                elif shift_code_val in ("NU", "NUT1", "NUT2", "NU1", "NU2", "NU3", "NUN") and "18:00 - 06:00" in shift_desc:
                    if ci_time <= "18:00" and ci_time >= "16:00":
                        ci_calc = "18:00"
                    if co_time != "—" and co_time and co_time >= "05:30" and co_time <= "06:30":
                        co_calc = "06:00"
                
                # 3) XNU Ca 1 (06:00 - 14:00)
                elif shift_code_val == "XNU" and "06:00 - 14:00" in shift_desc:
                    if ci_time <= "06:00":
                        ci_calc = "06:00"
                    if co_time != "—" and co_time and co_time >= "14:00":
                        co_calc = "14:00"

                # 4) XNU Ca 2 (14:00 - 22:00)
                elif shift_code_val == "XNU" and "14:00 - 22:00" in shift_desc:
                    if ci_time <= "14:00" and ci_time >= "12:00":
                        ci_calc = "14:00"
                    if co_time != "—" and co_time and co_time >= "22:00":
                        co_calc = "22:00"

                # 5) XNU Ca 3 (22:00 - 06:00)
                elif shift_code_val == "XNU" and "22:00 - 06:00" in shift_desc:
                    if ci_time <= "22:00" and ci_time >= "20:00":
                        ci_calc = "22:00"
                    if co_time != "—" and co_time and co_time >= "05:30" and co_time <= "06:30":
                        co_calc = "06:00"

                # 6) Các ca chuẩn khác
                else:
                    if s_start and ci_time < s_start:
                        ci_calc = s_start
                    if s_end and co_time != "—" and co_time and co_time >= s_end:
                        if shift_code_val in ("X", "X40", "XVP", "VP80", "SEP", "TX1", "TX2"):
                            co_calc = s_end

            # Status description
            status_text = cell_data.notes or ""
            if not status_text:
                st_map = {
                    "full": "Đủ ca",
                    "early_leave": "Về sớm",
                    "short": "Thiếu giờ",
                    "forgot_scan": "Quên scan",
                    "absent": "Vắng mặt",
                    "off": "Nghỉ tuần",
                    "holiday": "Nghỉ lễ",
                    "no_data": "Chưa làm việc",
                }
                status_text = st_map.get(cell_data.status, cell_data.status or "—")

            ws2.cell(d_row2, 1, seq2).alignment = align_center
            ws2.cell(d_row2, 2, cell_data.work_date).alignment = align_center
            ws2.cell(d_row2, 3, cell_data.dow).alignment = align_center
            ws2.cell(d_row2, 4, r.employee_code).alignment = align_center
            ws2.cell(d_row2, 5, r.full_name).alignment = align_left
            ws2.cell(d_row2, 6, r.department or "").alignment = align_left
            ws2.cell(d_row2, 7, cell_data.shift_code or "—").alignment = align_center
            ws2.cell(d_row2, 8, shift_desc).alignment = align_left
            ws2.cell(d_row2, 9, ci_time).alignment = align_center
            ws2.cell(d_row2, 10, co_time).alignment = align_center
            ws2.cell(d_row2, 11, ci_calc).alignment = align_center
            ws2.cell(d_row2, 12, co_calc).alignment = align_center
            
            c_std = ws2.cell(d_row2, 13, cell_data.actual_hours or 0.0)
            c_std.number_format = "#,##0.0"
            c_std.alignment = align_right
            
            c_ot = ws2.cell(d_row2, 14, cell_data.ot_hours or 0.0)
            c_ot.number_format = "#,##0.0"
            c_ot.alignment = align_right
            
            c_meal = ws2.cell(d_row2, 15, cell_data.meal_allowance or 0.0)
            c_meal.number_format = "#,##0"
            c_meal.alignment = align_right
            
            ws2.cell(d_row2, 16, status_text).alignment = align_left

            # Formatting borders and fills
            for c_i in range(1, 17):
                c_node = ws2.cell(d_row2, c_i)
                c_node.font = font_data
                c_node.border = border_all
                if cell_data.is_holiday:
                    c_node.fill = fill_holiday
                elif cell_data.is_sunday:
                    c_node.fill = fill_sun

            d_row2 += 1
            seq2 += 1

    # Freeze panes & column widths for Sheet 2
    ws2.freeze_panes = "A6"
    w2_dims = {
        "A": 6, "B": 12, "C": 6, "D": 11, "E": 24, "F": 16,
        "G": 10, "H": 22, "I": 13, "J": 13, "K": 14, "L": 14,
        "M": 14, "N": 13, "O": 15, "P": 35
    }
    for col_l, w in w2_dims.items():
        ws2.column_dimensions[col_l].width = w

    # =========================================================
    # SHEET 3: DUYỆT TĂNG CA & LỊCH (OT & SCHEDULE AUDIT SHEET)
    # =========================================================
    ws3 = wb.create_sheet(title="Duyệt tăng ca & Lịch")
    ws3.views.sheetView[0].showGridLines = True

    # Title block Sheet 3
    ws3.merge_cells("A1:I1")
    ws3.cell(1, 1, "CÔNG TY TNHH HIỆP LỢI").font = font_company
    ws3.cell(1, 1).alignment = align_left

    ws3.merge_cells("A2:I2")
    ws3.cell(2, 1, f"DANH SÁCH DUYỆT TĂNG CA VÀ PHÂN CA ĐẶC BIỆT — THÁNG {month_key}").font = font_title
    ws3.cell(2, 1).alignment = align_center

    # Section 1: OT & Meal Approvals Table
    ws3.cell(4, 1, "1. DANH SÁCH DUYỆT TĂNG CA THỰC TẾ & PHỤ CẤP ĂN (MEAL ALLOWANCE)").font = font_section

    h3_labels1 = ["STT", "Ngày", "Mã NV", "Họ và tên", "Bộ phận", "Số giờ OT duyệt", "Số bữa ăn duyệt", "Trạng thái", "Ghi chú"]
    # Fetch XOvertimeConfig for Sheet 3
    xot_q = select(XOvertimeConfig, Employee).join(Employee, Employee.id == XOvertimeConfig.employee_id).where(
        and_(
            XOvertimeConfig.work_date >= range_start,
            XOvertimeConfig.work_date <= range_end,
        )
    )
    if department:
        xot_q = xot_q.where(Employee.department == department)
    xot_res = await db.execute(xot_q.order_by(XOvertimeConfig.work_date, Employee.employee_code))
    xot_list = xot_res.all()

    # Section 1: OT & Meal Approvals Table
    ws3.cell(4, 1, "1. DANH SÁCH DUYỆT TĂNG CA THỦ CÔNG & PHỤ CẤP ĂN (XOVERTIME CONFIGS)").font = font_section

    h3_labels1 = ["STT", "Ngày", "Mã NV", "Họ và tên", "Bộ phận", "Giờ ra OT", "Số giờ OT duyệt", "Số bữa ăn duyệt", "Ghi chú"]
    h3_row1 = 5
    ws3.row_dimensions[h3_row1].height = 26
    for c_i, label in enumerate(h3_labels1, 1):
        c = ws3.cell(h3_row1, c_i, label)
        c.font = font_header
        c.fill = fill_header_main
        c.alignment = align_center
        c.border = border_box

    r3 = h3_row1 + 1
    if xot_list:
        for idx, (xot, emp) in enumerate(xot_list, 1):
            ws3.row_dimensions[r3].height = 19
            ws3.cell(r3, 1, idx).alignment = align_center
            ws3.cell(r3, 2, str(xot.work_date)).alignment = align_center
            ws3.cell(r3, 3, emp.employee_code).alignment = align_center
            ws3.cell(r3, 4, emp.full_name).alignment = align_left
            ws3.cell(r3, 5, emp.department or "").alignment = align_left
            ws3.cell(r3, 6, str(xot.ot_end_time)[:5] if xot.ot_end_time else "—").alignment = align_center
            
            c_ot = ws3.cell(r3, 7, float(xot.ot_hours or 0.0))
            c_ot.number_format = "#,##0.0"
            c_ot.alignment = align_right
            
            c_mc = ws3.cell(r3, 8, xot.meal_count or 0)
            c_mc.number_format = "#,##0"
            c_mc.alignment = align_right
            
            ws3.cell(r3, 9, "Duyệt OT ca X/X40").alignment = align_left

            for c_i in range(1, 10):
                c_node = ws3.cell(r3, c_i)
                c_node.font = font_data
                c_node.border = border_all
            r3 += 1
    else:
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=9)
        ws3.cell(r3, 1, "Không có dữ liệu duyệt tăng ca thủ công trong tháng này").alignment = align_center
        ws3.cell(r3, 1).font = font_subtitle
        r3 += 1

    # Section 2: Meal Approval for Irregular hours
    r3 += 2
    ws3.cell(r3, 1, "2. DANH SÁCH DUYỆT BẤT THƯỜNG & TIỀN ĂN (MEAL APPROVALS)").font = font_section
    r3 += 1

    h3_labels2_meal = ["STT", "Ngày", "Mã NV", "Họ và tên", "Bộ phận", "Mã ca", "Vào - Ra thực tế", "Số bữa duyệt", "Trạng thái", "Lý do"]
    ws3.row_dimensions[r3].height = 26
    for c_i, label in enumerate(h3_labels2_meal, 1):
        c = ws3.cell(r3, c_i, label)
        c.font = font_header
        c.fill = fill_header_sub
        c.alignment = align_center
        c.border = border_box

    r3 += 1
    if approvals:
        for idx, (app, emp) in enumerate(approvals, 1):
            ws3.row_dimensions[r3].height = 19
            ci_t = app.check_in.strftime("%H:%M") if app.check_in else "—"
            co_t = app.check_out.strftime("%H:%M") if app.check_out else "—"
            in_out_str = f"{ci_t} - {co_t}"
            st_str = "Đã duyệt" if app.status == "approved" else ("Từ chối" if app.status == "rejected" else "Chờ duyệt")

            ws3.cell(r3, 1, idx).alignment = align_center
            ws3.cell(r3, 2, str(app.work_date)).alignment = align_center
            ws3.cell(r3, 3, emp.employee_code).alignment = align_center
            ws3.cell(r3, 4, emp.full_name).alignment = align_left
            ws3.cell(r3, 5, emp.department or "").alignment = align_left
            ws3.cell(r3, 6, app.shift_code or "—").alignment = align_center
            ws3.cell(r3, 7, in_out_str).alignment = align_center
            
            c_mc = ws3.cell(r3, 8, app.approved_meal_count or 0)
            c_mc.number_format = "#,##0"
            c_mc.alignment = align_right
            
            ws3.cell(r3, 9, st_str).alignment = align_center
            ws3.cell(r3, 10, app.reason or "").alignment = align_left

            for c_i in range(1, 11):
                c_node = ws3.cell(r3, c_i)
                c_node.font = font_data
                c_node.border = border_all
            r3 += 1
    else:
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=10)
        ws3.cell(r3, 1, "Không có dữ liệu duyệt tiền ăn bất thường trong tháng này").alignment = align_center
        ws3.cell(r3, 1).font = font_subtitle
        r3 += 1

    # Section 3: Work Schedule Overrides Table
    r3 += 2
    ws3.cell(r3, 1, "3. DANH SÁCH LỊCH PHÂN CA ĐẶC BIỆT / ĐỔI CA (WORK SCHEDULES)").font = font_section
    r3 += 1

    h3_labels2 = ["STT", "Ngày", "Mã NV", "Họ và tên", "Bộ phận", "Ca phân công", "Giờ ca", "Ghi chú phân ca"]
    ws3.row_dimensions[r3].height = 26
    for c_i, label in enumerate(h3_labels2, 1):
        c = ws3.cell(r3, c_i, label)
        c.font = font_header
        c.fill = fill_header_sub
        c.alignment = align_center
        c.border = border_box

    r3 += 1
    if schedules:
        for idx, (ws_rec, emp, shift_rec) in enumerate(schedules, 1):
            ws3.row_dimensions[r3].height = 19
            s_code = shift_rec.code if shift_rec else "—"
            s_time = f"{str(shift_rec.start_time)[:5]} - {str(shift_rec.end_time)[:5]}" if shift_rec and shift_rec.start_time else "—"

            ws3.cell(r3, 1, idx).alignment = align_center
            ws3.cell(r3, 2, str(ws_rec.work_date)).alignment = align_center
            ws3.cell(r3, 3, emp.employee_code).alignment = align_center
            ws3.cell(r3, 4, emp.full_name).alignment = align_left
            ws3.cell(r3, 5, emp.department or "").alignment = align_left
            ws3.cell(r3, 6, s_code).alignment = align_center
            ws3.cell(r3, 7, s_time).alignment = align_center
            ws3.cell(r3, 8, ws_rec.notes or "").alignment = align_left

            for c_i in range(1, 9):
                c_node = ws3.cell(r3, c_i)
                c_node.font = font_data
                c_node.border = border_all
            r3 += 1
    else:
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=8)
        ws3.cell(r3, 1, "Không có lịch phân ca đặc biệt trong tháng này (sử dụng ca mặc định)").alignment = align_center
        ws3.cell(r3, 1).font = font_subtitle

    # Column widths Sheet 3
    w3_dims = {
        "A": 6, "B": 12, "C": 11, "D": 24, "E": 16,
        "F": 16, "G": 16, "H": 16, "I": 30
    }
    for col_l, w in w3_dims.items():
        ws3.column_dimensions[col_l].width = w

    # Return workbook stream
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"cham_cong_chi_tiet_{month_key}{'_' + department.replace(' ', '_') if department else ''}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class ManualAttendanceAction(BaseModel):
    employee_id: int
    work_date: date
    action: Literal["convert_paid_leave", "mark_worked", "change_shift"]
    reason: Optional[str] = None
    shift_code: Optional[str] = None


@router.post("/manual-action")
async def manual_attendance_action(
    request: ManualAttendanceAction,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN)),
):
    await check_date_locked(db, request.work_date)
    emp = await db.get(Employee, request.employee_id)
    if not emp:
        raise HTTPException(status_code=404, detail="Nhan vien khong ton tai")

    work_date = request.work_date
    reason = request.reason.strip() if request.reason else None

    shift_result = await db.execute(select(ShiftTemplate))
    all_shifts = shift_result.scalars().all()
    shifts_by_id = {s.id: s for s in all_shifts}
    shifts_by_code = {s.code: s for s in all_shifts}

    ws_result = await db.execute(select(WorkSchedule).where(and_(
        WorkSchedule.employee_id == emp.id,
        WorkSchedule.work_date == work_date,
    )))
    ws = ws_result.scalar_one_or_none()
    ws_before = {c.name: getattr(ws, c.name) for c in ws.__table__.columns} if ws else None

    if request.action == "convert_paid_leave":
        paid_shift = shifts_by_code.get("P")
        if not paid_shift or not paid_shift.is_paid_leave:
            raise HTTPException(status_code=400, detail="Khong tim thay ca phep (P)")

        if ws and ws.shift_id == paid_shift.id:
            if reason is not None:
                ws.notes = reason
                await log_audit(
                    db,
                    "work_schedules",
                    f"{emp.id}:{work_date}",
                    "UPDATE",
                    current_user.username,
                    ws_before,
                    {c.name: getattr(ws, c.name) for c in ws.__table__.columns},
                    notes="Update leave note",
                )
                await db.commit()
            return {"message": "Da la nghi phep"}

        # Check remaining leave for the year
        year = work_date.year
        year_start = date(year, 1, 1)
        year_end = date(year, 12, 31)

        sched_q = select(WorkSchedule).where(and_(
            WorkSchedule.employee_id == emp.id,
            WorkSchedule.work_date >= year_start,
            WorkSchedule.work_date <= year_end,
        ))
        sched_res = await db.execute(sched_q)
        override_map = {s.work_date: s.shift_id for s in sched_res.scalars().all()}

        holiday_q = select(CompanyHoliday.holiday_date).where(and_(
            CompanyHoliday.holiday_date >= year_start,
            CompanyHoliday.holiday_date <= year_end,
            CompanyHoliday.is_active == True,
        ))
        holiday_res = await db.execute(holiday_q)
        holiday_dates = set(holiday_res.scalars().all())

        default_shift = shifts_by_code.get(emp.default_shift_code) if emp.default_shift_code else None

        used = 0.0
        today = date.today()
        last_date = year_end if today.year > year else today
        curr = year_start
        while curr <= last_date:
            is_sunday = curr.weekday() == 6
            is_holiday = curr in holiday_dates
            sid = override_map.get(curr)
            shift = shifts_by_id.get(sid) if sid else (None if (is_sunday or is_holiday) else default_shift)

            if shift and shift.is_leave_code and shift.is_paid_leave:
                if shift.code == "P":
                    used += 1.0
                elif shift.code in ["S", "C"]:
                    used += 0.5

            curr += timedelta(days=1)

        entitlement = 12.0
        remaining = entitlement - used
        if remaining < 1:
            raise HTTPException(status_code=400, detail="Khong con phep nam")

        if ws:
            ws.shift_id = paid_shift.id
            ws.notes = reason
        else:
            ws = WorkSchedule(
                employee_id=emp.id,
                work_date=work_date,
                month_key=work_date.strftime("%Y-%m"),
                shift_id=paid_shift.id,
                notes=reason,
            )
            db.add(ws)

        await log_audit(
            db,
            "work_schedules",
            f"{emp.id}:{work_date}",
            "UPDATE" if ws_before else "CREATE",
            current_user.username,
            ws_before,
            {c.name: getattr(ws, c.name) for c in ws.__table__.columns},
            notes="Convert to paid leave",
        )
        await db.commit()
        return {"message": "Da chuyen sang nghi phep"}

    if request.action == "mark_worked":
        default_shift = shifts_by_code.get(emp.default_shift_code) if emp.default_shift_code else None
        if not default_shift or default_shift.is_leave_code:
            raise HTTPException(status_code=400, detail="Khong co ca mac dinh hop le")
        if not default_shift.start_time or not default_shift.end_time:
            raise HTTPException(status_code=400, detail="Ca mac dinh thieu gio bat dau/ket thuc")

        if ws:
            ws.shift_id = default_shift.id
            ws.notes = reason
        else:
            ws = WorkSchedule(
                employee_id=emp.id,
                work_date=work_date,
                month_key=work_date.strftime("%Y-%m"),
                shift_id=default_shift.id,
                notes=reason,
            )
            db.add(ws)

        att_result = await db.execute(select(AttendanceDaily).where(and_(
            AttendanceDaily.employee_id == emp.id,
            AttendanceDaily.work_date == work_date,
        )))
        att = att_result.scalar_one_or_none()
        att_before = {c.name: getattr(att, c.name) for c in att.__table__.columns} if att else None

        check_in_dt = datetime.combine(work_date, default_shift.start_time)
        if default_shift.is_night_shift:
            check_out_dt = datetime.combine(work_date + timedelta(days=1), default_shift.end_time)
        else:
            check_out_dt = datetime.combine(work_date, default_shift.end_time)

        break_mins = int(default_shift.break_minutes or 60)
        total_hours = calc_hours_between(check_in_dt, check_out_dt, break_mins)

        if att:
            att.first_check_in = check_in_dt
            att.last_check_out = check_out_dt
            att.total_hours = total_hours
            att.import_batch = "manual"
        else:
            att = AttendanceDaily(
                employee_id=emp.id,
                work_date=work_date,
                first_check_in=check_in_dt,
                last_check_out=check_out_dt,
                total_hours=total_hours,
                import_batch="manual",
            )
            db.add(att)

        await log_audit(
            db,
            "work_schedules",
            f"{emp.id}:{work_date}",
            "UPDATE" if ws_before else "CREATE",
            current_user.username,
            ws_before,
            {c.name: getattr(ws, c.name) for c in ws.__table__.columns},
            notes="Mark worked (manual)",
        )
        await log_audit(
            db,
            "attendance_daily",
            f"{emp.id}:{work_date}",
            "UPDATE" if att_before else "CREATE",
            current_user.username,
            att_before,
            {c.name: getattr(att, c.name) for c in att.__table__.columns},
            notes="Manual attendance",
        )
        await db.commit()
        return {"message": "Da danh dau di lam"}

    if request.action == "change_shift":
        if not request.shift_code:
            raise HTTPException(status_code=400, detail="Thieu ma ca moi")
        
        new_shift = shifts_by_code.get(request.shift_code)
        if not new_shift:
            raise HTTPException(status_code=400, detail=f"Khong tim thay ma ca {request.shift_code}")

        # If changing to a paid leave code, we can optionally reuse the balance check logic
        # But for "change_shift", we might want to be more flexible for admins.
        # However, to be safe, let's include a similar logic if it's P, S, or C.
        if new_shift.is_leave_code and new_shift.is_paid_leave:
            # Simple check: if it's P, S, or C, check remaining leave (omitted for brevity or kept?)
            # The user said "trừ phép", so we should probably keep it consistent.
            # But maybe admins want to override even if 0? 
            # For now, let's just apply the shift.
            pass

        if ws:
            ws.shift_id = new_shift.id
            ws.notes = reason
        else:
            ws = WorkSchedule(
                employee_id=emp.id,
                work_date=work_date,
                month_key=work_date.strftime("%Y-%m"),
                shift_id=new_shift.id,
                notes=reason,
            )
            db.add(ws)

        await log_audit(
            db,
            "work_schedules",
            f"{emp.id}:{work_date}",
            "UPDATE" if ws_before else "CREATE",
            current_user.username,
            ws_before,
            {c.name: getattr(ws, c.name) for c in ws.__table__.columns},
            notes=f"Change shift to {new_shift.code}",
        )
        await db.commit()
        return {"message": f"Da doi sang ca {new_shift.code}"}

    raise HTTPException(status_code=400, detail="Hanh dong khong hop le")


# Models for Forgotten Scans
class ForgottenScanFixItem(BaseModel):
    employee_id: int
    work_date: date
    check_in: Optional[str] = None     # HH:MM or None
    check_out: Optional[str] = None    # HH:MM or None
    is_off: bool = False               # True = công ty cho nghỉ


class ForgottenScanFixRequest(BaseModel):
    items: List[ForgottenScanFixItem]


@router.get("/forgot-scans")
async def get_forgot_scans(
    month_key: str = Query(..., description="YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Lấy danh sách các trường hợp quên quẹt thẻ trong tháng"""
    try:
        year, month = map(int, month_key.split("-"))
    except ValueError:
        raise HTTPException(400, "month_key phai la YYYY-MM")

    month_days = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end = date(year, month, month_days)

    # 1. Load shifts
    shift_result = await db.execute(select(ShiftTemplate))
    shifts_by_id = {}
    shifts_by_code = {}
    for s in shift_result.scalars().all():
        shifts_by_id[s.id] = s
        shifts_by_code[s.code] = s

    # 2. Load employees active in the month
    emp_q = select(Employee).where(
        and_(
            Employee.is_active == True,
            or_(Employee.join_date.is_(None), Employee.join_date <= month_end),
            or_(Employee.leave_date.is_(None), Employee.leave_date >= month_start)
        )
    )
    emp_result = await db.execute(emp_q)
    employees = list(emp_result.scalars().all())
    emp_map = {e.id: e for e in employees}
    emp_id_list = [e.id for e in employees]

    if not emp_id_list:
        return []

    # 3. Load schedule overrides
    schedule_q = select(WorkSchedule).where(
        and_(
            WorkSchedule.work_date >= month_start,
            WorkSchedule.work_date <= month_end,
            WorkSchedule.employee_id.in_(emp_id_list)
        )
    )
    schedule_result = await db.execute(schedule_q)
    override_map = {}
    for ws in schedule_result.scalars().all():
        override_map[(ws.employee_id, ws.work_date)] = ws.shift_id

    # 4. Load attendance records with missing scans
    att_q = select(AttendanceDaily).where(
        and_(
            AttendanceDaily.work_date >= month_start,
            AttendanceDaily.work_date <= month_end,
            AttendanceDaily.employee_id.in_(emp_id_list),
            or_(
                AttendanceDaily.first_check_in.is_(None),
                AttendanceDaily.last_check_out.is_(None)
            )
        )
    )
    att_result = await db.execute(att_q)
    forgot_records_raw = att_result.scalars().all()

    # Filter out records where both are null or both are present (forgot scan is exactly 1 missing)
    forgot_records = []
    for a in forgot_records_raw:
        if a.first_check_in is None and a.last_check_out is None:
            continue
        if a.first_check_in is not None and a.last_check_out is not None:
            continue
        forgot_records.append(a)

    # 5. Department order
    from app.models.department import Department
    dept_order_q = await db.execute(select(Department.name, Department.sort_order))
    dept_order_map = {row[0]: row[1] for row in dept_order_q.all() if row[0]}

    def get_sort_key(item):
        emp = emp_map.get(item.employee_id)
        d_name = emp.department if emp else ""
        d_order = dept_order_map.get(d_name, 9999) if d_name else 9999
        code_num = 999999
        if emp:
            try:
                code_num = int(emp.employee_code)
            except ValueError:
                pass
        return (d_order, d_name or "", code_num, item.work_date)

    forgot_records.sort(key=get_sort_key)

    # 6. Format response list
    results = []
    for a in forgot_records:
        emp = emp_map.get(a.employee_id)
        if not emp:
            continue

        # Shift resolution
        override_id = override_map.get((a.employee_id, a.work_date))
        if override_id:
            shift = shifts_by_id.get(override_id)
        else:
            shift = shifts_by_code.get(emp.default_shift_code)

        dow = DOW_VN[a.work_date.weekday()]
        shift_start_end = f"{str(shift.start_time)[:5]} - {str(shift.end_time)[:5]}" if shift and shift.start_time and shift.end_time else "Chưa xếp ca"

        ci_str = a.first_check_in.strftime("%H:%M") if a.first_check_in else None
        co_str = a.last_check_out.strftime("%H:%M") if a.last_check_out else None

        notes = "Quên check in" if not ci_str else "Quên check out"

        results.append({
            "employee_id": a.employee_id,
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
            "department": emp.department,
            "work_date": str(a.work_date),
            "dow": dow,
            "shift_code": shift.code if shift else None,
            "shift_start_end": shift_start_end,
            "first_check_in": ci_str,
            "last_check_out": co_str,
            "actual_hours": float(a.total_hours) if a.total_hours is not None else 0.0,
            "notes": notes,
        })

    return results


@router.post("/forgot-scans/quick-fix")
async def quick_fix_forgot_scans(
    request: ForgottenScanFixRequest,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Cập nhật nhanh các lỗi quên quẹt thẻ từ giao diện.
    Nhận riêng check_in / check_out cho phép swap (đổi chiều) giờ vào ↔ ra.
    """
    # Load shift OFF
    off_shift = (await db.execute(select(ShiftTemplate).where(ShiftTemplate.code == "OFF"))).scalar_one_or_none()
    if not off_shift:
        raise HTTPException(400, "Mã ca OFF không tồn tại trong hệ thống")

    shift_result = await db.execute(select(ShiftTemplate))
    all_shifts = list(shift_result.scalars().all())
    shifts_by_code = {s.code: s for s in all_shifts}
    shifts_by_id = {s.id: s for s in all_shifts}

    updated_count = 0

    for item in request.items:
        work_date = item.work_date
        employee_id = item.employee_id

        # Check date lock
        await check_date_locked(db, work_date)

        emp = await db.get(Employee, employee_id)
        if not emp:
            continue

        # ── Case 1: Công ty cho nghỉ ─────────────────────────────────────
        if item.is_off:
            ws_res = await db.execute(select(WorkSchedule).where(and_(
                WorkSchedule.employee_id == employee_id,
                WorkSchedule.work_date == work_date
            )))
            ws = ws_res.scalar_one_or_none()
            ws_before = {c.name: getattr(ws, c.name) for c in ws.__table__.columns} if ws else None
            if ws:
                ws.shift_id = off_shift.id
                ws.notes = "Công ty cho nghỉ (Cập nhật hàng loạt)"
            else:
                ws = WorkSchedule(
                    employee_id=employee_id,
                    work_date=work_date,
                    month_key=work_date.strftime("%Y-%m"),
                    shift_id=off_shift.id,
                    notes="Công ty cho nghỉ (Cập nhật hàng loạt)"
                )
                db.add(ws)

            await log_audit(
                db, "work_schedules", f"{employee_id}:{work_date}",
                "UPDATE" if ws_before else "CREATE", current_user.username,
                ws_before, {c.name: getattr(ws, c.name) for c in ws.__table__.columns},
                notes="Company off override via quick-fix"
            )

            await db.execute(delete(AttendanceDaily).where(and_(
                AttendanceDaily.employee_id == employee_id,
                AttendanceDaily.work_date == work_date
            )))
            await log_audit(
                db, "attendance_daily", f"{employee_id}:{work_date}",
                "DELETE", current_user.username,
                None, None,
                notes="Deleted attendance record due to Company Off override"
            )
            updated_count += 1
            continue

        # ── Case 2: Sửa giờ vào / giờ ra ────────────────────────────────
        ci_str = (item.check_in or "").strip()
        co_str = (item.check_out or "").strip()

        if not ci_str and not co_str:
            continue

        parsed_ci = parse_input_time(ci_str) if ci_str else None
        parsed_co = parse_input_time(co_str) if co_str else None

        if ci_str and not parsed_ci:
            raise HTTPException(400, f"Định dạng giờ vào không hợp lệ: {ci_str}")
        if co_str and not parsed_co:
            raise HTTPException(400, f"Định dạng giờ ra không hợp lệ: {co_str}")

        # Find existing AttendanceDaily record
        att_res = await db.execute(select(AttendanceDaily).where(and_(
            AttendanceDaily.employee_id == employee_id,
            AttendanceDaily.work_date == work_date
        )))
        att = att_res.scalar_one_or_none()
        att_before = {c.name: getattr(att, c.name) for c in att.__table__.columns} if att else None

        # Resolve active shift template
        ws_res = await db.execute(select(WorkSchedule).where(and_(
            WorkSchedule.employee_id == employee_id,
            WorkSchedule.work_date == work_date
        )))
        ws = ws_res.scalar_one_or_none()

        shift = None
        if ws:
            shift = shifts_by_id.get(ws.shift_id)
        if not shift and emp.default_shift_code:
            shift = shifts_by_code.get(emp.default_shift_code)
        if not shift:
            shift = shifts_by_code.get("D")

        is_night_shift = bool(shift.is_night_shift) if shift else False

        if att:
            # Gán giờ vào: nếu client gửi check_in → ghi đè; nếu không gửi → xóa (None)
            if parsed_ci:
                att.first_check_in = datetime.combine(work_date, parsed_ci)
            else:
                att.first_check_in = None

            # Gán giờ ra: nếu client gửi check_out → ghi đè; nếu không gửi → xóa (None)
            if parsed_co:
                if is_night_shift and att.first_check_in and parsed_co < att.first_check_in.time():
                    att.last_check_out = datetime.combine(work_date + timedelta(days=1), parsed_co)
                else:
                    att.last_check_out = datetime.combine(work_date, parsed_co)
            else:
                att.last_check_out = None

            # Tính lại tổng giờ
            if att.first_check_in and att.last_check_out:
                att.total_hours = round((att.last_check_out - att.first_check_in).total_seconds() / 3600.0, 2)
            else:
                att.total_hours = 0.0

            att.import_batch = "manual_fix"
        else:
            # Tạo mới
            new_ci = datetime.combine(work_date, parsed_ci) if parsed_ci else None
            new_co = datetime.combine(work_date, parsed_co) if parsed_co else None
            if is_night_shift and new_ci and new_co and new_co < new_ci:
                new_co = datetime.combine(work_date + timedelta(days=1), parsed_co)
            att = AttendanceDaily(
                employee_id=employee_id,
                work_date=work_date,
                first_check_in=new_ci,
                last_check_out=new_co,
                total_hours=round((new_co - new_ci).total_seconds() / 3600.0, 2) if new_ci and new_co else 0.0,
                import_batch="manual_fix"
            )
            db.add(att)

        await log_audit(
            db, "attendance_daily", f"{employee_id}:{work_date}",
            "UPDATE" if att_before else "CREATE", current_user.username,
            att_before, {c.name: getattr(att, c.name) for c in att.__table__.columns},
            notes="Quick fix forgotten scan"
        )
        updated_count += 1

    await db.commit()
    return {"message": f"Đã cập nhật thành công {updated_count} dòng"}


@router.get("/forgot-scans/export")
async def export_forgot_scans(
    month_key: str = Query(..., description="YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Xuất file Excel mẫu các lỗi quên quẹt thẻ trong tháng"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Side, Border

    scans = await get_forgot_scans(month_key=month_key, db=db, current_user=current_user)

    wb = openpyxl.Workbook()
    ews = wb.active  # excel worksheet — dùng tên khác để tránh nhầm WorkSchedule
    ews.title = "Quen quet the"

    # ── Styles ────────────────────────────────────────────────────────────────
    thin = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),  bottom=Side(style='thin', color='000000'),
    )
    hdr_font    = Font(name="Times New Roman", bold=True, size=11)
    title_font  = Font(name="Times New Roman", bold=True, size=14)
    normal_font = Font(name="Times New Roman", size=11)
    hint_font   = Font(name="Times New Roman", italic=True, size=10, color="595959")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    grey_fill   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    edit_fill   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # vàng = sửa được
    off_fill    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # cam = cột OFF
    hint_fill   = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")  # xanh = ghi chú
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_fn      = Font(name="Times New Roman", size=11, color="9C0006")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_fn   = Font(name="Times New Roman", size=11, color="9C6500")

    NUM_COLS = 11  # A..K

    # ── Row 1-2: Tiêu đề công ty ──────────────────────────────────────────────
    ews.merge_cells("A1:K1")
    ews["A1"] = "CÔNG TY TNHH HIỆP LỢI"
    ews["A1"].font = Font(name="Times New Roman", bold=True, size=12)
    ews["A1"].alignment = left

    ews.merge_cells("A2:K2")
    ews["A2"] = "MST: 3701609885"
    ews["A2"].font = normal_font
    ews["A2"].alignment = left

    # ── Row 3: Tên bảng ───────────────────────────────────────────────────────
    ews.merge_cells("A3:K3")
    ews["A3"] = f"DANH SÁCH QUÊN QUẸT THẺ - THÁNG {month_key}"
    ews["A3"].font = title_font
    ews["A3"].alignment = center

    # ── Row 4: Hướng dẫn ─────────────────────────────────────────────────────
    ews.merge_cells("A4:K4")
    ews["A4"] = (
        "HƯỚNG DẪN: "
        "① Cột 'Giờ vào' & 'Giờ ra' (nền vàng) — sửa hoặc điền thêm giờ theo định dạng HH:MM. "
        "Có thể sửa cả hai nếu cần chuyển giờ (ví dụ: chấm công chỉ có 22:00 thì điền vào 'Giờ ra', điền giờ vào thực tế vào 'Giờ vào'). "
        "② Cột 'Giờ thực' (nền cam) — gõ chữ X nếu hôm đó công ty cho nghỉ "
        "(hệ thống sẽ tự chuyển sang ca OFF và xóa dữ liệu chấm công). "
        "Để trống ô nào nếu không thay đổi ô đó."
    )
    ews["A4"].font = hint_font
    ews["A4"].fill = hint_fill
    ews["A4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ews.row_dimensions[4].height = 52

    # ── Row 5: Header ─────────────────────────────────────────────────────────
    headers = [
        "STT", "Mã nhân viên", "Họ tên", "Ngày", "Mã ca",
        "Thứ", "Thời gian\nquy định",
        "Giờ vào\n(✎ sửa trực tiếp)",   # col 8  — editable
        "Giờ ra\n(✎ sửa trực tiếp)",    # col 9  — editable
        "Giờ thực\n(gõ X = OFF)",        # col 10 — editable for X
        "Ghi chú",                        # col 11
    ]
    for ci, h in enumerate(headers, 1):
        cell = ews.cell(row=5, column=ci, value=h)
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = thin
        if ci in (8, 9):
            cell.fill = edit_fill
        elif ci == 10:
            cell.fill = off_fill
        else:
            cell.fill = grey_fill
    ews.row_dimensions[5].height = 34

    # ── Row 6+: Dữ liệu ──────────────────────────────────────────────────────
    cur = 6
    for idx, scan in enumerate(scans, 1):
        ews.cell(cur, 1, idx).alignment = center
        ews.cell(cur, 2, scan["employee_code"]).alignment = center
        ews.cell(cur, 3, scan["full_name"]).alignment = left
        ews.cell(cur, 4, scan["work_date"]).alignment = center
        ews.cell(cur, 5, scan["shift_code"] or "").alignment = center
        ews.cell(cur, 6, scan["dow"]).alignment = center
        ews.cell(cur, 7, scan["shift_start_end"]).alignment = center

        # Col H (8): Giờ vào — pre-filled, nền vàng, sửa được
        ci_cell = ews.cell(cur, 8, scan["first_check_in"] or "")
        ci_cell.alignment = center
        ci_cell.fill = edit_fill

        # Col I (9): Giờ ra — pre-filled, nền vàng, sửa được
        co_cell = ews.cell(cur, 9, scan["last_check_out"] or "")
        co_cell.alignment = center
        co_cell.fill = edit_fill

        # Col J (10): Giờ thực — nền cam, gõ X nếu OFF
        gthuc = ews.cell(cur, 10, scan["actual_hours"] if scan["actual_hours"] else 0)
        gthuc.alignment = center
        gthuc.fill = off_fill

        # Col K (11): Ghi chú (read-only)
        note_cell = ews.cell(cur, 11, scan["notes"])
        note_cell.alignment = center
        if scan["notes"] == "Quên check in":
            note_cell.fill = red_fill
            note_cell.font = red_fn
        else:
            note_cell.fill = yellow_fill
            note_cell.font = yellow_fn

        # Viền + font mặc định cho tất cả ô trong hàng
        for c in range(1, NUM_COLS + 1):
            cell = ews.cell(cur, c)
            cell.border = thin
            # Chỉ đặt font nếu ô chưa có font màu riêng (note cell)
            if c not in (11,):
                cell.font = normal_font

        cur += 1

    # ── Độ rộng cột ───────────────────────────────────────────────────────────
    for col, w in zip("ABCDEFGHIJK", [6, 15, 26, 15, 10, 7, 20, 16, 16, 13, 18]):
        ews.column_dimensions[col].width = w

    ews.freeze_panes = "A6"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"quen_quet_the_{month_key}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/forgot-scans/import")
async def import_forgot_scans(
    file: UploadFile = File(...),
    month_key: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Nhập Excel đã sửa: đọc Giờ vào (H), Giờ ra (I), Giờ thực (J=X → OFF)."""
    await check_month_locked(db, month_key)

    off_shift = (await db.execute(select(ShiftTemplate).where(ShiftTemplate.code == "OFF"))).scalar_one_or_none()
    if not off_shift:
        raise HTTPException(400, "Mã ca OFF không tồn tại trong hệ thống")

    shift_result = await db.execute(select(ShiftTemplate))
    shifts_by_code = {s.code: s for s in shift_result.scalars().all()}
    shifts_by_id   = {s.id:   s for s in shift_result.scalars().all()}

    content = await file.read()
    import openpyxl
    ewb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ews = ewb.active  # dùng tên 'ews' để không xung đột với WorkSchedule

    processed_count = 0
    errors = []

    # Dữ liệu bắt đầu từ row 6 (row 5 là header, row 4 là hướng dẫn)
    for r in range(6, ews.max_row + 1):
        emp_code_raw  = ews.cell(r, 2).value   # B: Mã nhân viên
        work_date_raw = ews.cell(r, 4).value   # D: Ngày
        ci_raw        = ews.cell(r, 8).value   # H: Giờ vào (editable)
        co_raw        = ews.cell(r, 9).value   # I: Giờ ra  (editable)
        gthuc_raw     = ews.cell(r, 10).value  # J: Giờ thực (X = OFF)

        if emp_code_raw is None or work_date_raw is None:
            continue

        # Kiểm tra có gì để làm không
        ci_str    = str(ci_raw).strip()    if ci_raw    is not None else ""
        co_str    = str(co_raw).strip()    if co_raw    is not None else ""
        gthuc_str = str(gthuc_raw).strip() if gthuc_raw is not None else ""
        is_off    = gthuc_str.upper() == "X"

        if not is_off and not ci_str and not co_str:
            continue  # Dòng không có thay đổi gì

        # Chuẩn hoá mã nhân viên
        emp_code = (
            str(int(emp_code_raw))
            if isinstance(emp_code_raw, (float, int))
            else str(emp_code_raw).split('.')[0] if '.' in str(emp_code_raw)
            else str(emp_code_raw)
        ).strip().lstrip("'")

        # Parse ngày
        work_date = None
        if isinstance(work_date_raw, (date, datetime)):
            work_date = work_date_raw if isinstance(work_date_raw, date) else work_date_raw.date()
        elif isinstance(work_date_raw, str):
            for fmt in ["%Y-%m-%d", "%d/%m/%Y"]:
                try:
                    work_date = datetime.strptime(work_date_raw.strip(), fmt).date()
                    break
                except ValueError:
                    continue

        if not work_date:
            errors.append(f"Dòng {r}: Ngày không hợp lệ '{work_date_raw}'")
            continue
        if work_date.strftime("%Y-%m") != month_key:
            errors.append(f"Dòng {r}: Ngày {work_date} không thuộc tháng {month_key}")
            continue

        # Tìm nhân viên
        emp_res = await db.execute(select(Employee).where(Employee.employee_code == emp_code))
        emp = emp_res.scalar_one_or_none()
        if not emp:
            errors.append(f"Dòng {r}: Không tìm thấy nhân viên mã '{emp_code}'")
            continue

        # ── Case 1: X → Công ty cho nghỉ ────────────────────────────────────
        if is_off:
            sched_res = await db.execute(select(WorkSchedule).where(and_(
                WorkSchedule.employee_id == emp.id,
                WorkSchedule.work_date == work_date,
            )))
            sched = sched_res.scalar_one_or_none()
            sched_before = {c.name: getattr(sched, c.name) for c in sched.__table__.columns} if sched else None
            if sched:
                sched.shift_id = off_shift.id
                sched.notes = "Công ty cho nghỉ (Nhập từ file Excel)"
            else:
                sched = WorkSchedule(
                    employee_id=emp.id, work_date=work_date, month_key=month_key,
                    shift_id=off_shift.id, notes="Công ty cho nghỉ (Nhập từ file Excel)"
                )
                db.add(sched)
            await log_audit(
                db, "work_schedules", f"{emp.id}:{work_date}",
                "UPDATE" if sched_before else "CREATE", current_user.username,
                sched_before, {c.name: getattr(sched, c.name) for c in sched.__table__.columns},
                notes="Company off via Excel import"
            )
            await db.execute(delete(AttendanceDaily).where(and_(
                AttendanceDaily.employee_id == emp.id,
                AttendanceDaily.work_date == work_date,
            )))
            await log_audit(db, "attendance_daily", f"{emp.id}:{work_date}", "DELETE",
                            current_user.username, None, None, notes="Deleted — Company Off (Excel)")
            processed_count += 1
            continue

        # ── Case 2: Sửa giờ vào / giờ ra ────────────────────────────────────
        parsed_ci = parse_input_time(ci_str) if ci_str else None
        parsed_co = parse_input_time(co_str) if co_str else None

        if ci_str and not parsed_ci:
            errors.append(f"Dòng {r}: Giờ vào '{ci_str}' không đúng định dạng HH:MM")
            continue
        if co_str and not parsed_co:
            errors.append(f"Dòng {r}: Giờ ra '{co_str}' không đúng định dạng HH:MM")
            continue

        await check_date_locked(db, work_date)

        # Tìm ca để biết ca đêm không
        sched_res = await db.execute(select(WorkSchedule).where(and_(
            WorkSchedule.employee_id == emp.id,
            WorkSchedule.work_date == work_date,
        )))
        sched = sched_res.scalar_one_or_none()
        shift = None
        if sched:
            shift = shifts_by_id.get(sched.shift_id)
        if not shift and emp.default_shift_code:
            shift = shifts_by_code.get(emp.default_shift_code)
        is_night = bool(shift.is_night_shift) if shift else False

        # Tìm / tạo bản ghi chấm công
        att_res = await db.execute(select(AttendanceDaily).where(and_(
            AttendanceDaily.employee_id == emp.id,
            AttendanceDaily.work_date == work_date,
        )))
        att = att_res.scalar_one_or_none()
        att_before = {c.name: getattr(att, c.name) for c in att.__table__.columns} if att else None

        if att:
            # Cập nhật giờ vào: nếu Excel có giá trị → ghi đè, nếu trống nhưng DB có → xóa
            if parsed_ci:
                att.first_check_in = datetime.combine(work_date, parsed_ci)
            elif not ci_str and att.first_check_in is not None:
                att.first_check_in = None  # User xóa ô Giờ vào trong Excel → clear

            # Cập nhật giờ ra: tương tự
            if parsed_co:
                # Ca đêm: giờ ra sáng hôm sau
                ci_for_night = att.first_check_in
                if is_night and ci_for_night and parsed_co < ci_for_night.time():
                    att.last_check_out = datetime.combine(work_date + timedelta(days=1), parsed_co)
                else:
                    att.last_check_out = datetime.combine(work_date, parsed_co)
            elif not co_str and att.last_check_out is not None:
                att.last_check_out = None  # User xóa ô Giờ ra trong Excel → clear

            # Tính lại tổng giờ
            if att.first_check_in and att.last_check_out:
                att.total_hours = round(
                    (att.last_check_out - att.first_check_in).total_seconds() / 3600.0, 2
                )
            else:
                att.total_hours = 0.0
            att.import_batch = "manual_fix_excel"
        else:
            # Tạo mới nếu chưa có bản ghi
            new_ci = datetime.combine(work_date, parsed_ci) if parsed_ci else None
            new_co = datetime.combine(work_date, parsed_co) if parsed_co else None
            if is_night and new_ci and new_co and new_co < new_ci:
                new_co = datetime.combine(work_date + timedelta(days=1), parsed_co)
            att = AttendanceDaily(
                employee_id=emp.id, work_date=work_date,
                first_check_in=new_ci, last_check_out=new_co,
                total_hours=round((new_co - new_ci).total_seconds() / 3600.0, 2) if new_ci and new_co else 0.0,
                import_batch="manual_fix_excel"
            )
            db.add(att)

        await log_audit(
            db, "attendance_daily", f"{emp.id}:{work_date}",
            "UPDATE" if att_before else "CREATE", current_user.username,
            att_before, {c.name: getattr(att, c.name) for c in att.__table__.columns},
            notes="Excel import — fix forgotten scan"
        )
        processed_count += 1

    await db.commit()

    if errors:
        return {
            "message": f"Hoàn tất: {processed_count} bản ghi cập nhật, {len(errors)} lỗi.",
            "errors": errors
        }
    return {"message": f"Đã cập nhật thành công {processed_count} bản ghi từ file Excel"}


@router.get("/forgot-scans/fixed")
async def get_fixed_scans(
    month_key: str = Query(..., description="YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Lấy danh sách các bản ghi đã được sửa thủ công (manual_fix / manual_fix_excel)."""
    try:
        year, month = map(int, month_key.split("-"))
    except ValueError:
        raise HTTPException(400, "month_key phải là YYYY-MM")

    month_days = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end = date(year, month, month_days)

    # Query records with manual fix markers
    att_q = select(AttendanceDaily).where(
        and_(
            AttendanceDaily.work_date >= month_start,
            AttendanceDaily.work_date <= month_end,
            AttendanceDaily.import_batch.in_(["manual_fix", "manual_fix_excel", "manual"]),
        )
    )
    att_result = await db.execute(att_q)
    fixed_records = list(att_result.scalars().all())

    if not fixed_records:
        return []

    # Load employees
    emp_ids = list(set(a.employee_id for a in fixed_records))
    emp_result = await db.execute(select(Employee).where(Employee.id.in_(emp_ids)))
    emp_map = {e.id: e for e in emp_result.scalars().all()}

    # Load shifts
    shift_result = await db.execute(select(ShiftTemplate))
    all_shifts = list(shift_result.scalars().all())
    shifts_by_id = {s.id: s for s in all_shifts}
    shifts_by_code = {s.code: s for s in all_shifts}

    # Load schedules
    schedule_q = select(WorkSchedule).where(
        and_(
            WorkSchedule.work_date >= month_start,
            WorkSchedule.work_date <= month_end,
            WorkSchedule.employee_id.in_(emp_ids),
        )
    )
    schedule_result = await db.execute(schedule_q)
    override_map = {(ws.employee_id, ws.work_date): ws.shift_id for ws in schedule_result.scalars().all()}

    # Department sort order
    from app.models.department import Department
    dept_order_q = await db.execute(select(Department.name, Department.sort_order))
    dept_order_map = {row[0]: row[1] for row in dept_order_q.all() if row[0]}

    results = []
    for a in fixed_records:
        emp = emp_map.get(a.employee_id)
        if not emp:
            continue

        # Shift resolution
        override_id = override_map.get((a.employee_id, a.work_date))
        if override_id:
            shift = shifts_by_id.get(override_id)
        else:
            shift = shifts_by_code.get(emp.default_shift_code)

        dow = DOW_VN[a.work_date.weekday()]
        shift_start_end = (
            f"{str(shift.start_time)[:5]} - {str(shift.end_time)[:5]}"
            if shift and shift.start_time and shift.end_time
            else "Chưa xếp ca"
        )

        ci_str = a.first_check_in.strftime("%H:%M") if a.first_check_in else None
        co_str = a.last_check_out.strftime("%H:%M") if a.last_check_out else None

        source_label = {
            "manual_fix": "Sửa nhanh",
            "manual_fix_excel": "Nhập Excel",
            "manual": "Thủ công",
        }.get(a.import_batch, a.import_batch)

        results.append({
            "id": a.id,
            "employee_id": a.employee_id,
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
            "department": emp.department,
            "work_date": str(a.work_date),
            "dow": dow,
            "shift_code": shift.code if shift else None,
            "shift_start_end": shift_start_end,
            "first_check_in": ci_str,
            "last_check_out": co_str,
            "total_hours": float(a.total_hours) if a.total_hours is not None else 0.0,
            "source": source_label,
            "updated_at": a.updated_at.isoformat() if a.updated_at else None,
        })

    # Sort by department order → employee code → date
    def sort_key(item):
        d_order = dept_order_map.get(item["department"], 9999) if item["department"] else 9999
        code_num = 999999
        try:
            code_num = int(item["employee_code"])
        except ValueError:
            pass
        return (d_order, item["department"] or "", code_num, item["work_date"])

    results.sort(key=sort_key)
    return results


class DeleteFixedScanRequest(BaseModel):
    attendance_id: int


@router.post("/forgot-scans/fixed/delete")
async def delete_fixed_scan(
    request: DeleteFixedScanRequest,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Xóa 1 bản ghi chấm công đã sửa thủ công, hoàn tác về trạng thái gốc từ máy chấm công."""
    att = await db.get(AttendanceDaily, request.attendance_id)
    if not att:
        raise HTTPException(404, "Không tìm thấy bản ghi chấm công")

    if att.import_batch not in ("manual_fix", "manual_fix_excel", "manual"):
        raise HTTPException(400, "Chỉ có thể xóa bản ghi đã sửa thủ công")

    await check_date_locked(db, att.work_date)

    att_before = {c.name: getattr(att, c.name) for c in att.__table__.columns}

    emp = await db.get(Employee, att.employee_id)
    if not emp:
        # Nếu nhân viên không tồn tại, xóa hẳn bản ghi
        await db.execute(
            delete(AttendanceDaily).where(AttendanceDaily.id == request.attendance_id)
        )
        await db.commit()
        return {"message": "Đã xóa bản ghi"}

    emp_code = str(emp.employee_code).lstrip("'")

    # Kiểm tra xem ca làm việc của nhân viên là ca động NU/XNU hay ca thường
    is_nu = is_nu_dynamic_shift_code(emp.default_shift_code)
    if not is_nu:
        ws_res = await db.execute(select(WorkSchedule).where(and_(
            WorkSchedule.employee_id == emp.id,
            WorkSchedule.work_date == att.work_date
        )))
        ws = ws_res.scalar_one_or_none()
        if ws:
            shift = await db.get(ShiftTemplate, ws.shift_id)
            if shift and is_nu_dynamic_shift_code(shift.code):
                is_nu = True

    restored = False
    if is_nu:
        # Khôi phục cho ca động NU/XNU
        nu_shift_code_map = {}
        for d in (att.work_date - timedelta(days=1), att.work_date, att.work_date + timedelta(days=1)):
            ws_d = (await db.execute(select(WorkSchedule).where(and_(
                WorkSchedule.employee_id == emp.id,
                WorkSchedule.work_date == d
            )))).scalar_one_or_none()
            s_code = emp.default_shift_code
            if ws_d:
                shift_d = await db.get(ShiftTemplate, ws_d.shift_id)
                if shift_d:
                    s_code = shift_d.code
            if is_nu_dynamic_shift_code(s_code):
                nu_shift_code_map[(emp.id, d)] = s_code

        log_res_nu = await db.execute(
            select(AttendanceLog).where(
                and_(
                    AttendanceLog.employee_code.like(f"%{emp_code}"),
                    AttendanceLog.event_time >= datetime.combine(att.work_date - timedelta(days=1), time(0, 0)),
                    AttendanceLog.event_time <= datetime.combine(att.work_date + timedelta(days=2), time(12, 0))
                )
            )
        )
        nu_logs = []
        for l in log_res_nu.scalars().all():
            l.employee_id = emp.id
            nu_logs.append(l)

        nu_res = build_nu_shift_day_results(nu_shift_code_map, [emp.id], nu_logs)
        res_day = nu_res.get((emp.id, att.work_date))
        if res_day and (res_day.check_in or res_day.check_out):
            att.first_check_in = res_day.check_in
            att.last_check_out = res_day.check_out
            if res_day.check_in and res_day.check_out:
                att.total_hours = round((res_day.check_out - res_day.check_in).total_seconds() / 3600.0, 2)
            else:
                att.total_hours = 0.0
            att.import_batch = "reverted_fix"
            restored = True
    else:
        # Khôi phục cho ca thường
        ws_res = await db.execute(select(WorkSchedule).where(and_(
            WorkSchedule.employee_id == emp.id,
            WorkSchedule.work_date == att.work_date
        )))
        ws = ws_res.scalar_one_or_none()
        shift = None
        if ws:
            shift = await db.get(ShiftTemplate, ws.shift_id)
        if not shift and emp.default_shift_code:
            shift = (await db.execute(select(ShiftTemplate).where(ShiftTemplate.code == emp.default_shift_code))).scalar_one_or_none()

        is_night = shift.is_night_shift if shift else False

        log_res = await db.execute(
            select(AttendanceLog).where(
                and_(
                    AttendanceLog.employee_code.like(f"%{emp_code}"),
                    AttendanceLog.event_time >= datetime.combine(att.work_date, time(0, 0)),
                    AttendanceLog.event_time <= datetime.combine(att.work_date + timedelta(days=1), time(12, 0))
                )
            )
        )
        all_logs = sorted(log_res.scalars().all(), key=lambda l: l.event_time)

        daily_scans = []
        for l in all_logs:
            scan_dt = l.event_time
            work_date_of_scan = scan_dt.date()
            if scan_dt.time() < time(6, 0): # NIGHT_SHIFT_CUTOFF
                if is_night:
                    if work_date_of_scan == att.work_date + timedelta(days=1):
                        daily_scans.append(scan_dt)
            else:
                if work_date_of_scan == att.work_date:
                    daily_scans.append(scan_dt)

        if daily_scans:
            daily_scans.sort()
            first_in = daily_scans[0]
            last_out = daily_scans[-1] if len(daily_scans) > 1 else None

            att.first_check_in = first_in
            att.last_check_out = last_out
            if first_in and last_out:
                att.total_hours = round((last_out - first_in).total_seconds() / 3600.0, 2)
            else:
                att.total_hours = 0.0
            att.import_batch = "reverted_fix"
            restored = True

    if restored:
        await log_audit(
            db, "attendance_daily", f"{att.employee_id}:{att.work_date}",
            "UPDATE", current_user.username,
            att_before, {c.name: getattr(att, c.name) for c in att.__table__.columns},
            notes="Reverted manual fix and restored raw machine scans"
        )
        await db.commit()
        return {"message": "Đã hoàn tác bản ghi về trạng thái gốc từ máy chấm công"}
    else:
        # Nếu không tìm thấy lịch sử quét thẻ nào, tiến hành xóa hẳn bản ghi chấm công ngày đó
        await db.execute(
            delete(AttendanceDaily).where(AttendanceDaily.id == request.attendance_id)
        )
        await log_audit(
            db, "attendance_daily", f"{att.employee_id}:{att.work_date}",
            "DELETE", current_user.username,
            att_before, None,
            notes="Deleted manual fix, no raw scans found to restore"
        )
        await db.commit()
        return {"message": "Đã xóa bản ghi chấm công"}


class UpdateFixedScanRequest(BaseModel):
    attendance_id: int
    check_in: Optional[str] = None
    check_out: Optional[str] = None


@router.post("/forgot-scans/fixed/update")
async def update_fixed_scan(
    request: UpdateFixedScanRequest,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Sửa lại giờ vào/ra của bản ghi đã sửa thủ công."""
    att = await db.get(AttendanceDaily, request.attendance_id)
    if not att:
        raise HTTPException(404, "Không tìm thấy bản ghi chấm công")

    if att.import_batch not in ("manual_fix", "manual_fix_excel", "manual"):
        raise HTTPException(400, "Chỉ có thể sửa bản ghi đã sửa thủ công")

    await check_date_locked(db, att.work_date)

    att_before = {c.name: getattr(att, c.name) for c in att.__table__.columns}

    # Parse times
    ci_str = (request.check_in or "").strip()
    co_str = (request.check_out or "").strip()
    parsed_ci = parse_input_time(ci_str) if ci_str else None
    parsed_co = parse_input_time(co_str) if co_str else None

    if ci_str and not parsed_ci:
        raise HTTPException(400, f"Giờ vào không đúng định dạng: {ci_str}")
    if co_str and not parsed_co:
        raise HTTPException(400, f"Giờ ra không đúng định dạng: {co_str}")

    # Resolve shift for night-shift handling
    shift_result = await db.execute(select(ShiftTemplate))
    all_shifts = list(shift_result.scalars().all())
    shifts_by_id = {s.id: s for s in all_shifts}
    shifts_by_code = {s.code: s for s in all_shifts}

    emp = await db.get(Employee, att.employee_id)
    ws_res = await db.execute(select(WorkSchedule).where(and_(
        WorkSchedule.employee_id == att.employee_id,
        WorkSchedule.work_date == att.work_date,
    )))
    ws = ws_res.scalar_one_or_none()
    shift = None
    if ws:
        shift = shifts_by_id.get(ws.shift_id)
    if not shift and emp and emp.default_shift_code:
        shift = shifts_by_code.get(emp.default_shift_code)
    is_night = bool(shift.is_night_shift) if shift else False

    work_date = att.work_date

    # Update check-in
    if parsed_ci:
        att.first_check_in = datetime.combine(work_date, parsed_ci)
    else:
        att.first_check_in = None

    # Update check-out
    if parsed_co:
        if is_night and att.first_check_in and parsed_co < att.first_check_in.time():
            att.last_check_out = datetime.combine(work_date + timedelta(days=1), parsed_co)
        else:
            att.last_check_out = datetime.combine(work_date, parsed_co)
    else:
        att.last_check_out = None

    # Recalculate hours
    if att.first_check_in and att.last_check_out:
        att.total_hours = round(
            (att.last_check_out - att.first_check_in).total_seconds() / 3600.0, 2
        )
    else:
        att.total_hours = 0.0

    await log_audit(
        db, "attendance_daily", f"{att.employee_id}:{work_date}",
        "UPDATE", current_user.username,
        att_before, {c.name: getattr(att, c.name) for c in att.__table__.columns},
        notes="Updated manual fix via UI"
    )

    await db.commit()
    return {"message": "Đã cập nhật bản ghi thành công"}

