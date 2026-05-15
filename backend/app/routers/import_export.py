from collections import defaultdict
from typing import Optional
from datetime import datetime, date, time, timedelta
import calendar
import json
import uuid
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, text, delete
from app.database import get_db
from app.models.attendance import AttendanceLog, AttendanceDaily
from app.models.employee import Employee
from app.models.shift import ShiftTemplate
from app.models.schedule import WorkSchedule
from app.models.user import AppUser, UserRole
from app.middleware.auth import require_roles
from app.utils.audit_helper import log_audit
from app.services.nu_shift import is_nu_dynamic_shift_code, build_nu_shift_day_results

router = APIRouter(prefix="/import-export", tags=["Import/Export"])

NIGHT_SHIFT_CUTOFF = time(6, 0)  # Gio truoc 6h sang tinh la ca dem hom truoc


@router.post("/attendance")
async def import_attendance(
    file: UploadFile = File(...),
    month_key: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Import file cham cong tu may cham cong.
    Format: C1=Ma NV, C2=Ten, C3=Bo phan, C4=Thoi gian (datetime)
    Bo phan trong file se bo qua, lay theo nhan vien trong DB.
    Se xu ly: loai trung, nhom theo ngay, tinh first_check_in/last_check_out, luu vao attendance_daily.
    Ca dem: scan truoc 6h sang => tinh la ca hom truoc."""
    try:
        content = await file.read()
        filename = file.filename.lower()
        
        # 0. Detect file type and parse data
        # Standard format: Column 1=Code, 2=Name, 3=Dept, 4=Time
        data_rows = []
        
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            ws = wb[wb.sheetnames[0]]
            for r in range(2, ws.max_row + 1):
                c1 = ws.cell(r, 1).value
                c2 = ws.cell(r, 2).value
                c3 = ws.cell(r, 3).value
                c4 = ws.cell(r, 4).value
                if c1 is not None:
                    data_rows.append((c1, c2, c3, c4))
        elif filename.endswith(".csv"):
            import csv
            import io
            # Try different encodings
            text_content = ""
            for encoding in ["utf-8-sig", "utf-8", "latin1", "cp1252"]:
                try:
                    text_content = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if not text_content:
                raise HTTPException(400, "Khong the doc file CSV (sai encoding)")
                
            # Try different delimiters
            dialect = None
            try:
                # Use a larger sample for sniffer
                dialect = csv.Sniffer().sniff(text_content[:2048])
            except:
                pass
            
            reader = csv.reader(io.StringIO(text_content), dialect=dialect or "excel")
            rows = list(reader)
            # Skip header if needed (assume header exists if first col is not numeric)
            start_idx = 0
            if rows and len(rows[0]) > 0 and not str(rows[0][0]).isdigit():
                start_idx = 1
                
            for r in rows[start_idx:]:
                if not r: continue
                if len(r) >= 4:
                    data_rows.append((r[0], r[1], r[2], r[3]))
                elif len(r) >= 2: # Min requirement
                    data_rows.append((r[0], r[1], None, r[2] if len(r) > 2 else None))
        else:
            raise HTTPException(400, "Dinh dang file khong duoc ho tro. Vui long dung .xlsx hoặc .csv")

        batch_id = str(uuid.uuid4())[:8]

        # 1. Thu thap thong tin nhan vien tu du lieu
        file_employees = {} # code -> name
        for c_raw, n_raw, _, _ in data_rows:
            if c_raw is not None and n_raw is not None:
                c = str(int(c_raw) if isinstance(c_raw, (float, int)) else str(c_raw).split('.')[0] if '.' in str(c_raw) else c_raw).strip().lstrip("'")
                file_employees[c] = str(n_raw).strip()

        # 2. Dong bo bang Employee
        all_emps = (await db.execute(select(Employee))).scalars().all()
        active_emps = [e for e in all_emps if e.is_active]
        codes_in_file = set(file_employees.keys())
        
        # Xac dinh ngay moc (dau thang duoc chon)
        try:
            y, m = map(int, month_key.split("-"))
            ref_date = date(y, m, 1)
        except:
            ref_date = date.today().replace(day=1)

        # A. Cho nghi viec nhung nguoi KHONG co trong file
        for emp in active_emps:
            if str(emp.employee_code) not in codes_in_file:
                emp.is_active = False
                emp.leave_date = ref_date
                
        # B. Xu ly nhan vien trong file (moi hoac cap nhat)
        for code, name in file_employees.items():
            # Tim theo code (uu tien active)
            emp = next((e for e in active_emps if str(e.employee_code).lstrip("'") == code), None)
            if emp:
                if emp.full_name != name:
                    emp.full_name = name
                continue
                
            emp_any = next((e for e in all_emps if str(e.employee_code).lstrip("'") == code), None)
            if emp_any:
                emp_any.is_active = True
                emp_any.leave_date = None
                emp_any.full_name = name
            else:
                new_emp = Employee(
                    employee_code=code,
                    full_name=name,
                    is_active=True,
                    join_date=ref_date
                )
                db.add(new_emp)

        await db.commit()

        # 2.5 Clear old data for this month
        y, m = map(int, month_key.split("-"))
        days_in_month = calendar.monthrange(y, m)[1]
        month_start = date(y, m, 1)
        month_end = date(y, m, days_in_month)
        
        # Delete Daily records
        await db.execute(delete(AttendanceDaily).where(and_(
            AttendanceDaily.work_date >= month_start,
            AttendanceDaily.work_date <= month_end
        )))
        
        # Delete Raw Logs
        # Include a bit of buffer for night shifts (up to 12:00 on the 1st of the next month)
        log_start = datetime.combine(month_start, time(0, 0))
        log_end = datetime.combine(month_end + timedelta(days=1), time(12, 0))
        await db.execute(delete(AttendanceLog).where(and_(
            AttendanceLog.event_time >= log_start,
            AttendanceLog.event_time <= log_end
        )))
        
        await db.commit()

        # Load lai employee code mapping
        emp_result = await db.execute(select(Employee).where(Employee.is_active == True))
        emp_map = {str(e.employee_code).lstrip("'"): e for e in emp_result.scalars().all()}

        # Load shifts
        shift_result = await db.execute(select(ShiftTemplate))
        shifts_by_code = {s.code: s for s in shift_result.scalars().all()}
        shifts_by_id = {s.id: s for s in shift_result.scalars().all()}

        # Parse raw scans
        raw_scans = {}  # emp_code -> set of datetimes
        skipped_employees = set()
        total_rows = 0

        for emp_code_raw, emp_name, _, scan_time in data_rows:
            if emp_code_raw is None:
                continue

            emp_code = str(int(emp_code_raw) if isinstance(emp_code_raw, (float, int)) else str(emp_code_raw).split('.')[0] if '.' in str(emp_code_raw) else emp_code_raw).strip().lstrip("'")

            if emp_code not in emp_map:
                skipped_employees.add(emp_code)
                continue

            emp = emp_map[emp_code]

            if scan_time is None:
                continue

            if isinstance(scan_time, str):
                # Try various formats
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"]:
                    try:
                        scan_time = datetime.strptime(scan_time, fmt)
                        break
                    except ValueError:
                        continue

            if not isinstance(scan_time, datetime):
                continue

            total_rows += 1
            if emp_code not in raw_scans:
                raw_scans[emp_code] = set()
            raw_scans[emp_code].add(scan_time)

            log = AttendanceLog(
                employee_code=emp_code,
                employee_name=str(emp_name or ""),
                department=emp.department,
                event_time=scan_time,
                source_file=file.filename,
                import_batch=batch_id,
            )
            db.add(log)

        # 1.5 Validate months in file against month_key
        # We check if the majority of scans or at least a significant portion match the month_key
        # For simplicity and according to user request: if month_key is not in the set of months found, error out.
        file_months = set()
        for emp_code, scans in raw_scans.items():
            for s_dt in scans:
                file_months.add(s_dt.strftime("%Y-%m"))
        
        if file_months and month_key not in file_months:
            sorted_months = sorted(list(file_months))
            raise HTTPException(
                status_code=400, 
                detail=f"Dữ liệu trong file thuộc tháng {', '.join(sorted_months)}, không khớp với tháng {month_key} đã chọn. Vui lòng kiểm tra lại!"
            )

        # Process: group by employee + work_date
        # For night shifts: if scan is before NIGHT_SHIFT_CUTOFF (6am), it belongs to previous day
        processed = 0
        updated = 0

        # Load schedule data for night shift detection
        schedule_result = await db.execute(select(WorkSchedule))
        schedule_map = {}
        for ws_rec in schedule_result.scalars().all():
            schedule_map[(ws_rec.employee_id, ws_rec.work_date)] = ws_rec.shift_id

        # Grouping scans by work_date
        # For non-NU shifts, use standard 6am cutoff.
        # For NU shifts, use build_nu_shift_day_results logic.
        
        # 1. Identify NU employees and their potential dates
        nu_emp_ids = []
        for emp_code, scans in raw_scans.items():
            emp = emp_map[emp_code]
            if is_nu_dynamic_shift_code(emp.default_shift_code):
                nu_emp_ids.append(emp.id)
            else:
                # Check if they have NU in schedule
                has_nu_sched = any(
                    is_nu_dynamic_shift_code(shifts_by_id[sid].code if sid in shifts_by_id else "") 
                    for (eid, _), sid in schedule_map.items() if eid == emp.id
                )
                if has_nu_sched:
                    nu_emp_ids.append(emp.id)
        
        # 2. Get NU results if any
        nu_results_all = {}
        if nu_emp_ids:
            # Prepare logs for build_nu_shift_day_results
            nu_logs = []
            nu_code_map = {}
            for emp_code, scans in raw_scans.items():
                emp = emp_map[emp_code]
                if emp.id not in nu_emp_ids: continue
                
                for s_dt in scans:
                    nu_logs.append(type('Log', (), {'employee_id': emp.id, 'event_time': s_dt}))
                
                # Determine codes for all dates in scans
                emp_dates = {s_dt.date() for s_dt in scans}
                # Also include previous and next day to be safe
                all_dates = set()
                for d in emp_dates:
                    all_dates.add(d)
                    all_dates.add(d - timedelta(days=1))
                    all_dates.add(d + timedelta(days=1))
                
                for d in all_dates:
                    sid = schedule_map.get((emp.id, d))
                    s_code = emp.default_shift_code
                    if sid and sid in shifts_by_id:
                        s_code = shifts_by_id[sid].code
                    
                    if is_nu_dynamic_shift_code(s_code):
                        nu_code_map[(emp.id, d)] = s_code
            
            nu_results_all = build_nu_shift_day_results(nu_code_map, nu_emp_ids, nu_logs)

        # 3. Final grouping and upsert
        for emp_code, scans in raw_scans.items():
            emp = emp_map[emp_code]
            
            # Map scans to work_dates
            # For NU: use nu_results_all to get the "correct" check times
            # For others: use 6am cutoff
            
            daily_data = {} # work_date -> (first_in, last_out)
            
            if emp.id in nu_emp_ids:
                # Use results from build_nu_shift_day_results
                emp_results = {k[1]: v for k, v in nu_results_all.items() if k[0] == emp.id}
                for w_date, res in emp_results.items():
                    if res.check_in or res.check_out:
                        daily_data[w_date] = (res.check_in, res.check_out)
            else:
                # Standard 6am cutoff
                daily_scans = defaultdict(list)
                for scan_dt in scans:
                    work_date = scan_dt.date()
                    if scan_dt.time() < NIGHT_SHIFT_CUTOFF:
                        # Simple rule: if before 6am and not NU, check if yesterday was night shift
                        prev_date = work_date - timedelta(days=1)
                        sid = schedule_map.get((emp.id, prev_date))
                        shift = shifts_by_id.get(sid) if sid else shifts_by_code.get(emp.default_shift_code)
                        if shift and shift.is_night_shift:
                            work_date = prev_date
                    daily_scans[work_date].append(scan_dt)
                
                for w_date, d_scans in daily_scans.items():
                    d_scans.sort()
                    daily_data[w_date] = (d_scans[0], d_scans[-1] if len(d_scans) > 1 else None)

            for work_date, (first_in, last_out) in daily_data.items():
                total_hours = 0.0
                if first_in and last_out and last_out > first_in:
                    total_hours = round((last_out - first_in).total_seconds() / 3600.0, 2)

                # Upsert
                existing = await db.execute(
                    select(AttendanceDaily).where(
                        and_(AttendanceDaily.employee_id == emp.id, AttendanceDaily.work_date == work_date)
                    )
                )
                att = existing.scalar_one_or_none()
                if att:
                    att.first_check_in = first_in
                    att.last_check_out = last_out
                    att.total_hours = total_hours
                    att.import_batch = batch_id
                    updated += 1
                else:
                    att = AttendanceDaily(
                        employee_id=emp.id,
                        work_date=work_date,
                        first_check_in=first_in,
                        last_check_out=last_out,
                        total_hours=total_hours,
                        import_batch=batch_id,
                    )
                    db.add(att)
                    processed += 1

        await db.commit()

        # Ghi nhật ký
        await log_audit(
            db, "attendance", batch_id, "IMPORT", current_user.username,
            notes=f"Import {file.filename} - {month_key}. {processed} moi, {updated} cap nhat. {len(raw_scans)} NV."
        )
        await db.commit()

        return {
            "message": f"Import thanh cong! {processed} moi, {updated} cap nhat",
            "batch_id": batch_id,
            "total_raw_rows": total_rows,
            "employees_processed": len(raw_scans),
            "days_created": processed,
            "days_updated": updated,
            "skipped_employees": list(skipped_employees)[:10] if skipped_employees else [],
            "filename": file.filename,
        }
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Internal Server Error during import: {str(e)}")


@router.get("/backup")
async def backup_database(
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN)),
):
    """Backup toan bo du lieu cac bang chinh thanh JSON"""
    tables = [
        ("employees", "SELECT * FROM employees"),
        ("shift_templates", "SELECT * FROM shift_templates"),
        ("work_schedules", "SELECT * FROM work_schedules"),
        ("attendance_daily", "SELECT * FROM attendance_daily"),
        ("attendance_logs", "SELECT * FROM attendance_logs"),
        ("company_holidays", "SELECT * FROM company_holidays"),
        ("app_users", "SELECT id, username, role, full_name, is_active FROM app_users"),
    ]

    backup_data = {
        "version": "1.0",
        "created_at": datetime.utcnow().isoformat(),
        "created_by": current_user.username,
        "tables": {},
    }

    for table_name, query in tables:
        try:
            result = await db.execute(text(query))
            rows = result.fetchall()
            columns = result.keys()
            data = []
            for row in rows:
                row_dict = {}
                for i, col in enumerate(columns):
                    val = row[i]
                    if isinstance(val, (datetime, date)):
                        val = val.isoformat()
                    elif hasattr(val, '__str__') and not isinstance(val, (str, int, float, bool, type(None))):
                        val = str(val)
                    row_dict[col] = val
                data.append(row_dict)
            backup_data["tables"][table_name] = data
        except Exception as e:
            backup_data["tables"][table_name] = {"error": str(e)}

    json_bytes = json.dumps(backup_data, ensure_ascii=False, indent=2).encode("utf-8")
    filename = f"hieploi_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"

    return StreamingResponse(
        BytesIO(json_bytes),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/restore")
async def restore_database(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN)),
):
    """Restore du lieu tu file backup JSON. Chi restore employees, shifts, schedules, holidays."""
    content = await file.read()
    try:
        backup = json.loads(content.decode("utf-8"))
    except Exception:
        raise HTTPException(400, "File khong hop le. Phai la JSON backup.")

    if "tables" not in backup:
        raise HTTPException(400, "File backup khong co du lieu tables")

    restored = {}

    # Restore order matters (foreign keys)
    # 1. shift_templates
    if "shift_templates" in backup["tables"]:
        data = backup["tables"]["shift_templates"]
        if isinstance(data, list):
            count = 0
            for row in data:
                existing = await db.execute(
                    select(ShiftTemplate).where(ShiftTemplate.code == row.get("code"))
                )
                if not existing.scalar_one_or_none():
                    shift = ShiftTemplate(
                        code=row["code"],
                        name=row.get("name"),
                        start_time=row.get("start_time"),
                        end_time=row.get("end_time"),
                        standard_hours=row.get("standard_hours"),
                        break_minutes=row.get("break_minutes"),
                        default_overtime_hours=row.get("default_overtime_hours"),
                        meal_allowance=row.get("meal_allowance", 0),
                        meal_count=row.get("meal_count", 0),
                        is_night_shift=row.get("is_night_shift", False),
                        is_leave_code=row.get("is_leave_code", False),
                        is_paid_leave=row.get("is_paid_leave", False),
                    )
                    db.add(shift)
                    count += 1
            restored["shift_templates"] = count

    # 2. company_holidays
    from app.models.holiday import CompanyHoliday
    if "company_holidays" in backup["tables"]:
        data = backup["tables"]["company_holidays"]
        if isinstance(data, list):
            count = 0
            for row in data:
                h_date = date.fromisoformat(row["holiday_date"]) if row.get("holiday_date") else None
                if h_date:
                    existing = await db.execute(
                        select(CompanyHoliday).where(CompanyHoliday.holiday_date == h_date)
                    )
                    if not existing.scalar_one_or_none():
                        h = CompanyHoliday(
                            holiday_date=h_date,
                            name=row.get("name", ""),
                            holiday_type=row.get("holiday_type", "company"),
                            is_active=row.get("is_active", True),
                            notes=row.get("notes"),
                        )
                        db.add(h)
                        count += 1
            restored["company_holidays"] = count

    await db.commit()

    return {
        "message": "Restore thanh cong!",
        "restored": restored,
        "backup_version": backup.get("version"),
        "backup_date": backup.get("created_at"),
    }
@router.get("/export-meal-allowance")
async def export_meal_allowance(
    start_date: date = Query(...),
    end_date: date = Query(...),
    night_allowance: float = Query(100000),
    department: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Xuat file Excel tien com + boi duong ca dem theo format mau"""
    import openpyxl
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

    # 1. Lay du lieu (giong ben meal_allowance.py nhung goi truc tiep)
    from app.routers.meal_allowance import get_meal_allowance
    data = await get_meal_allowance(start_date, end_date, department, night_allowance, db, current_user)
    
    # 2. Tao Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tien Com - Boi Duong"

    # Style
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_font = Font(name="Times New Roman", bold=True, size=11)
    title_font = Font(name="Times New Roman", bold=True, size=14)
    normal_font = Font(name="Times New Roman", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    dash_number_format = "#,##0;-#,##0;-;@"

    # Row 1-3: Headers
    ws.merge_cells("A1:Q1")
    ws["A1"] = "CÔNG TY TNHH HIỆP LỢI"
    ws["A1"].font = Font(name="Times New Roman", bold=True, size=12)

    ws.merge_cells("A2:Q2")
    ws["A2"] = "MST: 3701609885"
    ws["A2"].font = Font(name="Times New Roman", size=11)

    ws.merge_cells("A3:Q3")
    date_str = f"TỪ NGÀY {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    ws["A3"] = f"TIỀN CƠM VÀ TIỀN BỒI DƯỠNG TĂNG CA ĐÊM {date_str}"
    ws["A3"].font = title_font
    ws["A3"].alignment = center

    ws.merge_cells("A4:Q4")
    zh_date_str = f"由 {start_date.year}年 {start_date.month}月 {start_date.day} 日 至 {end_date.year}年 {end_date.month}月 {end_date.day}日"
    ws["A4"] = f"廠房員工伙食費與晚班補貼費{zh_date_str}"
    ws["A4"].font = Font(name="Times New Roman", bold=True, size=12)
    ws["A4"].alignment = center

    # Row 6: Vietnamese Header
    headers_vn = [
        "STT", "MSNV", "HỌ VÀ TÊN", "Số Bữa", "Tiền Cơm", "Cộng Tiền cơm",
        "Số Đêm", "Bồi Dưỡng Ca Đêm", "Cộng tiền bồi dưỡng Đêm", "Số Bữa",
        "Tiền BD Đi Phụ Xe Tải 30.000đ /Bửa- Đi nối sợi 70.000đ /ngày", "Cộng tiền bồi dưỡng", "TIỀN ĐIỆN", "TIỀN THỰC LÃNH",
        "SỐ TIỀN THIẾU", "SỐ TIỀN DƯ", "KÍ NHẬN"
    ]
    # Row 7: Chinese Header
    headers_zh = [
        "次序", "MSNV", "姓名", "餐數", "每餐,餐費", "小計", 
        "晚數", "大晚班補貼", "小計", "餐數", 
        "補貼跟車 / 穿扣", "小計", "NaN", "合計",
        "NaN", "NaN", "簽名"
    ]
    
    for col_idx, header in enumerate(headers_vn, 1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = gray_fill
        
    for col_idx, header in enumerate(headers_zh, 1):
        cell = ws.cell(row=7, column=col_idx)
        if header != "NaN" and header != "MSNV":
            cell.value = header
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = gray_fill

    # Row 8+: Data
    current_row = 8
    # Sort by employee_code numerically
    sorted_rows = sorted(data.rows, key=lambda x: int(x.employee_code) if str(x.employee_code).isdigit() else 999999)
    
    for idx, item in enumerate(sorted_rows, 1):
        # STT, MSNV, Name
        ws.cell(current_row, 1, idx).font = normal_font
        ws.cell(current_row, 2, item.employee_code).font = normal_font
        ws.cell(current_row, 3, item.full_name).font = normal_font
        
        # Meal
        ws.cell(current_row, 4, item.meal_count).font = normal_font
        ws.cell(current_row, 5, item.meal_rate).font = normal_font
        # Formula: Cell F = D * E
        ws.cell(current_row, 6, f"=D{current_row}*E{current_row}").font = normal_font
        
        # Night
        ws.cell(current_row, 7, item.night_shifts).font = normal_font
        if item.night_shifts > 0:
            ws.cell(current_row, 8, night_allowance).font = normal_font
        else:
            ws.cell(current_row, 8, "-").font = normal_font
            
        # Formula: Cell I = G * H
        # Since H might be "-", we use a formula that treats non-numeric as 0
        ws.cell(current_row, 9, f"=G{current_row}*IF(ISNUMBER(H{current_row}),H{current_row},0)").font = normal_font
        
        # Phụ xe (Empty with formula)
        ws.cell(current_row, 10, None).font = normal_font
        ws.cell(current_row, 11, None).font = normal_font
        # Formula: Cell L = J * K
        ws.cell(current_row, 12, f"=J{current_row}*K{current_row}").font = normal_font
        
        # Tiền điện (Empty)
        ws.cell(current_row, 13, None).font = normal_font
        
        # Thực lãnh
        # Formula: Cell N = F + I + L - M + O - P
        ws.cell(current_row, 14, f"=F{current_row}+I{current_row}+L{current_row}-M{current_row}+O{current_row}-P{current_row}").font = Font(name="Times New Roman", bold=True)
        
        # Thiếu, Dư, Kí nhận
        ws.cell(current_row, 15, None).font = normal_font
        ws.cell(current_row, 16, None).font = normal_font
        ws.cell(current_row, 17, None).font = normal_font

        # Borders for all cells in row
        numeric_cols = {4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}
        right_align_cols = {5, 6, 8, 9, 11, 12, 13, 14, 15, 16}
        for c in range(1, 18):
            cell = ws.cell(current_row, c)
            cell.border = border
            if c in numeric_cols:
                cell.number_format = dash_number_format
            if c in right_align_cols:
                cell.alignment = right
            else:
                cell.alignment = center

        current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    for c in "DEFGHIJKLM":
        ws.column_dimensions[c].width = 12
    ws.column_dimensions["N"].width = 15
    ws.column_dimensions["Q"].width = 15

    # Export
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"TienAn_BoiDuong_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
