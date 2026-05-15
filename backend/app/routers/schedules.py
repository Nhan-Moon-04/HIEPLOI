from typing import List, Optional
from datetime import date
import calendar
import re
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, delete, cast, Integer
from app.database import get_db
from app.models.schedule import WorkSchedule
from app.models.employee import Employee
from app.models.shift import ShiftTemplate
from app.models.x_overtime import XOvertimeConfig
from app.models.user import AppUser, UserRole
from app.middleware.auth import get_current_user, require_roles
from pydantic import BaseModel
from decimal import Decimal

router = APIRouter(prefix="/schedules", tags=["Schedules"])

DOW_VN = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]


class ScheduleCell(BaseModel):
    employee_id: int
    employee_code: str
    full_name: str
    department: Optional[str] = None
    default_shift_code: Optional[str] = None
    days: dict


class ScheduleMonthResponse(BaseModel):
    month_key: str
    year: int
    month: int
    days_in_month: int
    weekdays: dict
    rows: List[ScheduleCell]


class ScheduleUpdateRequest(BaseModel):
    employee_id: int
    day: int
    shift_code: Optional[str] = None


@router.get("")
async def get_schedule(
    month_key: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    try:
        year, month = map(int, month_key.split("-"))
    except ValueError:
        raise HTTPException(400, "month_key phai la YYYY-MM")

    dim = calendar.monthrange(year, month)[1]
    weekdays = {}
    for d in range(1, dim + 1):
        weekdays[str(d)] = DOW_VN[date(year, month, d).weekday()]

    first_day = date(year, month, 1)
    last_day = date(year, month, dim)

    emp_q = select(Employee).where(
        Employee.is_active == True,
    ).order_by(Employee.employee_code)
    emps = (await db.execute(emp_q)).scalars().all()

    sched_q = select(WorkSchedule).where(WorkSchedule.month_key == month_key)
    scheds = (await db.execute(sched_q)).scalars().all()

    shift_q = select(ShiftTemplate)
    shifts = {s.id: s.code for s in (await db.execute(shift_q)).scalars().all()}

    omap = {}
    for ws in scheds:
        omap[(ws.employee_id, ws.work_date.day)] = shifts.get(ws.shift_id, "?")

    rows = []
    for emp in emps:
        days = {}
        for d in range(1, dim + 1):
            key = (emp.id, d)
            days[str(d)] = omap.get(key, None)
        rows.append(ScheduleCell(
            employee_id=emp.id,
            employee_code=emp.employee_code,
            full_name=emp.full_name,
            department=emp.department,
            default_shift_code=emp.default_shift_code,
            days=days,
        ))

    return ScheduleMonthResponse(
        month_key=month_key, year=year, month=month,
        days_in_month=dim, weekdays=weekdays, rows=rows,
    )


@router.put("/cell")
async def update_schedule_cell(
    request: ScheduleUpdateRequest,
    month_key: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    try:
        year, month = map(int, month_key.split("-"))
    except ValueError:
        raise HTTPException(400, "Invalid month_key")

    work_date = date(year, month, request.day)

    if request.shift_code is None:
        await db.execute(delete(WorkSchedule).where(and_(
            WorkSchedule.employee_id == request.employee_id,
            WorkSchedule.work_date == work_date,
        )))
        await db.commit()
        return {"message": "Reverted to default"}

    shift_r = await db.execute(select(ShiftTemplate).where(ShiftTemplate.code == request.shift_code))
    shift = shift_r.scalar_one_or_none()
    if not shift:
        raise HTTPException(400, f"Shift '{request.shift_code}' not found")

    existing = await db.execute(select(WorkSchedule).where(and_(
        WorkSchedule.employee_id == request.employee_id,
        WorkSchedule.work_date == work_date,
    )))
    ws = existing.scalar_one_or_none()
    if ws:
        ws.shift_id = shift.id
    else:
        ws = WorkSchedule(
            employee_id=request.employee_id,
            work_date=work_date,
            month_key=month_key,
            shift_id=shift.id,
        )
        db.add(ws)
    await db.commit()
    return {"message": "Updated"}


@router.post("/import")
async def import_schedule(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

    target_sheet = None
    for name in wb.sheetnames:
        if "lich" in name.lower() or "schedule" in name.lower():
            target_sheet = name
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]

    title = str(ws.cell(1, 1).value or "")
    month_key = None
    m = re.search(r"(\d{1,2})\s*/\s*(\d{4})", title)
    if m:
        month_key = f"{int(m.group(2))}-{int(m.group(1)):02d}"
    if not month_key:
        raise HTTPException(400, "Khong tim thay thang/nam trong tieu de. Format: 'LICH LAM VIEC THANG MM/YYYY'")

    year, month = map(int, month_key.split("-"))
    dim = calendar.monthrange(year, month)[1]

    day_cols = {}
    for c in range(4, ws.max_column + 1):
        dv = ws.cell(2, c).value
        if dv is not None:
            try:
                d = int(dv)
                if 1 <= d <= dim:
                    day_cols[c] = d
            except (ValueError, TypeError):
                pass

    shift_r = await db.execute(select(ShiftTemplate))
    shift_map = {}
    for s in shift_r.scalars().all():
        shift_map[s.code.upper()] = s.id
        shift_map[s.code.lower()] = s.id

    emp_r = await db.execute(select(Employee))
    emp_map = {e.employee_code: e.id for e in emp_r.scalars().all()}

    created = 0
    updated = 0
    skipped = 0
    unknown_shifts = set()

    for r in range(4, ws.max_row + 1):
        emp_code = ws.cell(r, 2).value
        if emp_code is None:
            continue
        emp_code = str(int(emp_code) if isinstance(emp_code, float) else emp_code).strip()
        if emp_code not in emp_map:
            continue

        emp_id = emp_map[emp_code]

        for col, day in day_cols.items():
            cell_val = ws.cell(r, col).value
            if cell_val is None or str(cell_val).strip() == "":
                continue

            sc = str(cell_val).strip().upper()
            if sc in ("O", "OFF"):
                sc = "OFF"

            if sc not in shift_map:
                unknown_shifts.add(sc)
                skipped += 1
                continue

            shift_id = shift_map[sc]
            wd = date(year, month, day)

            ex = await db.execute(select(WorkSchedule).where(and_(
                WorkSchedule.employee_id == emp_id,
                WorkSchedule.work_date == wd,
            )))
            rec = ex.scalar_one_or_none()
            if rec:
                rec.shift_id = shift_id
                updated += 1
            else:
                db.add(WorkSchedule(
                    employee_id=emp_id, work_date=wd,
                    month_key=month_key, shift_id=shift_id,
                ))
                created += 1

    await db.commit()

    return {
        "message": f"Import xong! Tao {created}, cap nhat {updated}, bo qua {skipped}",
        "month_key": month_key,
        "sheet": target_sheet,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "unknown_shifts": list(unknown_shifts),
    }


# ─── X Overtime Config ────────────────────────────────────────────────────────

class XOvertimeConfigRequest(BaseModel):
    employee_id: int
    work_date: date
    ot_end_time: Optional[str] = None
    ot_hours: Optional[Decimal] = None
    meal_count: Optional[int] = None


@router.get("/x-overtime")
async def get_x_overtime_config(
    month_key: str,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    year, month = map(int, month_key.split("-"))
    start_date = date(year, month, 1)
    end_date = date(year, month, calendar.monthrange(year, month)[1])
    
    result = await db.execute(
        select(XOvertimeConfig).where(
            XOvertimeConfig.work_date >= start_date,
            XOvertimeConfig.work_date <= end_date,
        )
    )
    cfgs = result.scalars().all()
    return [
        {
            "employee_id": c.employee_id,
            "work_date": c.work_date,
            "ot_end_time": c.ot_end_time.strftime("%H:%M") if c.ot_end_time else None,
            "ot_hours": c.ot_hours,
            "meal_count": c.meal_count,
        } for c in cfgs
    ]


@router.put("/x-overtime")
async def upsert_x_overtime_config(
    request: XOvertimeConfigRequest,
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    from datetime import time as dt_time

    result = await db.execute(
        select(XOvertimeConfig).where(
            XOvertimeConfig.employee_id == request.employee_id,
            XOvertimeConfig.work_date == request.work_date,
        )
    )
    cfg = result.scalar_one_or_none()

    parsed_end_time = None
    if request.ot_end_time:
        parts = request.ot_end_time.split(":")
        parsed_end_time = dt_time(int(parts[0]), int(parts[1]))

    if cfg:
        cfg.ot_end_time = parsed_end_time
        cfg.ot_hours = request.ot_hours
        cfg.meal_count = request.meal_count or 0
    else:
        cfg = XOvertimeConfig(
            employee_id=request.employee_id,
            work_date=request.work_date,
            ot_end_time=parsed_end_time,
            ot_hours=request.ot_hours,
            meal_count=request.meal_count or 0,
        )
        db.add(cfg)

    await db.commit()
    return {"message": "Da luu config tang ca X", "employee_id": request.employee_id, "work_date": request.work_date}


@router.delete("/x-overtime")
async def delete_x_overtime_config(
    employee_id: int = Query(...),
    work_date: date = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: AppUser = Depends(require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)),
):
    """Xóa config tăng ca X của nhân viên theo ngày"""
    result = await db.execute(
        select(XOvertimeConfig).where(
            XOvertimeConfig.employee_id == employee_id,
            XOvertimeConfig.work_date == work_date,
        )
    )
    cfg = result.scalar_one_or_none()
    if cfg:
        await db.delete(cfg)
        await db.commit()
    return {"message": "Da xoa"}
